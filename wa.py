import os
import sys
import time
import argparse
import random
import shutil
import threading
import webbrowser
from datetime import datetime

try:
    import openpyxl
    import pandas as pd
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from webdriver_manager.chrome import ChromeDriverManager
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
except ImportError:
    print("❗️ [!] Please install requirements: pip install selenium webdriver-manager openpyxl pandas")
    sys.exit(1)

SESSION_DIR = os.path.join(os.getcwd(), "wa_session")
DEBUG_DIR = os.path.join(os.getcwd(), "debug_logs")
LOG_FILE = os.path.join(os.getcwd(), "wa_log.txt")
os.makedirs(DEBUG_DIR, exist_ok=True)

def log(msg):
    print(msg)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(f"{datetime.now()} {msg}\n")

def clear_session():
    if os.path.exists(SESSION_DIR):
        shutil.rmtree(SESSION_DIR)
    log("🗑️ [*] Chrome session cleared.")

def wait_random(min_sec=1, max_sec=20):
    t = random.uniform(min_sec, max_sec)
    log(f"⏳ [*] Waiting {t:.2f}s before next action... ⏳")
    time.sleep(t)

def split_multi_input(input_string):
    if not input_string:
        return []
    return [x.strip() for x in input_string.split(",") if x.strip()]

def col_letter_to_index(letter):
    if letter.isalpha():
        letter = letter.upper()
        index = 0
        for c in letter:
            index = index * 26 + (ord(c) - ord('A') + 1)
        return index - 1
    return None

def get_numbers_from_excel(file_path, col):
    # Try both by header (name) and by column letter
    if file_path.lower().endswith('.csv'):
        df = pd.read_csv(file_path, dtype=str)
        if col.isalpha():
            col_idx = col_letter_to_index(col)
            if col_idx >= len(df.columns):
                raise Exception(f"❗️ ERROR: CSV file has only {len(df.columns)} columns, can't get column {col.upper()}.")
            numbers = df.iloc[:, col_idx].astype(str).tolist()
            log(f"📥 [+] Imported {len(numbers)} contacts from CSV column {col.upper()}.")
        else:
            if col not in df.columns:
                raise Exception(f"❗️ ERROR: CSV column '{col}' not found! Columns are: {list(df.columns)}")
            numbers = df[col].astype(str).tolist()
            log(f"📥 [+] Imported {len(numbers)} contacts from CSV column '{col}'.")
        return numbers
    else:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        header = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]
        numbers = []
        col_idx = None
        if col.isalpha():
            col_idx = col_letter_to_index(col)
            if col_idx >= len(header):
                raise Exception(f"❗️ ERROR: Excel has only {len(header)} columns, can't get column {col.upper()}.")
        else:
            for idx, name in enumerate(header):
                if name == col.strip().lower():
                    col_idx = idx
                    break
            if col_idx is None:
                raise Exception(f"❗️ Column '{col}' not found in Excel file! Columns are: {header}")
        for row in ws.iter_rows(min_row=2):
            if col_idx >= len(row):
                continue
            val = row[col_idx].value
            if val:
                numbers.append(str(val).strip())
        log(f"📥 [+] Imported {len(numbers)} contacts from Excel column {col}.")
        return numbers

def get_chrome_driver(profile_path=None, headless=False):
    from selenium.webdriver.chrome.options import Options
    options = Options()
    if profile_path:
        options.add_argument(f'--user-data-dir={profile_path}')
        options.add_argument('--profile-directory=Default')
    if headless:
        options.add_argument('--headless')
        options.add_argument('--window-size=1920,1080')
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--lang=en')
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    try:
        driver.maximize_window()
    except Exception:
        pass
    log("🚗 [*] Chrome launched and maximized.")
    return driver

def wait_for_element(driver, xpath, timeout=15):
    try:
        return WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.XPATH, xpath)))
    except Exception:
        return None

def safe_click_attach(driver):
    attach_btn = None
    try:
        attach_btn = WebDriverWait(driver, 20).until(
            EC.visibility_of_element_located((By.XPATH, '//span[@data-icon="plus-rounded"]'))
        )
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", attach_btn)
        time.sleep(0.7)
        try:
            attach_btn.click()
        except Exception:
            driver.execute_script("arguments[0].click();", attach_btn)
    except Exception:
        attach_btn = wait_for_element(driver, '//div[@title="Attach"]', 5)
        if attach_btn:
            try:
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", attach_btn)
                time.sleep(0.7)
                attach_btn.click()
            except Exception:
                driver.execute_script("arguments[0].click();", attach_btn)
        else:
            filename = f"debug_no_attach_{int(time.time())}.png"
            driver.save_screenshot(os.path.join(DEBUG_DIR, filename))
            log(f"❗️ [!] ERROR: Attach button not found. Screenshot saved as {filename}")
            log(driver.page_source[:1000])
            raise Exception("Attach button not found.")
    time.sleep(2)

def login_whatsapp():
    log("🔄 [*] Starting WhatsApp Web login... 🚀")
    clear_session()
    driver = get_chrome_driver(profile_path=SESSION_DIR)
    driver.get("https://web.whatsapp.com/")
    log("🔲 [*] Waiting for QR code in the opened browser (press ENTER any time to force success)...")

    enter_pressed = {'value': False}
    def monitor_enter():
        input()
        enter_pressed['value'] = True
    t = threading.Thread(target=monitor_enter, daemon=True)
    t.start()

    qr_xpath = '//canvas[@aria-label="Scan this QR code to link a device!"]'
    qr_found = False
    for attempt in range(1, 4):
        if enter_pressed['value']:
            log("🟢 [*] Manual ENTER detected. Forcing login success!")
            driver.quit()
            return
        try:
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.XPATH, qr_xpath))
            )
            qr_found = True
            log(f"✅ [*] QR Code loaded (try {attempt}/3). Please scan now! 📱 (or press ENTER to bypass)")
            break
        except Exception:
            log(f"⏳ [*] QR code not detected after {15*attempt}s. Retrying..." if attempt < 3 else "")
    if not qr_found:
        log("⚠️ [!] QR code did not appear in 45s. Closing browser.\n"
            "WARNING: All log messages before absl::InitializeLog() is called are written to STDERR\n"
            "❌ [!] QR code did not appear in 45s. Is WhatsApp down or blocked? 😢")
        driver.quit()
        return

    login_complete = False
    try:
        log("👀 [*] Waiting for you to scan QR... 🕵️‍♂️ (or press ENTER to bypass)")
        for _ in range(120):
            if enter_pressed['value']:
                log("🟢 [*] Manual ENTER detected. Forcing login success!")
                driver.quit()
                return
            if not driver.find_elements(By.XPATH, qr_xpath):
                log("🎉 [*] QR scanned, loading chats...")
                break
            time.sleep(1)
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, '//div[@role="grid"]'))
        )
        log("⏳ [*] Chats loaded. Waiting extra 10 seconds for session stability... 💾")
        for _ in range(10):
            if enter_pressed['value']:
                log("🟢 [*] Manual ENTER detected. Forcing login success!")
                driver.quit()
                return
            time.sleep(1)
        login_complete = True
    except Exception:
        pass

    if login_complete:
        log("✅ [+] Login successful. Closing browser... 🚀")
    else:
        if not enter_pressed['value']:
            log("❌ [!] Login not detected in time. Closing browser.")
    driver.quit()

def send_messages(numbers, message):
    driver = get_chrome_driver(profile_path=SESSION_DIR)
    success, failure = 0, 0
    for idx, number in enumerate(numbers, start=1):
        sent = False
        log(f"💬 [*] Sending message to {number} ({idx}/{len(numbers)})...")
        url = f"https://web.whatsapp.com/send?phone={number}&text={message}"
        driver.get(url)
        if not wait_for_element(driver, '//div[@role="grid"]', timeout=15):
            log(f"❗️ [!] ERROR: Chat not loaded for {number}. Not retrying (maybe invalid number).")
            failure += 1
        else:
            for attempt in range(1, 4):
                try:
                    input_box = wait_for_element(driver, '//div[@contenteditable="true"][@data-tab="10"]', 10)
                    if input_box:
                        input_box.send_keys('\n')
                        log(f"✅ [+] Message sent to {number}! ✉️ (try {attempt}/3)")
                        sent = True
                        success += 1
                        break
                    else:
                        log(f"❗️ [!] ERROR: Message box not found for {number}. (try {attempt}/3)")
                except Exception as e:
                    log(f"❗️ [!] ERROR: {e} (try {attempt}/3)")
                if attempt < 3:
                    log(f"🔁 [*] Retrying message send in 5 seconds...")
                    time.sleep(5)
            if not sent:
                log(f"❌ [!] Failed to send message to {number} after 3 tries.")
                failure += 1
        wait_random()
    driver.quit()
    return success, failure

def send_attachments_any(files, numbers, message):
    driver = get_chrome_driver(profile_path=SESSION_DIR)
    success, failure = 0, 0
    for idx, number in enumerate(numbers, start=1):
        sent = False
        file_path = files[min(idx-1, len(files)-1)]
        if not os.path.isfile(file_path):
            log(f"❗️ [!] Attachment file not found: {file_path}")
            failure += 1
            continue
        log(f"📎 [*] Sending attachment to {number} ({idx}/{len(numbers)})...")
        url = f"https://web.whatsapp.com/send?phone={number}"
        driver.get(url)
        if not wait_for_element(driver, '//div[@role="grid"]', timeout=15):
            log(f"❗️ [!] ERROR: Chat not loaded for {number}. Not retrying.")
            failure += 1
            continue
        for attempt in range(1, 4):
            try:
                # Message first, then file, then wait, then send
                if message:
                    input_box = wait_for_element(driver, '//div[@contenteditable="true"][@data-tab="10"]', 10)
                    if input_box:
                        input_box.clear()
                        input_box.send_keys(message)
                        log(f"📝 [*] Entered message for {number}, waiting 2s before attaching...")
                        time.sleep(2)
                safe_click_attach(driver)
                file_input = None
                try:
                    file_input = driver.find_element(By.XPATH, '//input[contains(@accept,"image") or contains(@accept,"video")]')
                except Exception:
                    try:
                        file_input = driver.find_element(By.XPATH, '//input[@accept="*"]')
                    except Exception:
                        file_input = driver.find_element(By.XPATH, '//input[@type="file"]')
                if not file_input:
                    raise Exception("Attachment file input not found.")
                file_input.send_keys(os.path.abspath(file_path))
                log(f"📎 [*] Attached file for {number}, waiting 2s before sending...")
                time.sleep(2)
                send_btn = wait_for_element(driver, '//div[@role="button" and @aria-label="Send"]', 10)
                if not send_btn:
                    send_btn = wait_for_element(driver, '//span[@data-icon="send"]', 5)
                if not send_btn:
                    raise Exception("Send button not found.")
                send_btn.click()
                log(f"✅ [+] Attachment sent to {number}! 📎 (try {attempt}/3)")
                sent = True
                success += 1
                break
            except Exception as e:
                log(f"❗️ [!] ERROR: {e} (try {attempt}/3)")
            if attempt < 3:
                log(f"🔁 [*] Retrying attachment send in 5 seconds...")
                time.sleep(5)
        if not sent:
            log(f"❌ [!] Failed to send attachment to {number} after 3 tries.")
            failure += 1
        wait_random()
    driver.quit()
    return success, failure

def send_from_excel(excel_path, col, message, attachment_path):
    numbers = get_numbers_from_excel(excel_path, col)
    log(f"📋 [*] Action: MSG. Total contacts: {len(numbers)}")
    files = split_multi_input(attachment_path) if attachment_path else []
    if files:
        s, f = send_attachments_any(files, numbers, message)
    else:
        s, f = send_messages(numbers, message)
    generate_html_report(s, f)

def generate_html_report(success, failure):
    total = success + failure
    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>WabulkXpress Messaging Analytics</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<style>
    * {{
        box-sizing: border-box;
        margin: 0;
        padding: 0;
    }}
    body {{
        font-family: 'Arial', sans-serif;
        color: #e0e0e0;
        position: relative;
        min-height: 100vh;
        overflow: hidden;
        padding: 20px;
    }}
    body::before {{
        content: "";
        background: #121212;
        background-size: cover;
        filter: blur(8px);
        position: absolute;
        top: 0;
        left: 0;
        right: 0;
        bottom: 0;
        z-index: -2;
    }}
    body::after {{
        content: "";
        position: absolute;
        top: 0;
        left: 0;
        right: 0;
        bottom: 0;
        background-color: rgba(0, 0, 0, 0.6);
        z-index: -1;
    }}
    .container {{
        display: flex;
        align-items: flex-start;
        justify-content: space-between;
        padding: 30px;
        gap: 20px;
        border-radius: 10px;
        background-color: rgba(30, 30, 30, 0.9);
        max-width: 1200px;
        margin: auto;
        box-shadow: 0 0 20px rgba(0,0,0,0.5);
        transition: transform 0.3s ease;
    }}
    .container:hover {{
        transform: scale(1.02);
    }}
    .info {{
        flex: 1;
        padding: 20px;
        background: rgba(0, 0, 0, 0.3);
        border-radius: 10px;
        margin-right: 20px;
    }}
    .info h2 {{
        margin-bottom: 15px;
    }}
    .info p {{
        margin-bottom: 10px;
        line-height: 1.5;
    }}
    .chart-container {{
        flex: 1;
        position: relative;
        max-width: 400px;
        background: rgba(0, 0, 0, 0.3);
        padding: 15px;
        border-radius: 10px;
        box-shadow: 0 0 15px rgba(0,0,0,0.4);
        transition: box-shadow 0.3s ease;
    }}
    .chart-container:hover {{
        box-shadow: 0 0 25px rgba(0,0,0,0.6);
    }}
    h1 {{
        margin-bottom: 20px;
        text-align: center;
        font-size: 2em;
    }}
    button {{
        padding: 10px 20px;
        border: none;
        border-radius: 5px;
        background-color: #0078D7;
        color: #fff;
        cursor: pointer;
        transition: background-color 0.3s ease, transform 0.3s ease;
        margin-top: 15px;
    }}
    button:hover {{
        background-color: #005fa3;
        transform: scale(1.05);
    }}
    @media (max-width: 768px) {{
        .container {{
            flex-direction: column;
        }}
        .chart-container {{
            margin-top: 20px;
            max-width: 100%;
        }}
    }}
</style>
</head>
<body onload="window.focus();">
<h1>WabulkXpress Messaging Analytics</h1>
<div class="container">
    <div class="info">
        <h2>Message Summary</h2>
        <p>Total Messages: <strong id="totalCount"></strong></p>
        <p>Success: <strong id="successCount"></strong></p>
        <p>Failure: <strong id="failureCount"></strong></p>
        <button onclick="window.close();">Close Report</button>
    </div>
    <div class="chart-container">
        <canvas id="pieChart"></canvas>
    </div>
</div>
<script>
    const total = {total};
    const success = {success};
    const failure = {failure};
    document.getElementById('totalCount').textContent = total;
    document.getElementById('successCount').textContent = success;
    document.getElementById('failureCount').textContent = failure;
    const ctx = document.getElementById('pieChart').getContext('2d');
    const data = {{
        labels: ['Success', 'Failure'],
        datasets: [{{
            data: [success, failure],
            backgroundColor: ['#4CAF50', '#F44336'],
            borderColor: ['#2E7D32', '#C62828'],
            borderWidth: 2,
        }}]
    }};
    const options = {{
        cutout: '70%',
        responsive: true,
        plugins: {{
            legend: {{
                position: 'bottom',
                labels: {{
                    color: '#e0e0e0'
                }}
            }}
        }}
    }};
    new Chart(ctx, {{
        type: 'doughnut',
        data: data,
        options: options
    }});
</script>
</body>
</html>"""
    report_path = os.path.join(os.getcwd(), "Report.html")
    try:
        with open(report_path, "w", encoding="utf-8") as f:
            f.write(html_content)
        webbrowser.open("file:///" + report_path)
        log(f"📊 [*] HTML Report generated: {report_path}")
    except Exception as e:
        log(f"Error generating/opening HTML report: {e}")

def main():
    parser = argparse.ArgumentParser(description='💬 WhatsApp Automation CLI 💬')
    subparsers = parser.add_subparsers(dest='command', help='Sub-commands')

    subparsers.add_parser('login', help='Login to WhatsApp Web')

    msg_parser = subparsers.add_parser('msg', help='Send message, attachment, or both')
    msg_parser.add_argument('arg1', nargs='?', type=str, help='Attachment file OR Phone number(s)')
    msg_parser.add_argument('arg2', nargs='?', type=str, help='Phone number(s) OR Message')
    msg_parser.add_argument('arg3', nargs='?', type=str, help='Message (optional)')
    msg_parser.add_argument('-exl', type=str, help='Excel/CSV file')
    msg_parser.add_argument('-col', type=str, help='Column in Excel/CSV for phone numbers (letter or header name)')
    msg_parser.add_argument('-fileloc', type=str, help='Attachment file path(s) for each Excel/CSV row, comma-separated')

    args = parser.parse_args()

    if args.command == 'login':
        login_whatsapp()
    elif args.command == 'msg':
        if args.exl and args.col:
            send_from_excel(args.exl, args.col, args.arg3 if args.arg3 else args.arg2, args.fileloc)
        else:
            if args.arg1 and args.arg2 and not args.arg3:
                if (args.arg1.strip().startswith("+") or args.arg1.strip()[0].isdigit()) and not os.path.isfile(args.arg1):
                    numbers = split_multi_input(args.arg1)
                    message = args.arg2
                    s, f = send_messages(numbers, message)
                    generate_html_report(s, f)
                else:
                    files = split_multi_input(args.arg1)
                    numbers = split_multi_input(args.arg2)
                    s, f = send_attachments_any(files, numbers, None)
                    generate_html_report(s, f)
            elif args.arg1 and args.arg2 and args.arg3:
                files = split_multi_input(args.arg1)
                numbers = split_multi_input(args.arg2)
                message = args.arg3
                s, f = send_attachments_any(files, numbers, message)
                generate_html_report(s, f)
            elif args.arg1 and not args.arg2:
                numbers = split_multi_input(args.arg1)
                print("❗️ [!] No message or attachment to send.")
            else:
                parser.print_help()
    else:
        parser.print_help()

if __name__ == '__main__':
    main()
