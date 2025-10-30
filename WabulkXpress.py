#!/usr/bin/env python
import os
import re
import sys
import ctypes
import google.generativeai as genai
import shutil
import time
import random
import math
import threading
import queue
import webbrowser
import requests
import tkinterdnd2
import pyautogui
from tkinter import messagebox
import openpyxl
from io import StringIO
import gc
from tkinter import filedialog, messagebox, colorchooser, Label
from PIL import Image, ImageTk, ImageDraw, ImageFont, ImageOps
from io import BytesIO
import win32clipboard
import customtkinter as ctk
import tkinter as tk
from datetime import datetime, timedelta
import csv
import struct
import urllib.parse
import logging
import sys
import json
logger = logging.getLogger(__name__)
import logging
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from dotenv import load_dotenv
load_dotenv()
logging.getLogger('libav').setLevel(logging.ERROR)

# ----------------------- THEME COLORS -----------------------
DARK_BG = "#000000"
LIGHT_BG = "#FFFFFF"
DARK_FG = "#111111"
LIGHT_FG = "#F5F5F5"
DARK_INNER = "#222222"
LIGHT_INNER = "#EAEAEA"
# ------------------------------------------------------------

# ----------------------- GLOBAL CONSTANTS & PATHS -----------------------
# Application version and GitHub URLs
CURRENT_VERSION = "25"
GITHUB_API_URL = "https://api.github.com/repos/ParthSancheti/WabulkXpress/releases/latest"
GITHUB_RELEASES_URL = "https://github.com/ParthSancheti/WabulkXpress/"
GITHUB_PROFILE_URL="https://github.com/ParthSancheti/"

# File and folder paths
BIN_FOLDER = os.path.join(os.getcwd(), "bin")
LICENSE_FILE = os.path.join(BIN_FOLDER, "license.dat")
TITLE_ICON_PATH = os.path.join(BIN_FOLDER, "loco.ico")
LOGO_PATH = os.path.join(BIN_FOLDER, "Logo.png")
MY_PROFILE_PIC = os.path.join(BIN_FOLDER, "profile_pic.jpg")
OUTPUT_IMG_FOLDER = os.path.join(os.getcwd(), "output_img")
SESSION_DIR = os.path.join(os.getcwd(), "selenium_session")

# Schedule Popup Icons
UP_ARROW_LIGHT_PATH = os.path.join(BIN_FOLDER, "up_arrow.png")
UP_ARROW_DARK_PATH = os.path.join(BIN_FOLDER, "up_arrow_dark.png")
DOWN_ARROW_LIGHT_PATH = os.path.join(BIN_FOLDER, "down_arrow.png")
UPLOAD_ICON_LIGHT_PATH = os.path.join(BIN_FOLDER, "upload_icon_light.png")
UPLOAD_ICON_DARK_PATH = os.path.join(BIN_FOLDER, "upload_icon_dark.png")
UPDATE_CHECK_FILE = os.path.join(BIN_FOLDER, "update_check.dat")
DOWN_ARROW_DARK_PATH = os.path.join(BIN_FOLDER, "down_arrow_dark.png")
ERROR_IMAGE_PATH = os.path.join(BIN_FOLDER, "alert_image.jpg")
PRODUCT_IMAGE_PATH = os.path.join(BIN_FOLDER, "product_key_image.png")
WHATSAPP_NUMBER = "+918007579299"

# Delay settings
DEFAULT_MIN_DELAY = 1
DEFAULT_MAX_DELAY = 10
VIDEO_PATH = os.path.join(BIN_FOLDER, "woi.mp4")
LOADING_GIF_PATH = os.path.join(BIN_FOLDER, "lod.gif")
AI_ICON_PATH = os.path.join(BIN_FOLDER, "ai_icon.png")

# Ensure necessary directories exist
if not os.path.exists(OUTPUT_IMG_FOLDER):
    os.makedirs(OUTPUT_IMG_FOLDER)
if not os.path.exists(SESSION_DIR):
    os.makedirs(SESSION_DIR)

# ----------------------- SELENIUM HELPER FUNCTIONS -----------------------
class CustomRoundedButton(ctk.CTkButton):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.configure(corner_radius=20, fg_color="green", font=("Arial", 18, "bold"))
        self.bind("<Enter>", self.on_hover)
        self.bind("<Leave>", self.on_leave)
        self.bind("<ButtonPress>", self.on_press)
        self.bind("<ButtonRelease>", self.on_release)
    def on_hover(self, event):
        self.configure(fg_color="#45A049")
    def on_leave(self, event):
        self.configure(fg_color="green")
    def on_press(self, event):
        self.configure(fg_color="red")
    def on_release(self, event):
        self.configure(fg_color="green")

class GuiLogger:
    def __init__(self, gui):
        self.gui = gui
    def log(self, message):
        self.gui.after(0, lambda: self.gui.log_live(message))
def normalize_phone(phone):
    phone = re.sub(r"[^\d+]", "", phone.strip())
    if not phone.startswith("+"):
        phone = "+91" + phone
    return phone
def selenium_login(gui_logger):
    gui_logger.log("🔄 Initializing Login Process...")
    options = webdriver.ChromeOptions()
    options.add_argument(f'--user-data-dir={SESSION_DIR}')
    options.add_argument('--profile-directory=Default')
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--lang=en')

    driver = None # Initialize driver to None
    try:
        gui_logger.log("🚗 Starting Chrome Driver...")
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
        driver.maximize_window()
        gui_logger.log("🌐 Navigating to WhatsApp Web...")
        driver.get("https://web.whatsapp.com")

        gui_logger.log("⏳ Waiting for WhatsApp to load chats (up to 60s)...")
        WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.XPATH, '//div[@role="grid"] | //canvas[@aria-label="Scan me!"]')) # Wait for chat list OR QR code
        )

        # Check if QR code is visible
        try:
            qr_canvas = driver.find_element(By.XPATH, '//canvas[@aria-label="Scan me!"]')
            if qr_canvas.is_displayed():
                gui_logger.log("📱 QR code scan required. Please scan within 5 minutes.")
                # Wait longer for the QR scan to complete
                WebDriverWait(driver, 300).until(
                    EC.presence_of_element_located((By.XPATH, '//div[@role="grid"]')) # Wait for chat list after scan
                )
                gui_logger.log("✅ WhatsApp Web logged in successfully after scan!")
            else:
                 # QR not visible, assume already logged in
                 gui_logger.log("✅ WhatsApp Web logged in successfully!")
        except Exception:
             # QR element not found, assume already logged in
             gui_logger.log("✅ WhatsApp Web logged in successfully!")

    except TimeoutException:
        gui_logger.log("❌ Timed out waiting for WhatsApp to load or QR scan.")
    except Exception as e:
        gui_logger.log(f"❌ Login Error: {e}")
    finally:
        if driver:
            try:
                driver.quit()
                gui_logger.log("🚪 Login browser closed.")
            except Exception as e:
                gui_logger.log(f"⚠️ Error closing login browser: {e}")

#-------------

# Helper function for Document/Media distinction
# This MUST be defined BEFORE selenium_send_bulk in your script.
def _is_media_file(file_path):
    """
    Checks file extension against known media types. Returns False for documents (PDF, DOCX, etc.).
    """
    if not file_path: return False
    ext = os.path.splitext(file_path)[1].lower()
    return ext in ['.png', '.jpg', '.jpeg', '.gif', '.mp4', '.mov', '.avi', '.3gp', '.webm', '.webp']
    
# NOTE: The copy_text_to_clipboard function must also be defined globally in your script.

def selenium_send_bulk(numbers, messages, attachments, min_delay, max_delay, gui_logger):
    # Imports assumed to be available globally in the main script context
    from selenium.webdriver.common.keys import Keys
    from selenium.common.exceptions import TimeoutException, WebDriverException 
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    
    # Reference to the main WabulkXpressApp instance 
    main_app = gui_logger.gui 
    
    # --- GLOBAL CONSTANTS (Assumed to be global in your script) ---
    SESSION_DIR = os.path.join(os.getcwd(), "selenium_session")
    # -----------------------------------------------------------------

    options = webdriver.ChromeOptions()
    options.add_argument(f'--user-data-dir={SESSION_DIR}')
    options.add_argument('--profile-directory=Default')
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--lang=en')
    
    try:
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
        driver.maximize_window()
        print("DEBUG [selenium_send_bulk]: WebDriver initialized successfully.")
    except Exception as e:
        gui_logger.log(f"❌ Critical Error: Failed to initialize WebDriver: {e}")
        return 0, len(numbers)
        
    success, failure = 0, 0
    INPUT_BOX_XPATH = '//div[@contenteditable="true"][@aria-placeholder="Type a message"]'
    LAST_MESSAGE_STATUS_XPATH = (
        '(//div[contains(@class, "message-out")]//span[contains(@data-icon, "msg-check") '
        'or contains(@data-icon, "msg-dblcheck") '
        'or contains(@data-icon, "msg-dblcheck-ack")])[last()]'
    )
    CHAT_PANEL_XPATH = '//div[@data-testid="conversation-panel-body"] | //div[@role="grid"]'
    
    for idx, number in enumerate(numbers):
        if not gui_logger.gui.sending:
            gui_logger.log("🛑 Stop signal received. Halting process...")
            break

        msg = messages[idx] if isinstance(messages, list) else messages
        attachment = attachments[idx] if attachments and idx < len(attachments) else None
        gui_logger.log(f"💬 [{idx+1}/{len(numbers)}] Sending to {number}...")
        
        url = f"https://web.whatsapp.com/send?phone={number}"
        
        try:
            driver.get(url)
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, CHAT_PANEL_XPATH))
            )
            time.sleep(random.uniform(1.5, 2.5))

        except TimeoutException as te:
            gui_logger.log(f"❗️ Chat panel did not load for {number} within 30s. Triggering network check.")
            # Network Backup Check 
            if hasattr(main_app, 'network_check_and_wait') and not main_app.network_check_and_wait(current_index=idx, total_count=len(numbers)):
                 failure += len(numbers) - idx
                 break
            else:
                 idx -= 1
                 failure -= 1
                 continue
        except WebDriverException as e:
            gui_logger.log(f"❌ CRITICAL Error during chat load for {number}: {e}")
            if hasattr(main_app, 'network_check_and_wait') and not main_app.network_check_and_wait(current_index=idx, total_count=len(numbers)):
                failure += len(numbers) - idx
                break
            else:
                idx -= 1
                failure -= 1
                continue
        except Exception as e:
             gui_logger.log(f"❗️ Error loading chat for {number}: {e}. Skipping.")
             failure += 1
             continue


        sent_something = False
        try:
            # --- 1. SEND ATTACHMENT (MEDIA OR DOCUMENT) ---
            if attachment and os.path.exists(attachment):
                is_media = _is_media_file(attachment) 
                
                try:
                    # 1. Click the main attachment button (pin icon)
                    attach_btn_xpath = '//span[@data-icon="plus-rounded"] | //span[@data-icon="attach-menu-plus"]'
                    attach_btn = WebDriverWait(driver, 15).until(
                        EC.element_to_be_clickable((By.XPATH, attach_btn_xpath))
                    )
                    attach_btn.click()
                    time.sleep(1.0) # Delay for menu to fully appear
                    
                    file_input = None
                    if is_media:
                        # 1a. Media Path: Find the image/video input (Accepts media types)
                        file_input_xpath = '//input[@accept="image/*,video/mp4,video/3gpp,video/quicktime"]' 
                        file_input = WebDriverWait(driver, 5).until(
                            EC.presence_of_element_located((By.XPATH, file_input_xpath))
                        )
                    else:
                        file_input_xpath = '//input[@accept="image/*,video/mp4,video/3gpp,video/quicktime"] | //input[@type="file"]'
                        print(f"DEBUG [selenium_send_bulk]: Finding file input: {file_input_xpath}")
                        file_input = WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.XPATH, file_input_xpath))
                        )
                        print("DEBUG [selenium_send_bulk]: Sending keys to file input.")
                        file_input.send_keys(os.path.abspath(attachment))
                        time.sleep(0.5)
                        

                    # 2. Send the file path (Triggers file upload dialog)
                    if file_input:
                        abs_path = os.path.abspath(attachment)
                        file_input.send_keys(abs_path)
                    
                    time.sleep(2.0) # Delay for browser to process the file upload
                    
                    # --- Caption/Media Preview Interaction (ONLY for media) ---
                    if is_media and msg: 
                        try:
                            # Wait for the caption input box that appears for media files
                            caption_box_xpath = '//div[@data-testid="preview-caption-input-container"]//div[@contenteditable="true"]'
                            caption_box = WebDriverWait(driver, 10).until(
                                EC.presence_of_element_located((By.XPATH, caption_box_xpath))
                            )
                            
                            # Use File 2's clipboard paste logic for caption
                            if copy_text_to_clipboard(msg):
                                caption_box.send_keys(Keys.CONTROL, 'v')
                            else:
                                for ch in msg:
                                    caption_box.send_keys(ch)
                                    time.sleep(random.uniform(0.02, 0.08))

                            gui_logger.log(f"✍️ Added caption for {number}")
                            msg = None # CRITICAL: Clear message so it's only sent as a caption
                        except TimeoutException:
                            print("DEBUG: Caption box not found. Sending media without caption.")
                        
                    # NOTE: Document messages (msg) remain intact here.
                        
                    time.sleep(2) # Wait after processing/upload

                    # 3. Send via JS Click (Confirms the attachment modal for both documents and media)
                    send_btn_xpath = '//button[@aria-label="Send"]'
                    send_btn = WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, send_btn_xpath))
                    )
                    driver.execute_script("arguments[0].click();", send_btn)
                    gui_logger.log(f"📎 Attachment sending initiated to {number}")
                    sent_something = True

                    # Wait until the modal closes (staleness of the button)
                    WebDriverWait(driver, 15).until(EC.staleness_of(send_btn))
                    time.sleep(0.5)

                except TimeoutException as te:
                    gui_logger.log(f"❌ Timed out during attachment process for {number}: {te}")
                except Exception as attach_err:
                     gui_logger.log(f"❌ Error sending attachment to {number}: {attach_err}")
                     
            # --- 2. SEND MESSAGE (Executed if it was a document or media with no caption) ---
            if msg:
                try:
                    input_box = WebDriverWait(driver, 15).until(
                        EC.element_to_be_clickable((By.XPATH, INPUT_BOX_XPATH))
                    )
                    # Use JavaScript click as a fallback
                    try:
                        input_box.click()
                    except Exception:
                         driver.execute_script("arguments[0].click();", input_box)

                    time.sleep(0.5)

                    # Use your File 2 logic for typing/pasting the message content
                    def _has_non_bmp(s: str) -> bool: return any(ord(c) > 0xFFFF for c in s)
                    if '\n' in msg or _has_non_bmp(msg):
                        # Paste via clipboard (File 2 exact logic)
                        if copy_text_to_clipboard(msg):
                            input_box.send_keys(Keys.CONTROL, 'v')
                        else:
                            # Fallback for multi-line/emoji if clipboard fails
                            lines = msg.split('\n')
                            for i, line in enumerate(lines):
                                # Send the whole line at once to handle emojis
                                input_box.send_keys(line)
                                time.sleep(random.uniform(0.05, 0.1))
                                
                                if i < len(lines) - 1:
                                    # Send Shift+Enter for a new line
                                    input_box.send_keys(Keys.SHIFT, Keys.ENTER)
                                    time.sleep(0.1)
                    else:
                        # Type out simple, single-line messages (File 2 exact logic)
                        for char in msg:
                            input_box.send_keys(char)
                            time.sleep(random.uniform(0.02, 0.08))

                    time.sleep(0.5)
                    input_box.send_keys(Keys.ENTER)
                    gui_logger.log(f"✅ Message sending initiated to {number}")
                    sent_something = True

                except TimeoutException as te:
                    gui_logger.log(f"❌ Timed out waiting for message input box for {number}: {te}")
                except Exception as msg_err:
                     gui_logger.log(f"❌ Error typing/sending message to {number}: {msg_err}")

        except Exception as outer_err:
            gui_logger.log(f"❌ Unexpected error during send process for {number}: {outer_err}")
            # Network Backup Check
            if hasattr(main_app, 'network_check_and_wait') and not main_app.network_check_and_wait(current_index=idx, total_count=len(numbers)):
                 failure += len(numbers) - idx 
                 break 
            else:
                 idx -= 1
                 failure -= 1
                 continue


        # --- 3. WAIT FOR SEND CONFIRMATION ---
        action_succeeded = False 
        if sent_something:
            try:
                WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.XPATH, LAST_MESSAGE_STATUS_XPATH))
                )
                gui_logger.log(f"✔️ Send confirmed for {number}")
                action_succeeded = True 
            except TimeoutException:
                gui_logger.log(f"⚠️ Send confirmation checkmark not found for {number} within 30s.")
            except Exception as check_err:
                gui_logger.log(f"⚠️ Error checking send status for {number}: {check_err}")
                
        if action_succeeded:
            success += 1
        else:
            failure += 1

        # --- 4. DELAY LOGIC ---
        # Add 4s fixed delay after any attachment send for stability
        if attachment and os.path.exists(attachment) and sent_something:
            gui_logger.log("⏳ Adding fixed 4s delay after attachment...")
            time.sleep(4)
            
        if not gui_logger.gui.sending:
            gui_logger.log("🛑 Stop signal received. Halting process...")
            break

        random_delay = random.uniform(min_delay, max_delay)
        gui_logger.log(f"⏳ Waiting for {random_delay:.2f} seconds before next contact...")

        # Sleep incrementally to check for stop signal
        start_delay_time = time.time()
        while time.time() - start_delay_time < random_delay:
            if not gui_logger.gui.sending: break
            time.sleep(0.1) 

        if not gui_logger.gui.sending: 
            break

        # --- 5. 10 Contact Pause ---
        if (idx + 1) % 10 == 0 and (idx + 1) < len(numbers) and gui_logger.gui.sending:
            gui_logger.log(f"⏸️ 10 contacts processed. Pausing for 30 seconds...")
            start_pause_time = time.time()
            while time.time() - start_pause_time < 30:
                 if not gui_logger.gui.sending: break
                 time.sleep(0.1)

            if not gui_logger.gui.sending:
                 break

    # --- Cleanup ---
    try:
        driver.quit()
    except Exception as e:
        gui_logger.log(f"⚠️ Error quitting WebDriver: {e}")

    gui_logger.log(f"📊 Done: Success={success}, Failure={failure}")
    return success, failure










# ----------------------- UTILITY FUNCTIONS -----------------------

def network_check(gui_logger):
    """Performs a simple connectivity check using Google's public DNS."""
    try:
        # Using a reliable, fast target (Google's public DNS)
        requests.get("https://8.8.8.8", timeout=10)
        return True
    except requests.exceptions.RequestException as e:
        gui_logger.log(f"🌐 Network check failed: {e}")
        return False
    
def center_window(win):
    win.update_idletasks()
    width = win.winfo_width()
    height = win.winfo_height()
    x = (win.winfo_screenwidth() // 2) - (width // 2)
    y = (win.winfo_screenheight() // 2) - (height // 2)
    win.geometry(f"{width}x{height}+{x}+{y}")

def copy_text_to_clipboard(text):
    try:
        win32clipboard.OpenClipboard()
        win32clipboard.EmptyClipboard()
        win32clipboard.SetClipboardData(win32clipboard.CF_UNICODETEXT, text)
        win32clipboard.CloseClipboard()
        return True
    except Exception as e:
        print(f"Clipboard error: {e}")
        return False

def copy_file_to_clipboard(file_path):
    DROPFILES_FORMAT = "IiiIII"
    DROPFILES_SIZE = struct.calcsize(DROPFILES_FORMAT)
    offset = DROPFILES_SIZE
    file_list = file_path + "\0\0"
    dropfiles_struct = struct.pack(DROPFILES_FORMAT, DROPFILES_SIZE, 0, 0, 0, offset, 1)
    data = dropfiles_struct + file_list.encode("utf-16-le")
    try:
        win32clipboard.OpenClipboard()
        win32clipboard.EmptyClipboard()
        win32clipboard.SetClipboardData(win32clipboard.CF_HDROP, data)
        win32clipboard.CloseClipboard()
    except Exception as e:
        print(f"Error copying file to clipboard: {e}")

# ----------------------- UTILITY FUNCTIONS -----------------------
def round_corners(image_path, size=(200, 200), corner_radius=20):
    """Rounds the corners of an image and returns a CTkImage."""
    try:
        img = Image.open(image_path).resize(size, Image.Resampling.LANCZOS).convert("RGBA")
        mask = Image.new("L", size, 0)
        draw = ImageDraw.Draw(mask)
        # Draw rounded rectangle
        draw.rounded_rectangle((0, 0) + size, radius=corner_radius, fill=255)
        img.putalpha(mask)
        return ctk.CTkImage(light_image=img, dark_image=img, size=size)
    except Exception as e:
        print(f"Error rounding corners: {e}")
        return None

# ----------------------- CUSTOM WIDGET CLASSES -----------------------
class HoverHint(ctk.CTkToplevel):
    def __init__(self, widget, hint_text, image_path, *args, **kwargs):
        super().__init__(widget.master, *args, **kwargs)
        self.overrideredirect(True)
        self.geometry("350x150")
        self.configure(bg="transparent")
        
        self.frame = ctk.CTkFrame(self, corner_radius=15)
        self.frame.pack(expand=True, fill="both", padx=5, pady=5)
        self.text_frame = ctk.CTkFrame(self.frame, fg_color="transparent")
        self.text_frame.pack(side="left", fill="both", expand=True, padx=(15, 10), pady=10)
        
        self.hint_label = ctk.CTkLabel(self.text_frame, 
                                       text=hint_text, 
                                       anchor="w", 
                                       justify="left", 
                                       wraplength=170,
                                       font=("Arial", 14))
        self.hint_label.pack(expand=True, fill="both")
        try:
            box_width = 350
            image_width = int(box_width * 0.35)
            size = (image_width, image_width)
            corner_radius = 15

            img = Image.open(image_path).resize(size, Image.Resampling.LANCZOS).convert("RGBA")
            mask = Image.new("L", size, 0)
            draw = ImageDraw.Draw(mask)
            draw.rounded_rectangle((0, 0) + size, radius=corner_radius, fill=255)
            img.putalpha(mask)
            
            self.hint_image = ctk.CTkImage(light_image=img, dark_image=img, size=size)
            
        except Exception as e:
            print(f"Image load error: {e}")
            self.hint_image = None
            
        self.image_label = ctk.CTkLabel(self.frame, image=self.hint_image, text="")
        self.image_label.pack(side="right", padx=15, pady=15)
        
        self.withdraw()
        self.widget = widget
        self.widget.bind("<Enter>", self.show_hint)
        self.widget.bind("<Leave>", self.hide_hint)
        self.widget.bind("<Motion>", self.move_hint)
        
        self.update_theme()

    def update_theme(self):
        mode = ctk.get_appearance_mode().lower()
        fg_color = DARK_INNER if mode == "dark" else LIGHT_INNER
        text_color = "white" if mode == "dark" else "black"
        
        if self.winfo_exists():
            self.frame.configure(fg_color=fg_color)
            self.hint_label.configure(text_color=text_color)
        
    def show_hint(self, event=None):
        self.deiconify()
        self.lift()
        self.move_hint(event)

    def hide_hint(self, event=None):
        self.withdraw()

    def move_hint(self, event=None):
        if event:
            x = event.x_root + 10
            y = event.y_root + 10
            self.geometry(f"+{x}+{y}")

class AnimatedCTkButton(ctk.CTkButton):
    def __init__(self, *args, hover_fg_color="#0050a0", **kwargs):
        super().__init__(*args, **kwargs)
        self.hover_fg_color = hover_fg_color
        self.original_fg_color = self.cget("fg_color")
        self.bind("<Enter>", self.on_enter)
        self.bind("<Leave>", self.on_leave)
    def on_enter(self, event):
        self.configure(fg_color=self.hover_fg_color)
    def on_leave(self, event):
        self.configure(fg_color=self.original_fg_color)
class TkinterVideo(tk.Label):
    def __init__(self, master, path, scaled=True, keep_aspect=False, *args, **kwargs):
        super().__init__(master, *args, **kwargs)
        self.path = path
        self.scaled = scaled
        self.keep_aspect = keep_aspect
        self._stop = False
        self.frame_queue = queue.Queue()
        self.current_frame = None
        self._load_thread = threading.Thread(target=self._decode_video, daemon=True)
        self._load_thread.start()
        self.after(0, self._update_image)
    def _decode_video(self):
        try:
            import av
        except ImportError:
            print("av module not installed. Please run: pip install av")
            return
        try:
            container = av.open(self.path)
            stream = container.streams.video[0]
            stream.thread_type = "AUTO"
            delay = 1 / float(stream.average_rate)
            for frame in container.decode(stream):
                if self._stop:
                    break
                img = frame.to_image()
                self.frame_queue.put(img)
                time.sleep(delay)
            container.close()
        except Exception as e:
            print(f"Error in TkinterVideo: {e}")
    def _update_image(self):
        try:
            if not self.frame_queue.empty():
                img = self.frame_queue.get_nowait()
                if self.scaled:
                    w = self.winfo_width()
                    h = self.winfo_height()
                    if w and h:
                        if self.keep_aspect:
                            img = ImageOps.contain(img, (w, h))
                        else:
                            img = img.resize((w, h), Image.Resampling.LANCZOS)
                self.current_frame = ImageTk.PhotoImage(img)
                self.configure(image=self.current_frame)
        except Exception as e:
            print(f"Error updating image: {e}")
        if not self._stop:
            self.after(30, self._update_image)
    def stop(self):
        self._stop = True

    def pause(self):
        self._stop = True

    def play(self):
        if self._stop:
            self._stop = False
            self._load_thread = threading.Thread(target=self._decode_video, daemon=True)
            self._load_thread.start()
            self.after(0, self._update_image)

class ProgressPopup(ctk.CTkToplevel):
    def __init__(self, parent, title, total):
        super().__init__(parent)
        self.transient(parent)
        self.title(title)
        self.iconbitmap(TITLE_ICON_PATH)
        self.geometry("500x300")
        self.resizable(False, False)
        self.lift()
        self.attributes("-topmost", True)
        
        mode = ctk.get_appearance_mode().lower()
        bg_color = DARK_BG if mode == "dark" else LIGHT_BG
        fg_color = DARK_FG if mode == "dark" else LIGHT_FG
        
        self.configure(fg_color=bg_color)
        self.total = total
        self.current = 0
        frame = ctk.CTkFrame(self, corner_radius=10, fg_color=fg_color)
        frame.pack(expand=True, fill="both", padx=20, pady=20)
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)
        self.gif_label = ctk.CTkLabel(frame, text="")
        self.gif_label.grid(row=0, column=0, padx=10, pady=(30,10))
        self.load_gif(LOADING_GIF_PATH, size=(150, 150))
        self.progress_label = ctk.CTkLabel(frame, text=f"{self.current}/{self.total}", font=("Arial", 28, "bold"))
        self.progress_label.grid(row=1, column=0, padx=10, pady=(10,30))
        center_window(self)

    def load_gif(self, path, size=(150, 150)):
        try:
            image = Image.open(path)
            image = image.resize(size)
            self.gif_image = ImageTk.PhotoImage(image)
            self.gif_label.configure(image=self.gif_image)
        except Exception as e:
            print("Error loading GIF:", e)

    def update_progress(self, current):
        self.current = current
        if self.winfo_exists():
            self.after(0, lambda: self.progress_label.configure(text=f"{self.current}/{self.total}"))
            self.update_idletasks()

    def close(self):
        if self.winfo_exists():
            self.destroy()

class AnimatedGIF(tk.Label):
    def __init__(self, master, filename, delay=100):
        self.master = master
        self.filename = filename
        self.delay = delay
        im = Image.open(filename)
        self.frames = []
        try:
            for i in range(1000):
                im.seek(i)
                frame = ImageTk.PhotoImage(im.copy())
                self.frames.append(frame)
        except EOFError:
            pass
        self.idx = 0
        super().__init__(master, image=self.frames[0])
        self.after(self.delay, self.play)

    def play(self):
        self.idx = (self.idx + 1) % len(self.frames)
        self.configure(image=self.frames[self.idx])
        self.after(self.delay, self.play)
        
class AnimatedCTkButton(ctk.CTkButton):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.bind("<Enter>", self.on_enter)
        self.bind("<Leave>", self.on_leave)
        self.bind("<ButtonPress-1>", self.on_press)
        self.bind("<ButtonRelease-1>", self.on_release)
    def on_enter(self, event):
        pass
    def on_leave(self, event):
        pass
    def on_press(self, event):
        self.configure(fg_color="#388E3C")
    def on_release(self, event):
        self.configure(fg_color=self.cget("fg_color"))
def round_corners(image_path, size=(200, 200), corner_radius=20):
    try:
        img = Image.open(image_path).resize(size, Image.Resampling.LANCZOS).convert("RGBA")
        mask = Image.new("L", size, 0)
        draw = ImageDraw.Draw(mask)
        draw.rounded_rectangle((0, 0) + size, radius=corner_radius, fill=255)
        img.putalpha(mask)
        return ctk.CTkImage(light_image=img, dark_image=img, size=size)
    except Exception:
        return None
    
def center_window(win):
    win.update_idletasks()
    width = win.winfo_width()
    height = win.winfo_height()
    x = (win.winfo_screenwidth() // 2) - (width // 2)
    y = (win.winfo_screenheight() // 2) - (height // 2)
    win.geometry(f"{width}x{height}+{x}+{y}")


class MergeReplacePopup(ctk.CTkToplevel):
    """A themed popup to ask the user if they want to Merge or Replace existing data."""
    def __init__(self, master, launch_mapper_callback, *args, **kwargs):
        super().__init__(master, *args, **kwargs)
        self.title("Import Mode")
        self.geometry("380x200") # Slightly larger
        self.resizable(False, False)
        self.attributes("-topmost", True)
        self.transient(master)
        self.launch_mapper_callback = launch_mapper_callback # Callback to launch MappingPopup

        # --- Theme ---
        mode = ctk.get_appearance_mode().lower()
        bg_color = DARK_BG if mode == "dark" else LIGHT_BG
        fg_color = DARK_FG if mode == "dark" else LIGHT_FG
        
        self.configure(fg_color=bg_color)
        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self.bind("<Escape>", lambda e: self.destroy())
        
        # --- Layout ---
        container = ctk.CTkFrame(self, fg_color=fg_color, corner_radius=15)
        container.pack(fill="both", expand=True, padx=15, pady=15)
        container.grid_columnconfigure(0, weight=1)
        container.grid_rowconfigure((0, 1), weight=0)
        container.grid_rowconfigure(2, weight=1) # Spacer

        # 1. Title
        ctk.CTkLabel(container, text="Existing Data Found", font=("Arial", 18, "bold")).grid(row=0, column=0, pady=(10, 5))
        ctk.CTkLabel(container, text="How do you want to import the new data?", font=("Arial", 14)).grid(row=1, column=0, pady=(0, 15))
        
        # 2. Button Frame
        btn_frame = ctk.CTkFrame(container, fg_color="transparent")
        btn_frame.grid(row=3, column=0, pady=(10, 15), sticky="ew")
        btn_frame.grid_columnconfigure((0, 1), weight=1)

        # 3. Action Logic
        def on_select_mode(merge_mode):
            self.destroy() # Close this choice popup
            self.launch_mapper_callback(merge_mode) # Launch the next window

        # Button Style
        btn_font = ("Arial", 16, "bold")

        # 4. Merge Button (Green)
        merge_btn = ctk.CTkButton(
            btn_frame, 
            text="Merge\n(Add to existing)", 
            command=lambda: on_select_mode(merge_mode=True),
            fg_color="green",
            hover_color="#45A049",
            font=btn_font,
            height=60
        )
        merge_btn.grid(row=0, column=0, padx=(15, 7), sticky="ew")

        # 5. Replace Button (Red)
        replace_btn = ctk.CTkButton(
            btn_frame, 
            text="Replace\n(Clear old data)", 
            command=lambda: on_select_mode(merge_mode=False),
            fg_color="#D32F2F", # Red color
            hover_color="#B71C1C",
            font=btn_font,
            height=60
        )
        replace_btn.grid(row=0, column=1, padx=(7, 15), sticky="ew")
        
        center_window(self)
        


# ----------------------- (NEW) THEMED POPUP CLASS -----------------------
class ThemedPopup(ctk.CTkToplevel):
    """A general-purpose themed popup for errors, warnings, and info."""
    def __init__(self, master, title="Alert", message="Something happened.", icon_type="info", *args, **kwargs):
        super().__init__(master, *args, **kwargs)
        self.title(title)
        self.geometry("450x200") # Slightly wider for messages
        self.resizable(False, False)
        self.attributes("-topmost", True)
        self.transient(master) # Keep on top of parent

        # --- Theme ---
        mode = ctk.get_appearance_mode().lower()
        bg_color = DARK_BG if mode == "dark" else LIGHT_BG
        fg_color = DARK_FG if mode == "dark" else LIGHT_FG
        text_color = "white" if mode == "dark" else "black"

        self.configure(fg_color=bg_color)
        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self.bind("<Escape>", lambda e: self.destroy())

        # --- Layout ---
        container = ctk.CTkFrame(self, fg_color=fg_color, corner_radius=15)
        container.pack(fill="both", expand=True, padx=15, pady=15)
        container.grid_columnconfigure(0, weight=1)
        container.grid_rowconfigure(0, weight=0) # Icon row
        container.grid_rowconfigure(1, weight=1) # Message row
        container.grid_rowconfigure(2, weight=0) # Button row

        # --- Icon (Based on type) ---
        icon_path = ""
        if icon_type == "error":
            icon_path = os.path.join(BIN_FOLDER, "error_icon.png") # Create this image
        elif icon_type == "warning":
            icon_path = os.path.join(BIN_FOLDER, "warning_icon.png") # Create this image
        else: # Default to info
            icon_path = os.path.join(BIN_FOLDER, "info_icon.png") # Create this image

        icon_image = None
        icon_size = (50, 50)
        if os.path.exists(icon_path):
            try:
                img = Image.open(icon_path).resize(icon_size, Image.Resampling.LANCZOS)
                icon_image = ctk.CTkImage(light_image=img, dark_image=img, size=icon_size)
            except Exception as e:
                print(f"Error loading popup icon {icon_path}: {e}")

        icon_label = ctk.CTkLabel(container, image=icon_image, text="")
        icon_label.grid(row=0, column=0, pady=(10, 5))

        # --- Message ---
        message_label = ctk.CTkLabel(
            container,
            text=message,
            font=("Arial", 16),
            wraplength=380, # Allow text wrapping
            justify="center",
            text_color=text_color
        )
        message_label.grid(row=1, column=0, pady=10, padx=20, sticky="nsew")

        # --- OK Button ---
        ok_button = ctk.CTkButton(
            container,
            text="OK",
            width=120,
            height=40,
            font=("Arial", 16, "bold"),
            fg_color="#0078D7", # Use a consistent blue for OK
            hover_color="#005fa3",
            corner_radius=10,
            command=self.destroy
        )
        ok_button.grid(row=2, column=0, pady=(10, 15))

        center_window(self) # Center the popup
        # self.fade_in() # Optional: Add fade-in if desired

# ... (inside WabulkXpress.py)

# ----------------------- NEW THEMED POPUP CLASS (Enhanced AlertPopup) -----------------------
class AlertPopup(ctk.CTkToplevel):
    def __init__(self, master, msg, is_used=False, key=None, icon_type="error"):
        super().__init__(master)
        self.title("Alert")
        self.geometry("400x250")
        self.resizable(False, False)
        self.attributes("-topmost", True)
        self.transient(master)
        
        mode = ctk.get_appearance_mode().lower()
        # Use a near-pitch black background for the popup itself
        bg_color = DARK_BG if mode == "dark" else LIGHT_BG
        # Use a standard foreground color for the inner frame
        fg_color = DARK_FG if mode == "dark" else LIGHT_FG
        
        self.configure(fg_color=bg_color) 
        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self.bind("<Escape>", lambda e: self.destroy())
        
        # --- Layout ---
        self.container = ctk.CTkFrame(self, fg_color=fg_color, corner_radius=15) 
        self.container.pack(fill="both", expand=True, padx=15, pady=15)
        self.container.grid_columnconfigure(0, weight=1)
        self.container.grid_rowconfigure(0, weight=0) # Image
        self.container.grid_rowconfigure(1, weight=1) # Message
        self.container.grid_rowconfigure(2, weight=0) # Buttons

        # --- Image ---
        if icon_type == "error":
            img_path = ERROR_IMAGE_PATH
        elif icon_type == "key":
            img_path = PRODUCT_IMAGE_PATH
        else:
            img_path = ERROR_IMAGE_PATH # Default

        alert_img = round_corners(img_path, (80, 80), corner_radius=15)
        if alert_img:
            img_label = ctk.CTkLabel(self.container, image=alert_img, text="")
            img_label.grid(row=0, column=0, pady=(10, 5))
        
        # --- Message ---
        label = ctk.CTkLabel(self.container, text=msg, font=("Arial", 14, "bold"), wraplength=350, justify="center")
        label.grid(row=1, column=0, pady=5, padx=10, sticky="nsew")

        # --- Buttons ---
        buttons_frame = ctk.CTkFrame(self.container, fg_color="transparent")
        buttons_frame.grid(row=2, column=0, pady=(10, 5))
        buttons_frame.grid_columnconfigure((0, 1), weight=1)

        if is_used:
            help_btn = AnimatedCTkButton(
                buttons_frame, 
                text="Get Help", 
                width=100, height=40, font=("Arial", 16, "bold"),
                fg_color="#4CAF50", hover_color="#45A049", corner_radius=15,
                command=lambda: self.open_help(key)
            )
            help_btn.grid(row=0, column=0, padx=10)
        
        ok_btn_text = "Close" if is_used else "OK" # Use "Close" for Get Help scenario
        ok_btn_color = "#4CAF50" if is_used else "#333333" # Green if Get Help is not present
        ok_btn_hover = "#45A049" if is_used else "#555555"

        ok_btn = ctk.CTkButton(
            buttons_frame, 
            text=ok_btn_text, 
            width=100, height=40, font=("Arial", 16, "bold"),
            fg_color=ok_btn_color, hover_color=ok_btn_hover, corner_radius=15,
            command=self.destroy
        )
        # Position OK button to the right if Get Help is present, or center if not
        if is_used:
            ok_btn.grid(row=0, column=1, padx=10)
        else:
            # If no help button, center the OK button
            ok_btn.grid(row=0, column=0, columnspan=2, padx=10) 
            
        center_window(self)
        self.fade_in()

    def open_help(self, key):
        # ... (unchanged)
        if key:
            message = f"Hi! Please revoke my previous product key—I’ve changed devices. Key: {key}. Thanks! 🙏"
        else:
            message = "Hi! Please revoke my previous product key—I’ve changed devices. Thanks! 🙏"
        encoded = message.replace(" ", "%20")
        webbrowser.open(f"https://wa.me/{WHATSAPP_NUMBER}?text={encoded}")

    def fade_in(self, alpha=0.0):
        # ... (unchanged)
        if alpha < 1.0:
            self.attributes("-alpha", alpha)
            self.after(50, lambda: self.fade_in(alpha + 0.1))
        else:
            self.attributes("-alpha", 1.0)

# ----------------------- NEW IMPORT POPUP (PHASE 1) -----------------------
class NewImportPopup(ctk.CTkToplevel):
    def __init__(self, master, on_success_callback, *args, **kwargs):
        super().__init__(master, *args, **kwargs)
        self.title("Import Contacts")
        self.iconbitmap(TITLE_ICON_PATH)
        self.geometry("600x320")
        self.resizable(False, False)
        self.on_success_callback = on_success_callback
        self.wm_attributes("-topmost", True)
        self.transient(master)
        
        mode = ctk.get_appearance_mode().lower()
        bg_color = DARK_BG if mode == "dark" else LIGHT_BG
        fg_color = DARK_FG if mode == "dark" else LIGHT_FG
        inner_color = DARK_INNER if mode == "dark" else LIGHT_INNER
        self.configure(fg_color=bg_color) 
        
        # --- Main Layout (2 columns) ---
        self.grid_columnconfigure(0, weight=1) # Left controls
        self.grid_columnconfigure(1, weight=1) # Right image
        self.grid_rowconfigure(0, weight=1)
        
        # --- Left Frame (Controls) ---
        left_frame = ctk.CTkFrame(self, fg_color=fg_color, corner_radius=15)
        left_frame.grid(row=0, column=0, sticky="nsew", padx=(20, 10), pady=20)
        
        left_frame.grid_columnconfigure(0, weight=1)
        left_frame.grid_rowconfigure(0, weight=0) # Title
        left_frame.grid_rowconfigure(1, weight=1) # Drop box
        
        title_label = ctk.CTkLabel(left_frame, text="Import Contacts", font=("Arial", 24, "bold"))
        title_label.grid(row=0, column=0, pady=(10, 10))

        # --- Load Theme-Aware Upload Icon ---
        self.upload_icon = None
        try:
            img_light = Image.open(UPLOAD_ICON_LIGHT_PATH).resize((80, 80), Image.Resampling.LANCZOS)
            img_dark = Image.open(UPLOAD_ICON_DARK_PATH).resize((80, 80), Image.Resampling.LANCZOS)
            self.upload_icon = ctk.CTkImage(light_image=img_light, dark_image=img_dark, size=(80, 80))
        except Exception as e:
            print(f"Error loading upload icons: {e}")

        # "Drop Box" Button
        self.drop_box_btn = ctk.CTkButton(
            left_frame, 
            text="Click to Select File\nOr Paste Google Sheet URL\nOr Drag & Drop File" if self.upload_icon else "+\n\nClick, Paste, or Drop File",
            image=self.upload_icon,
            compound="top", # Image above text
            font=("Arial", 16), 
            command=self.select_file,
            corner_radius=10,
            fg_color=inner_color,
            text_color="gray",
            border_width=2,
            border_color="#555555",
            hover_color="#333333" if mode == "dark" else "#E0E0E0",
            height=180 # Taller box
        )
        self.drop_box_btn.grid(row=1, column=0, sticky="nsew", padx=20, pady=(10, 20))
        
        # --- Drag & Drop Bindings ---
        self.drop_box_btn.drop_target_register(tkinterdnd2.DND_FILES)
        self.drop_box_btn.dnd_bind('<<Drop>>', self.handle_drop)

        # --- Right Frame (Image) ---
        right_frame = ctk.CTkFrame(self, fg_color=fg_color, corner_radius=15)
        right_frame.grid(row=0, column=1, sticky="nsew", padx=(10, 20), pady=20)
        right_frame.grid_rowconfigure(0, weight=1)
        right_frame.grid_columnconfigure(0, weight=1)
        
        try:
            # Use the global round_corners function
            hint_img = round_corners(os.path.join(BIN_FOLDER, "hint_import.png"), (200, 200), corner_radius=15)
            if hint_img:
                img_label = ctk.CTkLabel(right_frame, image=hint_img, text="")
                img_label.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
        except Exception as e:
            print(f"Error loading hint_import.png: {e}")
            ctk.CTkLabel(right_frame, text="Image Error", font=("Arial", 16)).grid(row=0, column=0)

        center_window(self)
        # Bind paste event to the whole window
        self.bind("<Control-v>", self.handle_paste)
        self.bind("<Escape>", lambda e: self.destroy())

    def select_file(self):
        """Opens file dialog and triggers callback with the file path."""
        path = filedialog.askopenfilename(
            filetypes=[("Excel Files", "*.xlsx;*.xls"), ("CSV Files", "*.csv")],
            parent=self
        )
        if path:
            self.destroy() 
            self.on_success_callback(file_path=path)

    def handle_drop(self, event):
        """
        Handles a file being dropped onto the button.
        The fix ensures the action is queued on the main app's thread.
        """
        try:
            file_path = event.data
            # DND data can be wrapped in braces {}
            if file_path.startswith('{') and file_path.endswith('}'):
                file_path = file_path[1:-1]
            
            # Check for valid file types
            if file_path.endswith(('.xlsx', '.xls', '.csv')):
                
                # 1. Get the root application window instance (WabulkXpressApp)
                main_app_root = self.master.master 
                
                # 2. Define the successful next step (The Fix)
                def execute_next_step(path):
                    """
                    This function is guaranteed to run on the main thread
                    via main_app_root.after(0).
                    """
                    # Destroy THIS window immediately.
                    if self.winfo_exists():
                        self.destroy()
                        
                    # Call the original success callback, which is what 
                    # select_file() also calls. This launches the MappingPopup.
                    self.on_success_callback(file_path=path) 

                # 3. Queue the execution on the main app's thread
                main_app_root.after(0, lambda p=file_path: execute_next_step(p))

            else:
                ThemedPopup(self, title="File Error", message="Invalid file type. Please drop an .xlsx, .xls, or .csv file.", icon_type="error")
        except Exception as e:
            ThemedPopup(self, title="Drop Error", message=f"An unexpected error occurred: {e}", icon_type="error")

    def handle_paste(self, event=None):
        try:
            url = self.clipboard_get()
            if not url or not (url.startswith("http://") or url.startswith("https://")):
                ThemedPopup(self, title="Paste Error", message="Clipboard does not contain a valid URL.", icon_type="info")
                return
            
            if "docs.google.com/spreadsheets" not in url:
                ThemedPopup(self, title="Paste Warning", message="This doesn't look like a Google Sheet URL, but we'll try to fetch it anyway.", icon_type="warning")
            
            self._start_fetch(url)
            
        except tk.TclError:
            ThemedPopup(self, title="Paste Error", message="Could not read from clipboard.", icon_type="error")
        except Exception as e:
            ThemedPopup(self, title="Unexpected Error", message=f"An error occurred during paste: {e}", icon_type="error")


    def _start_fetch(self, url):
        """
        Starts the Google Sheet URL fetching process in a separate thread.
        """
        if not url:
            ThemedPopup(self, title="Error", message="No URL provided.", icon_type="error")
            return
            
        self.drop_box_btn.configure(text="Fetching URL...", state="disabled", image=None)
        
        main_app_instance = self.master.master 

        def download_thread_task():
            err_msg = None
            try:
                # 1. Validate and Extract Google Sheet IDs
                match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)(?:/.*gid=(\d+))?', url)
                
                if not match:
                    err_msg = "Invalid Google Sheet URL format. URL must contain '/spreadsheets/d/...'."
                
                if not err_msg:
                    sheet_id = match.group(1)
                    gid = match.group(2) or "0" 
                    
                    # 2. Build the public CSV export URL
                    csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"            
                    
                    # 3. Perform the network request 
                    response = requests.get(csv_url, timeout=15)
                    response.raise_for_status() 
                    data = response.text
                    
                    # --- On Success: Schedule on main thread ---
                    if main_app_instance:
                        main_app_instance.after(0, lambda: self.on_fetch_success(data))
                    return 
                    
            except requests.exceptions.Timeout:
                err_msg = "Request timed out (15s limit). Check your internet connection."
            except requests.exceptions.RequestException as e:
                err_msg = f"Failed to fetch URL: {e}"
            except Exception as e:
                err_msg = f"An unexpected error occurred during fetch: {e}"

            # --- On Failure: Schedule on main thread (safe communication) ---
            if err_msg and main_app_instance:
                main_app_instance.after(0, lambda msg=err_msg: self.on_fetch_error(msg))

        threading.Thread(target=download_thread_task, daemon=True).start()


    def on_fetch_success(self, data):
        """Callback on successful URL fetch. Triggers main callback with CSV data."""
        # This function is called via self.after(0) from the fetch thread
        self.on_success_callback(csv_data=data)
        self.destroy()

    def on_fetch_error(self, msg):
        """Callback on failed URL fetch. Shows error and re-enables UI."""
        # This function is called via self.after(0) from the fetch thread
        ThemedPopup(self, title="Error", message=msg, icon_type="error")
        self.drop_box_btn.configure(
            text="Click to Select File\nOr Paste Google Sheet URL\nOr Drag & Drop File", 
            state="normal", 
            image=self.upload_icon
        )


class MappingPopup(ctk.CTkToplevel):
    """
    Phase 2 Popup: Shows data preview and asks user to map columns and row ranges.
    
    FIX: The left control panel initialization is deferred to ensure
    it draws correctly without conflicting with the complex preview grid 
    or the drag-and-drop thread hand-off.
    """
    def __init__(self, master, on_import_callback, merge_mode, file_path=None, csv_data=None, *args, **kwargs):
        super().__init__(master, *args, **kwargs)
        self.title("Map Columns and Rows")
        self.iconbitmap(TITLE_ICON_PATH)
        self.geometry("1000x500") # Start with a slightly taller window for better preview
        self.minsize(800, 500)   # Set a minimum size
        self.resizable(True, True) # Allow resizing
        self.on_import_callback = on_import_callback
        self.merge_mode = merge_mode
        self.file_path = file_path
        self.csv_data = csv_data # Store in-memory CSV data
        self.is_csv = (csv_data is not None) or (file_path and file_path.endswith('.csv'))
        
        self.wm_attributes("-topmost", True)
        self.transient(master)

        # --- Theme Colors ---
        mode = ctk.get_appearance_mode().lower()
        bg_color = DARK_BG if mode == "dark" else LIGHT_BG
        fg_color = DARK_FG if mode == "dark" else LIGHT_FG
        self.inner_color = DARK_INNER if mode == "dark" else LIGHT_INNER
        
        self.configure(fg_color=bg_color) 
        
        # --- Main Layout (2 columns) ---
        self.grid_columnconfigure(0, weight=1) # Left controls
        self.grid_columnconfigure(1, weight=2) # Right preview (wider)
        self.grid_rowconfigure(0, weight=1)
        
        # --- Left Frame (Placeholder - DEFERRED INITIALIZATION FIX) ---
        self.left_frame = ctk.CTkFrame(self, corner_radius=15, fg_color=fg_color)
        self.left_frame.grid(row=0, column=0, sticky="nsew", padx=(20, 10), pady=20)
        # Placeholder Label for immediate feedback
        ctk.CTkLabel(self.left_frame, text="Loading Controls...", font=("Arial", 20)).pack(padx=20, pady=200) 
        
        # --- Right Frame (Preview) ---
        self.right_frame = ctk.CTkFrame(self, corner_radius=15, fg_color=fg_color)
        self.right_frame.grid(row=0, column=1, sticky="nsew", padx=(10, 20), pady=20)
        self.right_frame.grid_rowconfigure(1, weight=1)
        self.right_frame.grid_columnconfigure(0, weight=1)
        
        ctk.CTkLabel(self.right_frame, text="Data Preview (First 200 Rows)", font=("Arial", 18, "bold")).grid(row=0, column=0, pady=(10, 10))
        
        # Create a scrollable frame to hold the grid
        self.preview_grid_frame = ctk.CTkScrollableFrame(
            self.right_frame, 
            fg_color=self.inner_color, 
            border_width=1,
            border_color="#555555"
        )
        self.preview_grid_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 10))
        
        # --- Finalize ---
        self.load_preview_and_detect_headers() # Load data and the complex preview
        
        center_window(self)
        self.bind("<Return>", lambda e: self.on_import())
        self.bind("<Escape>", lambda e: self.destroy())

        # --- THE CORE FIX: Defer the control panel initialization ---
        self.after(50, self._initialize_controls) # 50ms delay for UI stability

    def toggle_advanced(self):
        """Shows or hides the advanced options frame for row ranges."""
        if self.advanced_visible:
            self.advanced_frame.grid_remove()
            self.toggle_advanced_btn.configure(text="Advanced Options ▾")
            self.advanced_visible = False
        else:
            # Grid it right below the toggle button
            self.advanced_frame.grid(row=6, column=0, columnspan=2, sticky="ew", padx=20)
            self.toggle_advanced_btn.configure(text="Advanced Options ▴")
            self.advanced_visible = True

    def _initialize_controls(self):
        """
        FIX IMPLEMENTATION: Creates and places all the column mapping fields and buttons.
        This is called via self.after(50) to ensure the window is stable.
        """
        
        # Clear the temporary loading label
        for widget in self.left_frame.winfo_children():
            widget.destroy()

        # Reconfigure the grid inside the left_frame
        self.left_frame.grid_columnconfigure(0, weight=1)
        self.left_frame.grid_rowconfigure(1, weight=0) # Field Frame
        self.left_frame.grid_rowconfigure(5, weight=0) # Adv Button
        self.left_frame.grid_rowconfigure(6, weight=0) # Adv Frame
        self.left_frame.grid_rowconfigure(7, weight=1) # Spacer
        self.left_frame.grid_rowconfigure(8, weight=0) # Import Button
        
        title_label = ctk.CTkLabel(self.left_frame, text="Map Data Columns", font=("Arial", 24, "bold"))
        title_label.grid(row=0, column=0, columnspan=2, pady=(10, 20), padx=20)

        # --- Column Mapping ---
        label_font = ("Arial", 16, "bold")
        entry_font = ("Arial", 16)
        entry_height = 40
        
        field_frame = ctk.CTkFrame(self.left_frame, fg_color="transparent")
        field_frame.grid(row=1, column=0, columnspan=2, sticky="ew", padx=20)
        field_frame.grid_columnconfigure(1, weight=1)

        # 1. Phone Column
        ctk.CTkLabel(field_frame, text="Phone Column:", font=label_font, text_color="green").grid(row=0, column=0, sticky="e", padx=(10,5), pady=10)
        self.phone_col_var = ctk.StringVar()
        self.phone_col_entry = ctk.CTkEntry(
            field_frame, textvariable=self.phone_col_var, 
            font=entry_font, height=entry_height, 
            fg_color=self.inner_color, border_width=2, border_color="green", corner_radius=10,
            placeholder_text="e.g., A or Phone"
        )
        self.phone_col_entry.grid(row=0, column=1, sticky="ew", padx=(5,10), pady=10)
        
        # 2. Name Column
        ctk.CTkLabel(field_frame, text="Name Column:", font=label_font, text_color="green").grid(row=1, column=0, sticky="e", padx=(10,5), pady=10)
        self.name_col_var = ctk.StringVar()
        self.name_col_entry = ctk.CTkEntry(
            field_frame, textvariable=self.name_col_var, 
            font=entry_font, height=entry_height, 
            fg_color=self.inner_color, border_width=2, border_color="green", corner_radius=10,
            placeholder_text="e.g., B or Name (Optional)"
        )
        self.name_col_entry.grid(row=1, column=1, sticky="ew", padx=(5,10), pady=10)

        # 3. Custom 1 Column
        ctk.CTkLabel(field_frame, text="Custom 1 Col:", font=label_font).grid(row=2, column=0, sticky="e", padx=(10,5), pady=10)
        self.custom1_col_var = ctk.StringVar()
        self.custom1_col_entry = ctk.CTkEntry(
            field_frame, textvariable=self.custom1_col_var, 
            font=entry_font, height=entry_height, 
            fg_color=self.inner_color, border_width=0, corner_radius=10,
            placeholder_text="e.g., C or Custom1 (Optional)"
        )
        self.custom1_col_entry.grid(row=2, column=1, sticky="ew", padx=(5,10), pady=10)

        # 4. Custom 2 Column
        ctk.CTkLabel(field_frame, text="Custom 2 Col:", font=label_font).grid(row=3, column=0, sticky="e", padx=(10,5), pady=10)
        self.custom2_col_var = ctk.StringVar()
        self.custom2_col_entry = ctk.CTkEntry(
            field_frame, textvariable=self.custom2_col_var, 
            font=entry_font, height=entry_height, 
            fg_color=self.inner_color, border_width=0, corner_radius=10,
            placeholder_text="e.g., D or Custom2 (Optional)"
        )
        self.custom2_col_entry.grid(row=3, column=1, sticky="ew", padx=(5,10), pady=10)
        
        # --- Advanced Options Toggle ---
        self.advanced_visible = False
        self.toggle_advanced_btn = ctk.CTkButton(
            self.left_frame, 
            text="Advanced Options ▾", 
            command=self.toggle_advanced,
            fg_color="transparent",
            hover_color=self.inner_color,
            text_color="gray"
        )
        self.toggle_advanced_btn.grid(row=5, column=0, columnspan=2, sticky="w", padx=20, pady=(10, 0))
        
        # --- Advanced Options Frame (Initially hidden) ---
        self.advanced_frame = ctk.CTkFrame(self.left_frame, fg_color="transparent")
        self.advanced_frame.grid_columnconfigure(1, weight=1)
        
        ctk.CTkLabel(self.advanced_frame, text="Row Range:", font=label_font).grid(row=0, column=0, sticky="e", padx=(10,5), pady=10)
        self.row_range_var = ctk.StringVar()
        self.row_range_entry = ctk.CTkEntry(
            self.advanced_frame, textvariable=self.row_range_var, 
            font=entry_font, height=entry_height, 
            fg_color=self.inner_color, border_width=0, corner_radius=10,
            placeholder_text="e.g., 2-10, 15, 21-30 (Optional)"
        )
        self.row_range_entry.grid(row=0, column=1, sticky="ew", padx=(5,10), pady=10)
        self.advanced_frame.grid_remove() # Hide initially

        # Spacer
        self.left_frame.grid_rowconfigure(7, weight=1) 
        ctk.CTkFrame(self.left_frame, fg_color="transparent").grid(row=7, column=0, sticky="nsew")

        # --- Import/Cancel Buttons ---
        button_frame = ctk.CTkFrame(self.left_frame, fg_color="transparent")
        button_frame.grid(row=8, column=0, columnspan=2, sticky="ew", padx=20, pady=(10, 10))
        button_frame.grid_columnconfigure((0,1), weight=1)
        
        self.cancel_btn = ctk.CTkButton(
            button_frame,
            text="Cancel",
            command=self.destroy,
            height=45,
            font=("Arial", 18, "bold"),
            fg_color="#555555",
            hover_color="#777777"
        )
        self.cancel_btn.grid(row=0, column=0, sticky="ew", padx=(0, 5))
        
        self.import_btn = ctk.CTkButton(
            button_frame, 
            text="Import Data", 
            command=self.on_import, 
            height=45,
            font=("Arial", 18, "bold"),
            fg_color="green",
            hover_color="#45A049"
        )
        self.import_btn.grid(row=0, column=1, sticky="ew", padx=(5, 0))
        
        # Finally, re-run header detection to populate the variables
        self.load_preview_and_detect_headers()


    def load_preview_and_detect_headers(self):
        """
        Loads the first 200 rows into a grid preview and auto-detects headers.
        Includes cosmetic fix for vertical spacing.
        (Logic remains as provided, no need to change internal data handling)
        """
        header = []
        preview_lines = []
        
        # --- Clean up the preview grid first ---
        for widget in self.preview_grid_frame.winfo_children():
            widget.destroy()
        
        try:
            # --- Data Loading Logic (Unchanged) ---
            if self.csv_data:
                f = StringIO(self.csv_data)
                try: dialect = csv.Sniffer().sniff(f.read(2048))
                except csv.Error: f.seek(0); dialect = 'excel'
                f.seek(0)
                reader = csv.reader(f, dialect)
                
                header = next(reader)
                preview_lines.append(header)
                for i, row in enumerate(reader):
                    if i >= 199: break
                    preview_lines.append(row)
                    
            elif self.file_path.endswith('.csv'):
                with open(self.file_path, newline='', encoding='utf-8-sig') as csvfile:
                    try: dialect = csv.Sniffer().sniff(csvfile.read(2048))
                    except csv.Error: csvfile.seek(0); dialect = 'excel'
                    csvfile.seek(0)
                    reader = csv.reader(csvfile, dialect)
                    header = next(reader)
                    preview_lines.append(header)
                    for i, row in enumerate(reader):
                        if i >= 199: break
                        preview_lines.append(row)
                        
            elif self.file_path.endswith(('.xlsx', '.xls')):
                wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
                sheet = wb.active
                header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
                header = [str(cell) if cell is not None else "" for cell in header_row]
                preview_lines.append(header)
                for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=201, values_only=True)):
                    preview_lines.append([str(cell) if cell is not None else "" for cell in row])

            # --- Populate Preview Grid ---
            if preview_lines:
                mode = ctk.get_appearance_mode().lower()
                border_color = "#444444" if mode == "dark" else "#CCCCCC"
                
                header_font = ctk.CTkFont(family="Arial", size=11, weight="bold") 
                data_font = ctk.CTkFont(family="Arial", size=11)
                
                max_cols = max(len(row) for row in preview_lines)
                
                # Create a placeholder frame for the top-left corner
                corner_frame = ctk.CTkFrame(self.preview_grid_frame, corner_radius=0, fg_color="transparent", border_width=1, border_color=border_color)
                corner_frame.grid(row=0, column=0, sticky="nsew")
                
                # 1. Column Letter Header (A, B, C...)
                for col_idx in range(max_cols):
                    col_letter = chr(65 + col_idx)
                    cell_frame = ctk.CTkFrame(self.preview_grid_frame, corner_radius=0, fg_color="transparent", border_width=1, border_color=border_color)
                    cell_frame.grid(row=0, column=col_idx + 1, sticky="nsew")
                    
                    # --- FIX: Reduced vertical padding to minimal (pady=0) ---
                    cell_label = ctk.CTkLabel(cell_frame, text=col_letter, font=header_font, text_color="green", padx=5, pady=0)
                    cell_label.pack(expand=True, fill="both") # Use pack to tightly fit content

                # 2. Data Rows (with row numbers)
                for row_idx, row in enumerate(preview_lines):
                    grid_row = row_idx + 1 
                    
                    # Row Number Cell
                    row_num_str = str(row_idx + 1)
                    row_num_frame = ctk.CTkFrame(self.preview_grid_frame, corner_radius=0, fg_color="transparent", border_width=1, border_color=border_color)
                    row_num_frame.grid(row=grid_row, column=0, sticky="nsew")
                    
                    # --- FIX: Reduced vertical padding to minimal (pady=0) ---
                    row_num_label = ctk.CTkLabel(row_num_frame, text=row_num_str, font=header_font, padx=5, pady=0)
                    row_num_label.pack(expand=True, fill="both") 
                    
                    # Data Cells
                    for col_idx in range(max_cols):
                        grid_col = col_idx + 1
                        cell_data = str(row[col_idx]) if col_idx < len(row) else ""
                        
                        cell_frame = ctk.CTkFrame(self.preview_grid_frame, corner_radius=0, fg_color="transparent", border_width=1, border_color=border_color)
                        cell_frame.grid(row=grid_row, column=grid_col, sticky="nsew")
                        
                        font_to_use = header_font if row_idx == 0 else data_font
                        
                        # Data cells should also have minimal vertical padding (pady=1)
                        cell_label = ctk.CTkLabel(cell_frame, text=cell_data, font=font_to_use, padx=5, pady=1, anchor="w")
                        cell_label.pack(expand=True, fill="x", padx=0, pady=0)
                        
                        self.preview_grid_frame.grid_columnconfigure(grid_col, weight=1)

            # --- Auto-detect Headers (Logic Unchanged, but only runs if controls exist) ---
            if hasattr(self, 'phone_col_entry'):
                placeholder_name = "Enter Column Name (e.g., Phone)"
                placeholder_letter = "Enter Column Letter (e.g., A)"
                
                if self.is_csv:
                    self.phone_col_entry.configure(placeholder_text=placeholder_name)
                    self.name_col_entry.configure(placeholder_text="e.g., Name (Optional)")
                    self.custom1_col_entry.configure(placeholder_text="e.g., Custom1 (Optional)")
                    self.custom2_col_entry.configure(placeholder_text="e.g., Custom2 (Optional)")
                    
                    for col_name in header:
                        if not col_name: continue
                        col_lower = col_name.lower()
                        if "phone" in col_lower or "mobile" in col_lower or "number" in col_lower:
                            if self.phone_col_var.get() == "": self.phone_col_var.set(col_name)
                        elif "name" in col_lower:
                            if self.name_col_var.get() == "": self.name_col_var.set(col_name)
                        elif "custom1" in col_lower or "var1" in col_lower:
                            if self.custom1_col_var.get() == "": self.custom1_col_var.set(col_name)
                        elif "custom2" in col_lower or "var2" in col_lower:
                            if self.custom2_col_var.get() == "": self.custom2_col_var.set(col_name)
                else:
                    self.phone_col_entry.configure(placeholder_text=placeholder_letter)
                    self.name_col_entry.configure(placeholder_text="e.g., B (Optional)")
                    self.custom1_col_entry.configure(placeholder_text="e.g., C (Optional)")
                    self.custom2_col_entry.configure(placeholder_text="e.g., D (Optional)")
                    
                    # Check for existing values before setting defaults
                    if not self.phone_col_var.get(): self.phone_col_var.set("A")
                    if not self.name_col_var.get(): self.name_col_var.set("B")
                    
                    for i, col_name in enumerate(header):
                        col_letter = chr(65 + i)
                        col_lower = str(col_name).lower()
                        # Only override if the field is currently empty
                        if "phone" in col_lower or "mobile" in col_lower or "number" in col_lower:
                            if not self.phone_col_var.get(): self.phone_col_var.set(col_letter)
                        elif "name" in col_lower:
                            if not self.name_col_var.get(): self.name_col_var.set(col_letter)


        except Exception as e:
            # --- Use ThemedPopup and clean up the preview grid to show the error message ---
            ThemedPopup(self, title="Preview Error", message=f"Failed to load data preview: {e}", icon_type="error")
            
            for widget in self.preview_grid_frame.winfo_children():
                widget.destroy()
            
            error_label = ctk.CTkLabel(self.preview_grid_frame, text=f"Error loading preview: {e}", text_color="red", wraplength=400)
            error_label.pack(padx=10, pady=10)
 
    def on_import(self):
        """Gathers all settings and passes them to the main app's callback."""
        phone_col = self.phone_col_var.get().strip()
        if not phone_col:
            # --- FIX: Replaced messagebox with ThemedPopup ---
            ThemedPopup(self, title="Error", message="Phone column is required.", icon_type="error")
            return
            
        name_col = self.name_col_var.get().strip()
        custom1_col = self.custom1_col_var.get().strip()
        custom2_col = self.custom2_col_var.get().strip()
        row_range_str = self.row_range_var.get().strip()
        
        # Call the main app's load function with all new parameters
        self.on_import_callback(
            file_path=self.file_path, 
            phone_col=phone_col, 
            name_col=name_col, 
            custom1_col=custom1_col, 
            custom2_col=custom2_col,
            merge_mode=self.merge_mode,
            row_range_str=row_range_str,
            csv_data=self.csv_data # Pass the in-memory data
        )
        self.destroy()
# ----------------------- (NEW) MAPPING POPUP (PHASE 2) -----------------------

class FirstRunPopup(ctk.CTkToplevel):
    def __init__(self, master, on_close_callback, *args, **kwargs):
        super().__init__(master, *args, **kwargs)
        self.validated = False  # Flag to track successful validation
        self.result_queue = queue.Queue() # Queue stores (is_valid, error_msg, is_used, key)
        self.title("Enter the Product Key")
        self.geometry("600x300")
        self.resizable(False, False)
        self.attributes("-topmost", True)
        
        mode = ctk.get_appearance_mode().lower()
        bg_color = DARK_BG if mode == "dark" else LIGHT_BG
        inner_color = DARK_INNER if mode == "dark" else LIGHT_INNER
        
        self.configure(fg_color=bg_color) # Theme
        self.on_close_callback = on_close_callback
        self.protocol("WM_DELETE_WINDOW", self.terminate)
        self.bind("<Escape>", lambda e: self.terminate())
        center_window(self)
        self.fade_in()
        self.container = ctk.CTkFrame(self, fg_color="transparent", width=600, height=300, corner_radius=0) # Theme
        self.container.pack(fill="both", expand=True)
        # Title
        title = ctk.CTkLabel(self.container, text="Enter the Product Key", font=("Arial", 24, "bold"),
                             fg_color="transparent")
        title.pack(anchor="w", pady=(40, 20), padx=40)
        # Entry
        self.product_key_entry = ctk.CTkEntry(self.container, width=320, height=50,
                                              placeholder_text="K39U-33W3",
                                              font=("Arial", 16), corner_radius=15,
                                              fg_color=inner_color, # Theme
                                              border_width=0) # Theme
        self.product_key_entry.pack(anchor="w", pady=10, padx=40)
        self.product_key_entry.bind("<FocusIn>", self.clear_placeholder)
        self.product_key_entry.bind("<Return>", lambda e: self.register_product_key())
        # Buttons Row - Left aligned
        buttons_frame = ctk.CTkFrame(self.container, fg_color="transparent", corner_radius=0)
        buttons_frame.pack(anchor="w", pady=30, padx=40)
        self.get_key_button = AnimatedCTkButton(buttons_frame, text="Get Product Key",
                                                width=140, height=50, font=("Arial", 20, "bold"),
                                                fg_color="#4CAF50", hover_color="#45A049", corner_radius=15,
                                                command=self.get_product_key)
        self.get_key_button.pack(side="left", padx=(0, 20))
        self.register_button = AnimatedCTkButton(buttons_frame, text="Register",
                                                 width=140, height=50, font=("Arial", 20, "bold"),
                                                 fg_color="#4CAF50", hover_color="#45A049", corner_radius=15,
                                                 command=self.register_product_key)
        self.register_button.pack(side="left")
        # Image - Right side
        self.product_img = round_corners(PRODUCT_IMAGE_PATH, (210, 210), corner_radius=15)
        if self.product_img:
            self.product_image_label = ctk.CTkLabel(self.container, image=self.product_img, text="")
            self.product_image_label.place(relx=0.85, rely=0.5, anchor="center")
            self.product_image_label.bind("<Enter>", self.animate_image_zoom_in)
            self.product_image_label.bind("<Leave>", self.animate_image_zoom_out)
        else:
            # Fallback label
            fallback_label = ctk.CTkLabel(self.container, text="Image Missing", text_color="#FFFFFF")
            fallback_label.place(relx=0.85, rely=0.5, anchor="center")
            

    def check_validation_result(self):
        """Checks the result queue from the validation thread."""
        try:
            # Check the queue without blocking
            # Queue now contains: (is_valid, error_msg, is_used, key)
            is_valid, error_msg, is_used, key = self.result_queue.get_nowait()

            # --- Result found ---
            if self.winfo_exists(): # Check if popup is still open
                if is_valid:
                    self.validated = True
                    self.show_introduction_video() # Proceed if valid
                else:
                    # --- Display Error Message synchronously ---
                    if error_msg:
                        # AlertPopup runs synchronously here because the main thread
                        # is blocked by .wait_window() and is processing the event loop
                        # via after() calls.
                        AlertPopup(self, error_msg, is_used=is_used, key=key)
                    
                    # Re-enable UI 
                    self.register_button.configure(state="normal", text="Register")
                    self.get_key_button.configure(state="normal")
                    self.product_key_entry.configure(state="normal")
                    self.product_key_entry.focus_set()
            else:
                print("DEBUG: Validation result received, but popup was closed.")

        except queue.Empty:
            # --- No result yet, check again soon ---
            if self.winfo_exists(): # Keep polling only if popup is open
                self.after(100, self.check_validation_result) # Check again after 100 milliseconds
        except Exception as e:
            # Handle unexpected errors during queue check/UI update
            print(f"ERROR checking validation result: {e}")
            if self.winfo_exists():
                # Display generic error and try to re-enable UI
                AlertPopup(self, f"An unexpected error occurred during validation: {e}")
                self.register_button.configure(state="normal", text="Register")
                self.get_key_button.configure(state="normal")
                self.product_key_entry.configure(state="normal")


    def clear_placeholder(self, event):
        if self.product_key_entry.get() == "Sample Product Key":
            self.product_key_entry.delete(0, "end")
            
    def animate_image_zoom_in(self, event):
        if hasattr(self, 'product_image_label') and self.product_img:
            zoomed_img = round_corners(PRODUCT_IMAGE_PATH, (240, 240), corner_radius=15)
            if zoomed_img:
                self.after(50, lambda: self.product_image_label.configure(image=zoomed_img))
                
    def animate_image_zoom_out(self, event):
        if hasattr(self, 'product_image_label') and self.product_img:
            self.after(50, lambda: self.product_image_label.configure(image=self.product_img))
            
    def fade_in(self, alpha=0.0):
        if alpha < 1.0:
            self.attributes("-alpha", alpha)
            self.after(50, lambda: self.fade_in(alpha + 0.1))
        else:
            self.attributes("-alpha", 1.0)
            
    def terminate(self):
        if not self.validated:  # If not validated, terminate the entire app
            self.master.quit()
            sys.exit(0)
        self.destroy()
        
    def get_product_key(self):
        message = "Hi! I would like to purchase WabulkXpress. What's the cost and how do I get a product key? 😊💸🚀"
        encoded = message.replace(" ", "%20")
        webbrowser.open(f"https://wa.me/{WHATSAPP_NUMBER}?text={encoded}")

    def write_license_file(self, key, expiry_date_str):
        """Helper function to write the license.dat file."""
        try:
            license_data = {"key": key, "expiry_date": expiry_date_str}
            with open(LICENSE_FILE, 'w') as f:
                json.dump(license_data, f, indent=2)
            logger.info(f"Successfully wrote license file. Key: {key}, Expiry: {expiry_date_str}")
            return True
        except Exception as e:
            logger.error(f"Failed to write local license file: {e}")
            # NOTE: We can't call AlertPopup here, it must be returned via queue.
            return False

    # WabulkXpress.py: Line ~1758 (within FirstRunPopup.validate_product_key)

    def validate_product_key(self, key):
        """
        Validates key against the new GAS logic. 
        Returns (bool is_valid, str/None error_message, bool is_used_or_expired, str/None key)
        """
        
        # 1. Re-check Format for general safety, but use the detailed error message
        if not key or not re.match(r'^[A-Z0-9]{4}-[A-Z0-9]{4}$', key):
             return False, "The product key format is incorrect. It must be XXXX-YYYY.", False, key
        
        encoded_key = urllib.parse.quote(key, safe='')
        
        try: # <--- PRIMARY TRY BLOCK STARTS HERE
            SCRIPT_ID = "AKfycbyXcXaweoMSFLjbUW7qILp7tqKeSAa5ZoJOGdZRWiNxp5s86zL401nUZNCRKdaz0qjm"
            url = f"https://script.google.com/macros/s/{SCRIPT_ID}/exec?key={encoded_key}"
            logger.info(f"Validating key '{key}' at URL: {url}")
            
            response = requests.get(url, timeout=30)
            response.raise_for_status()
            data = response.json()
            status = data.get("status")
            
            logger.info(f"API response: {data}")
            
            # --- VALIDATION LOGIC ---
            if status == "valid":
                used = data.get("used", 1) 
                
                if used == 0:
                    # --- Scenario A: First-time activation (Same as before) ---
                    validity_months = data.get("validity_months")
                    if not validity_months:
                        logger.error("Server Error: 'validity_months' missing.")
                        return False, "Server Error: Key is valid but validity period is missing. Contact support.", False, key
                    
                    try:
                        days_valid = int(float(validity_months) * 30.44) 
                        expiry_date_str = (datetime.now().date() + timedelta(days=days_valid)).strftime("%Y-%m-%d")
                        
                        if not self.write_license_file(key, expiry_date_str):
                            return False, "Error saving license data locally.", False, key 
                        
                    except Exception as e:
                        logger.error(f"Failed to calculate expiry or write file: {e}")
                        return False, f"Error saving license: {e}", False, key

                    # Mark the key as used in the Google Sheet
                    use_url = f"{url}&action=use"
                    use_response = requests.get(use_url, timeout=30)
                    use_data = use_response.json()
                    
                    if use_data.get("status") == "success":
                        logger.info(f"Key '{key}' successfully activated and marked as used.")
                        return True, None, False, key
                    else:
                        logger.warning(f"Failed to mark key as used: {use_data}")
                        try: os.remove(LICENSE_FILE)
                        except: pass
                        return False, f"Server error: {use_data.get('message', 'Could not mark key as used.')}", False, key

                elif used == 1:
                    # --- Scenario B: Key already active (NOW BLOCKED REGARDLESS OF EXPIRY) ---
                    expiry_date_str = data.get("expiry_date", "N/A")
                    
                    # Check if expired just to provide better messaging
                    is_expired = False
                    if expiry_date_str != "N/A":
                        try:
                            expiry_date = datetime.strptime(expiry_date_str, "%Y-%m-%d").date()
                            if datetime.now().date() > expiry_date:
                                is_expired = True
                        except ValueError:
                            pass # Keep message generic if date format is bad

                    message = (
                        f"This product key ({key}) has already been used and activated."
                        f"\nIf this is a new device, please use the 'Get Help' button below to request a key revocation."
                        + (" (Expired on " + expiry_date_str + ")" if is_expired else "")
                    )
                    
                    # Return FAILURE, but set is_used=True to show the Get Help button
                    logger.warning(f"Used key {key} detected. Blocking load.")
                    return False, message, True, key
            
            elif status == "expired":
                message = data.get("message", "This product key has expired.")
                return False, message, True, key
            
            else: # status == "invalid" or any other (e.g., 'Key not found')
                message = data.get("message", "The product key is invalid or not found.")
                return False, f"Invalid Key: {message}", False, key

        # --- EXCEPTION HANDLING (RETURNS ERRORS) ---
        except requests.exceptions.Timeout:
            logger.error("Validation request timed out.")
            return False, "Request timed out. Check internet or try again.", False, key
        except requests.exceptions.RequestException as e:
            logger.error(f"Network/HTTP error: {e}")
            return False, "Network Error: Check connection and try again.", False, key
        except json.JSONDecodeError:
            error_snippet = response.text[:200] if 'response' in locals() else "No response"
            logger.error(f"Invalid JSON. Snippet: {error_snippet}")
            return False, "Server returned invalid data. Contact support.", False, key
        except Exception as e:
            logger.error(f"Unexpected error: {e}", exc_info=True)
            return False, f"An unexpected error occurred: {e}", False, key
        
    def register_product_key(self, event=None):
        key = self.product_key_entry.get().strip()
        
        # --- SYNCHRONOUS PRE-CHECK (FIX) ---
        if not key or key == "Sample Product Key":
            ThemedPopup(self, title="Input Error", message="Please enter a valid product key.", icon_type="error")
            return
            
        if not re.match(r'^[A-Z0-9]{4}-[A-Z0-9]{4}$', key):
            # Synchronous error display for invalid format
            AlertPopup(self, "The product key format is incorrect. It must be XXXX-YYYY.", is_used=False, key=key)
            return
        # --- END SYNCHRONOUS PRE-CHECK ---

        # --- Disable UI elements ---
        self.register_button.configure(state="disabled", text="Validating...")
        self.get_key_button.configure(state="disabled")
        self.product_key_entry.configure(state="disabled")

        # --- Define the background task ---
        def validation_thread_task(product_key, result_q):
            # This function returns the full result tuple, including error message
            # The synchronous nature of the main thread processing the button click 
            # and then calling self.after(100, ...) allows this thread to run 
            # without crashing on API interaction failures.
            is_valid, error_msg, is_used, key = self.validate_product_key(product_key)
            result_q.put((is_valid, error_msg, is_used, key))

        # --- Start the background thread ---
        threading.Thread(
            target=validation_thread_task,
            args=(key, self.result_queue), 
            daemon=True
        ).start()

        # --- Start polling the queue on the main thread ---
        self.after(100, self.check_validation_result)

    def show_introduction_video(self):
        video_window = ctk.CTkToplevel(self)
        video_window.title("Welcome!")
        video_window.iconbitmap(TITLE_ICON_PATH)
        video_window.geometry("700x400")
        video_window.resizable(False, False)
        video_window.attributes("-topmost", True)
        video_window.transient(self)
        
        mode = ctk.get_appearance_mode().lower()
        bg_color = DARK_BG if mode == "dark" else LIGHT_BG
        
        video_window.configure(fg_color=bg_color)
        video_window.protocol("WM_DELETE_WINDOW", video_window.destroy)
        video_window.bind("<Escape>", lambda e: video_window.destroy())
        center_window(video_window)
        
        video_window.container = ctk.CTkFrame(video_window, width=700, height=400, fg_color="transparent")
        video_window.container.pack(fill="both", expand=True)
        
        if os.path.exists(VIDEO_PATH):
            video_window.video_player = TkinterVideo(video_window.container, VIDEO_PATH, scaled=True, keep_aspect=True)
            video_window.video_player.place(relx=0, rely=0, relwidth=1, relheight=1)
            video_window.video_player.play()
        else:
            video_window.instruction_label = ctk.CTkLabel(video_window.container, text="Video file not found", font=("Arial", 16))
            video_window.instruction_label.place(relx=0.5, rely=0.5, anchor="center")
            
        video_window.bottom_frame = ctk.CTkFrame(video_window.container, fg_color="transparent")
        video_window.bottom_frame.place(relx=0.5, rely=0.95, anchor="center")
        
        video_window.dont_show_var = ctk.BooleanVar(value=False)
        video_window.checkbox = ctk.CTkCheckBox(video_window.bottom_frame, text="Don't show this again", 
                                                variable=video_window.dont_show_var, font=("Arial", 14, "bold"))
        video_window.checkbox.pack(side="left", padx=20, pady=10)
        
        video_window.ok_button = ctk.CTkButton(video_window.bottom_frame, text="OK", fg_color="#29AE07", 
                                                hover_color="#009e05", font=("Arial", 16, "bold"), width=100, height=50,
                                                corner_radius=15, command=lambda: self.close_video_popup(video_window))
        video_window.ok_button.pack(side="left", padx=20, pady=10)
        video_window.protocol("WM_DELETE_WINDOW", lambda: self.close_video_popup(video_window))
        
    def close_video_popup(self, video_window):
        if hasattr(video_window, 'video_player'):
            try: video_window.video_player.stop()
            except: pass
        self.on_video_complete(video_window)
        
    def on_video_complete(self, video_window=None):
        if video_window:
            video_window.destroy()
        self.on_close_callback()
        self.destroy()

class ExcelTable(ctk.CTkScrollableFrame):
    def __init__(self, master, main_app, *args, **kwargs):
        super().__init__(master, *args, **kwargs)
        self.main_app = main_app
        # Remove scrollbar_fg_color to enable auto-hiding
        # Set button colors for the modern, rounded knob
        self.configure(
            corner_radius=15, 
            scrollbar_button_color="#555555",       # Knob color for dark mode
            scrollbar_button_hover_color="#333333"  # Knob hover color for dark mode
        )
        
        self.rows = []
        self.custom_cols_active = [] # Tracks which custom columns are shown
        self.undo_stack = [[]] # Start with an empty list as the initial state
        self.redo_stack = []
        self.header_frame = ctk.CTkFrame(self, corner_radius=10, fg_color="transparent")
        self.header_frame.pack(fill="x", padx=5, pady=3, ipady=5) # Added padding
        
        self.add_header() # Initial header
        self.prepopulate_rows(1) # Initial row

    def add_header(self):
        # Clear existing header
        for widget in self.header_frame.winfo_children():
            widget.destroy()

        # Define column widths
        col_widths = {
            "sno": 40,
            "phone": 150,
            "name": 150,
            "custom1": 100,
            "custom2": 100,
            "status": 60
        }
        
        # Calculate dynamic width for flexible columns (phone, name, custom)
        num_flex_cols = 2 + len(self.custom_cols_active)
        # Give status and sno fixed width, divide rest
        
        header_font = ("Arial", 16, "bold") # Bigger font

        ctk.CTkLabel(self.header_frame, text="S.No.", width=col_widths["sno"], anchor="center", font=header_font).pack(side="left", padx=5)
        
        phone_label = ctk.CTkLabel(self.header_frame, text="Phone Number", anchor="center", font=header_font)
        phone_label.pack(side="left", padx=5, fill="x", expand=True)
        
        name_label = ctk.CTkLabel(self.header_frame, text="Name", anchor="center", font=header_font)
        name_label.pack(side="left", padx=5, fill="x", expand=True)

        if "custom1" in self.custom_cols_active:
            c1_label = ctk.CTkLabel(self.header_frame, text="Custom 1", anchor="center", font=header_font)
            c1_label.pack(side="left", padx=5, fill="x", expand=True)

        if "custom2" in self.custom_cols_active:
            c2_label = ctk.CTkLabel(self.header_frame, text="Custom 2", anchor="center", font=header_font)
            c2_label.pack(side="left", padx=5, fill="x", expand=True)
            
        ctk.CTkLabel(self.header_frame, text="Status", width=col_widths["status"], anchor="center", font=header_font).pack(side="left", padx=5)

    def prepopulate_rows(self, count):
        for _ in range(count):
            self.add_row()

    def add_row(self, data=None):
        if data is None:
            data = {}
            
        row_frame = ctk.CTkFrame(self, corner_radius=10, fg_color="transparent")
        row_frame.pack(fill="x", padx=5, pady=3)
        
        sno_label = ctk.CTkLabel(row_frame, text=str(len(self.rows)+1), width=40, anchor="center")
        sno_label.pack(side="left", padx=5)
        
        mode = ctk.get_appearance_mode().lower()
        cell_color = DARK_INNER if mode == "dark" else LIGHT_INNER
        
        row_dict = {
            "sno": sno_label,
            "indicator_state": 0, 
            "row_frame": row_frame
        }

        # --- Phone Entry ---
        phone_var = ctk.StringVar(value=data.get("phone", ""))
        phone_entry = ctk.CTkEntry(row_frame, textvariable=phone_var, placeholder_text="Enter number", corner_radius=10, fg_color=cell_color, border_width=0)
        phone_entry.pack(side="left", padx=5, fill="x", expand=True)
        phone_entry.bind("<Return>", lambda event, widget=phone_entry, var=phone_var: self.validate_phone(widget, var))
        phone_entry.bind("<KeyRelease>", self.check_add_row)
        phone_entry.bind("<Control-v>", self.handle_paste)
        phone_entry.bind("<Control-z>", self.undo) # <-- MODIFIED
        phone_entry.bind("<Control-y>", self.redo) # <-- MODIFIED
        row_dict["phone"] = phone_entry
        row_dict["phone_var"] = phone_var

        # --- Name Entry ---
        name_var = ctk.StringVar(value=data.get("name", ""))
        name_entry = ctk.CTkEntry(row_frame, textvariable=name_var, placeholder_text="Enter name", corner_radius=10, fg_color=cell_color, border_width=0)
        name_entry.pack(side="left", padx=5, fill="x", expand=True)
        name_entry.bind("<Return>", lambda event, var=name_var: self.validate_name(var))
        name_entry.bind("<KeyRelease>", self.check_add_row)
        name_entry.bind("<Control-v>", self.handle_paste)
        name_entry.bind("<Control-z>", self.undo) # <-- MODIFIED
        name_entry.bind("<Control-y>", self.redo) # <-- MODIFIED
        row_dict["name"] = name_entry
        row_dict["name_var"] = name_var

        # --- Custom 1 Entry (Dynamic) ---
        if "custom1" in self.custom_cols_active:
            custom1_var = ctk.StringVar(value=data.get("custom1", ""))
            custom1_entry = ctk.CTkEntry(row_frame, textvariable=custom1_var, placeholder_text="Custom 1", corner_radius=10, fg_color=cell_color, border_width=0)
            custom1_entry.pack(side="left", padx=5, fill="x", expand=True)
            custom1_entry.bind("<KeyRelease>", self.check_add_row)
            custom1_entry.bind("<Control-v>", self.handle_paste)
            custom1_entry.bind("<Control-z>", self.undo) # <-- MODIFIED
            custom1_entry.bind("<Control-y>", self.redo) # <-- MODIFIED
            row_dict["custom1"] = custom1_entry
            row_dict["custom1_var"] = custom1_var
            
        # --- Custom 2 Entry (Dynamic) ---
        if "custom2" in self.custom_cols_active:
            custom2_var = ctk.StringVar(value=data.get("custom2", ""))
            custom2_entry = ctk.CTkEntry(row_frame, textvariable=custom2_var, placeholder_text="Custom 2", corner_radius=10, fg_color=cell_color, border_width=0)
            custom2_entry.pack(side="left", padx=5, fill="x", expand=True)
            custom2_entry.bind("<KeyRelease>", self.check_add_row)
            custom2_entry.bind("<Control-v>", self.handle_paste)
            custom2_entry.bind("<Control-z>", self.undo) # <-- MODIFIED
            custom2_entry.bind("<Control-y>", self.redo) # <-- MODIFIED
            row_dict["custom2"] = custom2_entry
            row_dict["custom2_var"] = custom2_var

        # --- Status Indicator ---
        indicator_bg = DARK_BG if mode == "dark" else LIGHT_BG # Match main BG for "grid" look
        indicator = tk.Canvas(row_frame, width=30, height=30, highlightthickness=0, bg=indicator_bg)
        indicator.create_oval(5, 5, 28, 28, fill="green", outline="")
        indicator.create_text(16, 17, text="✔", fill="white", font=("Arial", 12, "bold"))
        indicator.pack(side="left", padx=5)
        row_dict["indicator"] = indicator
        
        indicator.bind("<Button-1>", lambda e, r=row_dict: self.toggle_indicator(r))
        self.rows.append(row_dict)
    def update_row_numbers(self):
        for idx, row in enumerate(self.rows, start=1):
            row["sno"].configure(text=str(idx))

    def toggle_indicator(self, row_dict):
        self.save_state_for_undo() # Save state before the toggle/delete
        
        state = row_dict.get("indicator_state", 0)
        indicator = row_dict.get("indicator")
        if state == 0:
            indicator.delete("all")
            indicator.create_oval(5, 5, 28, 28, fill="red", outline="")
            indicator.create_text(16, 17, text="✖", fill="white", font=("Arial", 12, "bold"))
            row_dict["indicator_state"] = 1
            row_dict["skip"] = True
        elif state == 1:
            row_dict["row_frame"].destroy()
            self.rows.remove(row_dict)
            if not self.rows:
                self.add_row()
            self.update_row_numbers()

    def validate_phone(self, widget, var):
        # (This function is unchanged)
        text = var.get().strip()
        if not text:
            return
        original_text = text
        default_country = self.main_app.country_code_var.get()
        allowed_codes = ["+91", "+1", "+44", "+61", "+81", "+49", "+33", "+86", "+7"]
        text = text.replace(" ", "").replace("-", "")
        if text.startswith("+"):
            clean = re.sub(r"[^\d+]", "", text)
            digits_only = re.sub(r"\D", "", clean)
            if len(digits_only) < 10:
                self.main_app.log_live(f"⚠️ Invalid phone number detected: {original_text}")
            var.set(clean)
            widget.delete(0, "end")
            widget.insert(0, clean)
            return
        text = re.sub(r"^[^\d]+", "", text)
        matched = False
        for code in allowed_codes:
            code_digits = code.replace("+", "")
            if text.startswith(code_digits):
                text = "+" + text
                matched = True
                break
        if not matched:
            if default_country != "None":
                text = default_country + text
        final = re.sub(r"[^\d+]", "", text)
        digits_only = re.sub(r"\D", "", final)
        if len(digits_only) < 10 or not digits_only.isdigit():
            self.main_app.log_live(f"⚠️ Invalid phone number detected: {original_text}")
        var.set(final)
        widget.delete(0, "end")
        widget.insert(0, final)


    def validate_name(self, var):
        var.set(var.get().strip())

    def check_add_row(self, event):
        last_row = self.rows[-1]
        
        # Check if any field in the last row has text
        is_filled = False
        if last_row["phone_var"].get().strip() or last_row["name_var"].get().strip():
            is_filled = True
        
        if "custom1_var" in last_row and last_row["custom1_var"].get().strip():
            is_filled = True
            
        if "custom2_var" in last_row and last_row["custom2_var"].get().strip():
            is_filled = True

        if is_filled:
            self.add_row()

    def handle_paste(self, event=None):
        """Handles a Ctrl+V event to paste data from the clipboard."""
        try:
            # 1. Get data from clipboard
            clipboard_data = self.clipboard_get()
        except tk.TclError:
            self.main_app.log_live("📋 Nothing to paste.")
            return "break" # Stop default paste action

        if not clipboard_data:
            return "break" # Stop default paste action

        new_data_entries = []
        # 2. Split into rows
        rows = clipboard_data.strip().split('\n')
        
        self.main_app.log_live(f"📋 Parsing {len(rows)} rows from clipboard...")

        for row_str in rows:
            if not row_str.strip():
                continue # Skip empty lines
                
            # 3. Split each row by tabs (Excel format)
            columns = row_str.split('\t')
            entry = {}
            
            try:
                # Map columns by position:
                # Col 0 -> Phone
                # Col 1 -> Name
                # Col 2 -> Custom1
                # Col 3 -> Custom2
                entry["phone"] = columns[0].strip() if len(columns) > 0 else ""
                entry["name"] = columns[1].strip() if len(columns) > 1 else ""
                entry["custom1"] = columns[2].strip() if len(columns) > 2 else ""
                entry["custom2"] = columns[3].strip() if len(columns) > 3 else ""
            
            except IndexError:
                pass # Gracefully handle rows with fewer columns
            
            # Only add the entry if a phone number is present
            if entry.get("phone"):
                new_data_entries.append(entry)

        if not new_data_entries:
            self.main_app.log_live("📋 Paste complete. No valid new rows found.")
            return "break" # Stop default paste

        # 4. Get existing data from the table
        current_data = self.get_data()
        
        # 5. Merge old and new data
        combined_data = current_data + new_data_entries
        
        # 6. Reload the table with the combined data
        self.load_data(combined_data)
        
        self.main_app.log_live(f"📋 Pasted and merged {len(new_data_entries)} new contacts.")

        # 7. IMPORTANT: Stop the default paste action in the entry box
        return "break"
    
    def save_state_for_undo(self):
        """Saves the current data state for undo."""
        current_data = self.get_data()
        if not self.undo_stack or current_data != self.undo_stack[-1]:
            self.undo_stack.append(current_data)
            self.redo_stack.clear()

    def undo(self, event=None):
        """Loads the previous state from the undo stack."""
        if len(self.undo_stack) > 1: # Can't undo the initial empty state
            self.redo_stack.append(self.undo_stack.pop())
            data_to_load = self.undo_stack[-1]
            self._load_data_internal(data_to_load) # Use private loader
        return "break" # Stop default widget undo

    def redo(self, event=None):
        """Loads the next state from the redo stack."""
        if self.redo_stack:
            data_to_load = self.redo_stack.pop()
            self.undo_stack.append(data_to_load)
            self._load_data_internal(data_to_load) # Use private loader
        return "break" # Stop default widget undo
    
    def _load_data_internal(self, data):
        """
        Internal version of load_data that doesn't affect
        the undo/redo stacks.
        """
        # 1. Clear all existing widgets
        for child in self.winfo_children():
            if child != self.header_frame: # Keep header frame
                child.destroy()
        self.rows = []

        # 2. Determine active custom columns from the new data
        self.custom_cols_active = []
        if data: # Check if data is not empty
            if any(d.get("custom1") for d in data):
                self.custom_cols_active.append("custom1")
            if any(d.get("custom2") for d in data):
                self.custom_cols_active.append("custom2")
        
        # 3. Re-draw the header dynamically
        self.add_header()

        # 4. Load the new data into rows
        for entry in data:
            self.add_row(data=entry) # Pass data to add_row

        # 5. Add one final empty row
        self.add_row()
        
        # 6. Update S.No.
        self.update_row_numbers()

    def handle_paste(self, event=None):
        """Handles a Ctrl+V event to paste data from the clipboard."""
        try:
            # 1. Get data from clipboard
            clipboard_data = self.clipboard_get()
        except tk.TclError:
            self.main_app.log_live("📋 Nothing to paste.")
            return "break" # Stop default paste action

        if not clipboard_data:
            return "break" # Stop default paste action

        new_data_entries = []
        # 2. Split into rows
        rows = clipboard_data.strip().split('\n')
        
        self.main_app.log_live(f"📋 Parsing {len(rows)} rows from clipboard...")

        for row_str in rows:
            if not row_str.strip():
                continue # Skip empty lines
                
            # 3. Split each row by tabs (Excel format)
            columns = row_str.split('\t')
            entry = {}
            
            try:
                # Map columns by position:
                # Col 0 -> Phone
                # Col 1 -> Name
                # Col 2 -> Custom1
                # Col 3 -> Custom2
                entry["phone"] = columns[0].strip() if len(columns) > 0 else ""
                entry["name"] = columns[1].strip() if len(columns) > 1 else ""
                entry["custom1"] = columns[2].strip() if len(columns) > 2 else ""
                entry["custom2"] = columns[3].strip() if len(columns) > 3 else ""
            
            except IndexError:
                pass # Gracefully handle rows with fewer columns
            
            # Only add the entry if a phone number is present
            if entry.get("phone"):
                new_data_entries.append(entry)

        if not new_data_entries:
            self.main_app.log_live("📋 Paste complete. No valid new rows found.")
            return "break" # Stop default paste

        # 4. Get existing data from the table
        current_data = self.get_data()
        
        # 5. Merge old and new data
        combined_data = current_data + new_data_entries
        
        # 6. Reload the table with the combined data
        self.load_data(combined_data) # This will now save the undo state
        
        self.main_app.log_live(f"📋 Pasted and merged {len(new_data_entries)} new contacts.")

        # 7. IMPORTANT: Stop the default paste action in the entry box
        return "break"

    def load_data(self, data):
        self.save_state_for_undo() # Save the *previous* state
        self._load_data_internal(data) # Load the new state

    def get_data(self):
        data = []
        for row in self.rows:
            phone = row["phone_var"].get().strip()
            name = row["name_var"].get().strip()
            
            # Check custom fields only if they exist in the row_dict
            custom1 = ""
            if "custom1_var" in row:
                custom1 = row["custom1_var"].get().strip()
                
            custom2 = ""
            if "custom2_var" in row:
                custom2 = row["custom2_var"].get().strip()

            # Skip if all fields are empty
            if not phone and not name and not custom1 and not custom2:
                continue
                
            entry = {"phone": phone, "name": name}
            if "custom1" in self.custom_cols_active:
                entry["custom1"] = custom1
            if "custom2" in self.custom_cols_active:
                entry["custom2"] = custom2
                
            if "skip" in row:
                entry["skip"] = True
            data.append(entry)
        return data

class CustomImageWindow(ctk.CTkToplevel):
    def __init__(self, master, excel_data, *args, **kwargs):
        super().__init__(master, *args, **kwargs)
        self.title("Custom Image Generator")
        self.iconbitmap(TITLE_ICON_PATH)
        self.geometry("1000x600") # <-- Bigger window
        self.resizable(False, False)
        self.wm_attributes("-topmost", True)
        self.excel_data = excel_data
        
        mode = ctk.get_appearance_mode().lower()
        bg_color = DARK_BG if mode == "dark" else LIGHT_BG
        fg_color = DARK_FG if mode == "dark" else LIGHT_FG
        inner_color = DARK_INNER if mode == "dark" else LIGHT_INNER
        
        self.configure(padx=10, pady=10, fg_color=bg_color) # Theme
        self.template_image_path = None
        self.font_file_path = None
        self.last_click = (50, 50)
        self.font_size_var = ctk.StringVar(value="50")
        self.text_color_var = ctk.StringVar(value="black")
        self.ratio_options = ["Original", "4:3", "16:9", "5:8", "1:1", "3:2", "21:9"]
        self.ratio_var = ctk.StringVar(value="Original")

        # --- CONFIGURE TOP-LEVEL GRID ---
        self.grid_rowconfigure(0, weight=0) # Row 0 for top_frame (no expand)
        self.grid_rowconfigure(1, weight=1) # Row 1 for controls/preview (expand)
        self.grid_columnconfigure(0, weight=1) # Col 0 for controls
        self.grid_columnconfigure(1, weight=3) # Col 1 for preview

        # --- Top Frame (Ratio) ---
        top_frame = ctk.CTkFrame(self, corner_radius=15, fg_color=fg_color)
        # --- FIX: Use .grid() instead of .pack() ---
        top_frame.grid(row=0, column=0, columnspan=2, sticky="ew", padx=10, pady=10)
        
        ctk.CTkLabel(top_frame, text="Select Image Ratio:", font=("Arial", 16, "bold")).pack(side="left", padx=(10,5), pady=10)
        self.ratio_menu = ctk.CTkOptionMenu(
            top_frame, 
            values=self.ratio_options, 
            variable=self.ratio_var, 
            command=lambda x: self.update_preview(),
            height=40,
            font=("Arial", 16),
            dropdown_font=("Arial", 16),
            fg_color="green",
            button_color="#45A049",
            button_hover_color="#3E8E41"
        )
        self.ratio_menu.pack(side="left", padx=5, pady=10) # .pack() is OK here, it's inside top_frame

        # --- Main layout (already uses .grid()) ---
        self.control_frame = ctk.CTkFrame(self, corner_radius=15, fg_color=fg_color)
        self.control_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0,10))
        
        self.preview_frame = ctk.CTkFrame(self, corner_radius=15, fg_color=fg_color)
        self.preview_frame.grid(row=1, column=1, sticky="nsew", padx=(0,10), pady=(0,10))
        
        # --- Control Frame Widgets (Using .grid()) ---
        
        # Configure grid inside control_frame
        self.control_frame.grid_columnconfigure(0, weight=1) # Make column expandable
        
        ctk.CTkLabel(self.control_frame, text="Image Controls", font=("Arial", 22, "bold")).grid(
            row=0, column=0, pady=20, padx=20
        )
        
        # --- Button Styling ---
        btn_font = ("Arial", 16, "bold")
        btn_height = 45
        btn_radius = 10
        btn_fg = "green"
        btn_hover = "#45A049"
        
        self.select_template_btn = ctk.CTkButton(
            self.control_frame, text="Select Template Image", 
            command=self.select_template,
            font=btn_font, height=btn_height, corner_radius=btn_radius, fg_color=btn_fg, hover_color=btn_hover
        )
        self.select_template_btn.grid(row=1, column=0, pady=10, padx=20, sticky="ew")

        self.select_font_btn = ctk.CTkButton(
            self.control_frame, text="Select Font File (.ttf)", 
            command=self.select_font,
            font=btn_font, height=btn_height, corner_radius=btn_radius, fg_color=btn_fg, hover_color=btn_hover
        )
        self.select_font_btn.grid(row=2, column=0, pady=10, padx=20, sticky="ew")

        ctk.CTkLabel(self.control_frame, text="Font Size:", font=btn_font).grid(
            row=3, column=0, pady=(10, 5), padx=20, sticky="w"
        )
        self.font_size_entry = ctk.CTkEntry(
            self.control_frame, textvariable=self.font_size_var, 
            corner_radius=10, fg_color=inner_color, border_width=0,
            height=40, font=("Arial", 16)
        )
        self.font_size_entry.grid(row=4, column=0, pady=5, padx=20, sticky="ew")
        self.font_size_entry.bind("<KeyRelease>", lambda e: self.update_preview())
        self.font_size_entry.bind("<Return>", lambda e: self.generate_images_with_progress())

        ctk.CTkLabel(self.control_frame, text="Text Color:", font=btn_font).grid(
            row=5, column=0, pady=(10, 5), padx=20, sticky="w"
        )
        color_btn = ctk.CTkButton(
            self.control_frame, text="Choose Color", 
            command=self.choose_color,
            font=btn_font, height=btn_height, corner_radius=btn_radius, fg_color=btn_fg, hover_color=btn_hover
        )
        color_btn.grid(row=6, column=0, pady=5, padx=20, sticky="ew")

        self.set_position_btn = ctk.CTkButton(
            self.control_frame, text="Set Text Position\n(Click Preview Image)", 
            command=self.instruct_set_position,
            font=btn_font, height=50, corner_radius=btn_radius, fg_color=btn_fg, hover_color=btn_hover
        )
        self.set_position_btn.grid(row=7, column=0, pady=10, padx=20, sticky="ew")

        # Add a spacer frame that will expand and push the button down
        spacer_frame = ctk.CTkFrame(self.control_frame, fg_color="transparent")
        spacer_frame.grid(row=8, column=0, sticky="nsew", pady=0)
        self.control_frame.grid_rowconfigure(8, weight=1) # Configure the spacer's row to expand

        self.generate_btn = ctk.CTkButton(
            self.control_frame,
            text="Generate Images",
            command=self.generate_images_with_progress,
            font=("Arial", 18, "bold"), height=50, corner_radius=btn_radius,
            fg_color="#006400", # Darker Green for emphasis
            hover_color="#008000" # Brighter Green hover
        )
        self.generate_btn.grid(row=9, column=0, pady=20, padx=20, sticky="ew")

        # --- Preview Frame Widgets ---
        self.preview_frame.grid_rowconfigure(0, weight=1)
        self.preview_frame.grid_columnconfigure(0, weight=1)
        
        self.canvas = ctk.CTkCanvas(self.preview_frame, bg=inner_color, width=800, height=800, highlightthickness=0, bd=0)
        self.canvas.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        self.canvas.bind("<Button-1>", self.canvas_click)
        
        # Bind Enter key to the whole window
        self.bind("<Return>", lambda e: self.generate_images_with_progress())
        self.preview_image = None
        self.update_preview() # Initial call
        center_window(self) # Center after everything is drawn

    def choose_color(self):
        color = colorchooser.askcolor(title="Choose text color", parent=self)
        if color and color[1]:
            self.text_color_var.set(color[1])
            self.update_preview()
            
    def select_template(self):
        path = filedialog.askopenfilename(filetypes=[("Image Files", "*.png;*.jpg;*.jpeg")], parent=self)
        if path:
            self.template_image_path = path
            self.update_preview()
            
    def select_font(self):
        path = filedialog.askopenfilename(filetypes=[("Font Files", "*.ttf")], parent=self)
        if path:
            self.font_file_path = path
            self.update_preview() # Update preview when font changes
            
    def canvas_click(self, event):
        self.last_click = (event.x, event.y)
        self.update_preview()
        
    def instruct_set_position(self):
        messagebox.showinfo("Set Position", "Click on the preview image to set the text position.", parent=self)
        
    def update_preview(self):
        if not self.template_image_path:
            # --- NEW: Show placeholder text if no image ---
            self.canvas.delete("all")
            self.canvas.config(width=800, height=600) # Default size
            self.canvas.create_text(
                400, 300, 
                text="Please select a template image...", 
                font=("Arial", 24, "bold"), 
                fill="gray"
            )
            return
            
        try:
            # --- Get canvas size for scaling ---
            self.preview_frame.update_idletasks()
            canvas_width = self.canvas.winfo_width()
            canvas_height = self.canvas.winfo_height()
            
            if canvas_width < 100 or canvas_height < 100: # Min size
                canvas_width, canvas_height = 800, 600 # Fallback

            ratio = self.ratio_var.get()
            img_temp = Image.open(self.template_image_path).convert("RGB")
            
            if ratio == "4:3": new_size = (canvas_width, int(canvas_width * 3/4))
            elif ratio == "16:9": new_size = (canvas_width, int(canvas_width * 9/16))
            elif ratio == "5:8": new_size = (canvas_width, int(canvas_width * 8/5))
            elif ratio == "1:1": new_size = (canvas_width, canvas_width)
            elif ratio == "3:2": new_size = (canvas_width, int(canvas_width * 2/3))
            elif ratio == "21:9": new_size = (canvas_width, int(canvas_width * 9/21))
            else: # Original
                orig_size = img_temp.size
                ratio_val = min(canvas_width/orig_size[0], canvas_height/orig_size[1])
                new_size = (int(orig_size[0]*ratio_val), int(orig_size[1]*ratio_val))

            # Ensure image fits canvas
            if new_size[1] > canvas_height:
                new_size = (int(new_size[0] * (canvas_height / new_size[1])), canvas_height)
                
            img = img_temp.resize(new_size, Image.Resampling.LANCZOS)
            
            draw = ImageDraw.Draw(img)
            font_size = int(self.font_size_var.get() or 50)
            font_path = self.font_file_path if self.font_file_path else "arial.ttf"
            font = ImageFont.truetype(font_path, font_size)
            draw.text(self.last_click, "{User_Name}", font=font, fill=self.text_color_var.get())
            
            self.preview_image = ImageTk.PhotoImage(img)
            self.canvas.config(width=new_size[0], height=new_size[1]) # Resize canvas to image
            self.canvas.delete("all")
            self.canvas.create_image(0, 0, image=self.preview_image, anchor="nw")
            
        except Exception as e:
            messagebox.showerror("Preview Error", f"Error updating preview: {e}", parent=self)
            
    def generate_images_with_progress(self):
        total = len(self.excel_data)
        if total == 0:
            messagebox.showerror("Error", "No Excel data found. Please load contacts first.", parent=self)
            return
        if not self.template_image_path:
            messagebox.showerror("Error", "No template image selected.", parent=self)
            return
            
        prog = ProgressPopup(self, "Generating Images", total)
        prog.geometry(
            f"500x300+{self.winfo_rootx() + (self.winfo_width()-500)//2}"
            f"+{self.winfo_rooty() + (self.winfo_height()-300)//2}"
        )
        self.after(100, lambda: threading.Thread(target=self.generate_images, args=(prog,), daemon=True).start())
        
    def generate_images(self, prog):
        try:
            font_size = int(self.font_size_var.get())
        except ValueError:
            # --- FIX: Schedule UI calls on main thread ---
            self.after(0, lambda: messagebox.showerror("Error", "Invalid font size.", parent=self))
            self.after(0, prog.close)
            return
            
        font_path = self.font_file_path if self.font_file_path else "arial.ttf"
        text_color = self.text_color_var.get()
        text_pos = self.last_click
        ratio = self.ratio_var.get()
        
        # --- FIX: Store preview size to scale click position ---
        # We need to get this from the main thread *before* starting.
        # Let's pass the preview size to the thread.
        preview_width = self.canvas.winfo_width()
        preview_height = self.canvas.winfo_height()
        if preview_width < 100: preview_width = 800 # Fallback
        if preview_height < 100: preview_height = 600 # Fallback
        
        # Calculate click as a ratio (0.0 to 1.0)
        click_ratio_x = self.last_click[0] / preview_width
        click_ratio_y = self.last_click[1] / preview_height

        # --- Use a fixed high-res size for generation ---
        base_width = 1200 
        img_temp_orig = Image.open(self.template_image_path).convert("RGB")
        orig_size = img_temp_orig.size
        
        if ratio == "4:3": new_size = (base_width, int(base_width * 3/4))
        elif ratio == "16:9": new_size = (base_width, int(base_width * 9/16))
        elif ratio == "5:8": new_size = (base_width, int(base_width * 8/5))
        elif ratio == "1:1": new_size = (base_width, base_width)
        elif ratio == "3:2": new_size = (base_width, int(base_width * 2/3))
        elif ratio == "21:9": new_size = (base_width, int(base_width * 9/21))
        else: # Original
            new_size = (orig_size[0], orig_size[1]) # Use original full size
        
        img_temp_orig.close() # Close file handle

        count = 0
        for idx, entry in enumerate(self.excel_data, start=1):
            if not self.winfo_exists(): # Stop thread if window is closed
                return
            try:
                prog.update_progress(idx)
            except Exception:
                pass # Ignore if prog window was closed
                
            if entry.get("skip", False):
                continue
                
            phone = entry.get("phone", "").strip()
            if not phone:
                entry['image_path'] = None
                continue
                
            safe_phone = re.sub(r'[<>:"/\\|?*]', '_', phone)
            
            try:
                img = Image.open(self.template_image_path).convert("RGB").resize(new_size, Image.Resampling.LANCZOS)
                draw = ImageDraw.Draw(img)
                text_to_draw = entry.get("name", "").strip() or f"{idx}"
                
                # --- Scale font size and position ---
                # Scale font size relative to the new image width
                scaled_font_size = int(font_size * (new_size[0] / preview_width))
                font_obj = ImageFont.truetype(font_path, scaled_font_size)
                
                # Scale click position based on ratio
                scaled_pos_x = int(click_ratio_x * new_size[0])
                scaled_pos_y = int(click_ratio_y * new_size[1])
                scaled_pos = (scaled_pos_x, scaled_pos_y)
                
                draw.text(scaled_pos, text_to_draw, font=font_obj, fill=text_color)
                output_path = os.path.join(OUTPUT_IMG_FOLDER, f"{safe_phone}.png")
                img.save(output_path)
                entry['image_path'] = output_path
                count += 1
            except Exception as ex:
                print(f"Error generating image for phone {phone}: {ex}")
                
        # --- FIX: Schedule UI updates on the main thread ---
        def finish_generation():
            if not self.winfo_exists(): return # Don't do anything if window is gone
            messagebox.showinfo("Generation Complete", f"Generated {count} images in {OUTPUT_IMG_FOLDER}.", parent=self)
            try:
                prog.close()
            except:
                pass
            self.destroy() # This will now run on the main thread

        self.after(0, finish_generation)


class SchedulePopup(ctk.CTkToplevel):
    def __init__(self, master, on_schedule_set, *args, **kwargs):
        super().__init__(master, *args, **kwargs)
        self.title("Schedule Sending")
        self.iconbitmap(TITLE_ICON_PATH)
        self.geometry("450x355") # Increased size for bigger elements
        self.resizable(False, False)
        self.on_schedule_set = on_schedule_set
        self.wm_attributes("-topmost", True)
        self.transient(master) # Keep popup on top of main app

        # --- Theme Colors ---
        mode = ctk.get_appearance_mode().lower()
        bg_color = DARK_BG if mode == "dark" else LIGHT_BG
        fg_color = DARK_FG if mode == "dark" else LIGHT_FG
        inner_color = DARK_INNER if mode == "dark" else LIGHT_INNER
        self.btn_hover_color = "#333333" if mode == "dark" else "#E0E0E0"

        # --- Load Icons (Theme Aware) ---
        self.ICON_SIZE = (50, 50)
        self.BTN_SIZE = (80, 80)
        self.ENTRY_FONT = ("Arial", 48, "bold")
        self.LABEL_FONT = ("Arial", 16, "bold")
        
        try:
            # Load both light and dark versions
            up_img_light = Image.open(UP_ARROW_LIGHT_PATH).resize(self.ICON_SIZE, Image.Resampling.LANCZOS)
            up_img_dark = Image.open(UP_ARROW_DARK_PATH).resize(self.ICON_SIZE, Image.Resampling.LANCZOS)
            down_img_light = Image.open(DOWN_ARROW_LIGHT_PATH).resize(self.ICON_SIZE, Image.Resampling.LANCZOS)
            down_img_dark = Image.open(DOWN_ARROW_DARK_PATH).resize(self.ICON_SIZE, Image.Resampling.LANCZOS)

            # CTkImage will handle switching automatically
            self.up_icon = ctk.CTkImage(light_image=up_img_light, dark_image=up_img_dark, size=self.ICON_SIZE)
            self.down_icon = ctk.CTkImage(light_image=down_img_light, dark_image=down_img_dark, size=self.ICON_SIZE)
            
        except Exception as e:
            print(f"Error loading schedule icons: {e}")
            self.up_icon = None # Fallback
            self.down_icon = None # Fallback

        # --- Data Variables ---
        now = datetime.now()
        # Set default values based on current time
        current_hour_12 = now.hour % 12 if now.hour % 12 != 0 else 12
        self.hour_var = tk.StringVar(value=f"{current_hour_12:02d}")
        self.min_var = tk.StringVar(value=f"{now.minute:02d}")
        self.ampm_var = tk.StringVar(value="AM" if now.hour < 12 else "PM")

        # --- Layout ---
        self.configure(fg_color=bg_color) 
        
        # Main container to center everything
        container = ctk.CTkFrame(self, corner_radius=10, fg_color=fg_color)
        container.pack(expand=True, fill="both", padx=20, pady=20)

        container.rowconfigure(0, weight=3) # Time pickers
        container.rowconfigure(1, weight=1) # Set button
        container.rowconfigure(2, weight=1) # Cancel button
        container.columnconfigure(0, weight=1) # Center column

        # Frame for the 3 time pickers
        time_frame = ctk.CTkFrame(container, corner_radius=10, fg_color="transparent")
        time_frame.grid(row=0, column=0, sticky="nsew", pady=10)
        time_frame.columnconfigure((0, 1, 2), weight=1) # 3 columns for H, M, A/P
        time_frame.rowconfigure(0, weight=0) # Label row
        time_frame.rowconfigure(1, weight=1) # Picker row

        # --- Hour Picker ---
        ctk.CTkLabel(time_frame, text="Hour", font=self.LABEL_FONT).grid(row=0, column=0)
        hour_frame = self.create_spinner_widget(time_frame, self.hour_var, self.increment_hour, self.decrement_hour, inner_color)
        hour_frame.grid(row=1, column=0, padx=5)
        
        # --- Minute Picker ---
        ctk.CTkLabel(time_frame, text="Minute", font=self.LABEL_FONT).grid(row=0, column=1)
        min_frame = self.create_spinner_widget(time_frame, self.min_var, self.increment_min, self.decrement_min, inner_color)
        min_frame.grid(row=1, column=1, padx=5)

        # --- AM/PM Picker ---
        ctk.CTkLabel(time_frame, text="Period", font=self.LABEL_FONT).grid(row=0, column=2)
        ampm_frame = self.create_spinner_widget(time_frame, self.ampm_var, self.toggle_ampm, self.toggle_ampm, inner_color)
        ampm_frame.grid(row=1, column=2, padx=5)

        # --- Big Action Buttons ---
        self.set_btn = ctk.CTkButton(
            container, 
            text="Set Schedule", 
            corner_radius=100, 
            font=("Arial", 24, "bold"), 
            command=self.set_schedule,
            height=60,
            width=250,
            fg_color="green",
            hover_color="#45A049"
        )
        self.set_btn.grid(row=1, column=0, pady=10)

        self.cancel_btn = ctk.CTkButton(
            container, 
            text="Cancel", 
            corner_radius=100, 
            font=("Arial", 18, "bold"), 
            command=self.destroy,
            height=50,
            width=200,
            fg_color="#555555", # Neutral gray
            hover_color="#777777"
        )
        self.cancel_btn.grid(row=2, column=0, pady=5)

        center_window(self)
        # Bind Enter key to the whole window
        self.bind("<Return>", lambda e: self.set_schedule())    

    def create_spinner_widget(self, parent, text_var, up_cmd, down_cmd, inner_color):
        """Helper to create the up/down button and entry widget."""
        frame = ctk.CTkFrame(parent, fg_color="transparent")
        frame.rowconfigure((0, 2), weight=1) # Buttons
        frame.rowconfigure(1, weight=0) # Entry
        frame.columnconfigure(0, weight=1)

        # Up Button
        up_btn = ctk.CTkButton(
            frame, 
            text="", 
            image=self.up_icon, 
            width=self.BTN_SIZE[0], 
            height=self.BTN_SIZE[1], 
            fg_color="transparent", 
            hover_color=self.btn_hover_color,
            command=up_cmd
        )
        up_btn.grid(row=0, column=0, pady=5)

        # Entry Display
        entry = ctk.CTkEntry(
            frame, 
            width=120, 
            height=70, 
            corner_radius=10, 
            textvariable=text_var, 
            font=self.ENTRY_FONT, 
            fg_color=inner_color, 
            border_width=0,
            justify="center",
            state="readonly" # Use readonly state
        )
        entry.grid(row=1, column=0, pady=5)

        # Down Button
        down_btn = ctk.CTkButton(
            frame, 
            text="", 
            image=self.down_icon, 
            width=self.BTN_SIZE[0], 
            height=self.BTN_SIZE[1], 
            fg_color="transparent", 
            hover_color=self.btn_hover_color,
            command=down_cmd
        )
        down_btn.grid(row=2, column=0, pady=5)
        
        return frame

    def increment_hour(self):
        try:
            val = int(self.hour_var.get())
        except ValueError:
            val = 12
        val = 1 if val >= 12 else val + 1
        self.hour_var.set(f"{val:02d}") # Format to 2 digits

    def decrement_hour(self):
        try:
            val = int(self.hour_var.get())
        except ValueError:
            val = 1
        val = 12 if val <= 1 else val - 1
        self.hour_var.set(f"{val:02d}") # Format to 2 digits

    def increment_min(self):
        try:
            val = int(self.min_var.get())
        except ValueError:
            val = 0
        val = 0 if val >= 59 else val + 1
        self.min_var.set(f"{val:02d}") # Format to 2 digits

    def decrement_min(self):
        try:
            val = int(self.min_var.get())
        except ValueError:
            val = 0
        val = 59 if val <= 0 else val - 1
        self.min_var.set(f"{val:02d}") # Format to 2 digits

    def toggle_ampm(self):
        self.ampm_var.set("PM" if self.ampm_var.get().upper() == "AM" else "AM")

    def set_schedule(self):
        try:
            # Get 0-padded values
            hour_str = self.hour_var.get()
            min_str = self.min_var.get()
            
            # Validate format (should be 2 digits)
            if not (hour_str.isdigit() and (len(hour_str) == 2 or len(hour_str) == 1)): # Allow 1 or 2 digits
                 hour_str = f"{int(hour_str):02d}" # Re-format
            if not (min_str.isdigit() and (len(min_str) == 2 or len(min_str) == 1)):
                 min_str = f"{int(min_str):02d}" # Re-format
            
            # Final check
            if not (hour_str.isdigit() and len(hour_str) == 2):
                 raise ValueError("Hour format invalid.")
            if not (min_str.isdigit() and len(min_str) == 2):
                 raise ValueError("Minute format invalid.")

            hour = int(hour_str)
            minute = int(min_str)
            
            if not (1 <= hour <= 12):
                raise ValueError("Hour must be 1-12")
            if not (0 <= minute < 60):
                raise ValueError("Minute must be between 0 and 59")
        except ValueError as ve:
            messagebox.showerror("Invalid Time", f"Please enter valid time values: {ve}", parent=self)
            return
            
        ampm = self.ampm_var.get().upper()
        if ampm == "PM" and hour != 12:
            hour_24 = hour + 12
        elif ampm == "AM" and hour == 12:
            hour_24 = 0 # 12 AM is 00:00
        else:
            hour_24 = hour
            
        now = datetime.now()
        schedule_time = now.replace(hour=hour_24, minute=minute, second=0, microsecond=0)
        
        # If schedule time is in the past, set it for the next day
        if schedule_time <= now:
            schedule_time += timedelta(days=1)
            
        self.on_schedule_set(schedule_time)
        self.destroy()

class FindPopup(ctk.CTkToplevel):
    """A popup window for finding text in a target CTkTextbox."""
    def __init__(self, master, target_textbox, **kwargs):
        super().__init__(master, **kwargs)
        self.target = target_textbox
        self.master_app = master # Reference to WabulkXpressApp
        
        self.title("Find")
        self.geometry("350x130")
        self.resizable(False, False)
        self.attributes("-topmost", True)
        self.transient(master)
        
        # --- Theme Colors ---
        mode = ctk.get_appearance_mode().lower()
        bg_color = DARK_BG if mode == "dark" else LIGHT_BG
        fg_color = DARK_FG if mode == "dark" else LIGHT_FG
        inner_color = DARK_INNER if mode == "dark" else LIGHT_INNER
        
        self.configure(fg_color=bg_color)
        self.protocol("WM_DELETE_WINDOW", self.on_close_popup)
        self.bind("<Escape>", lambda e: self.on_close_popup())
        
        # --- Layout ---
        main_frame = ctk.CTkFrame(self, fg_color=fg_color, corner_radius=10)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        main_frame.grid_columnconfigure(0, weight=1)
        main_frame.grid_columnconfigure(1, weight=0)

        self.find_entry = ctk.CTkEntry(
            main_frame, 
            placeholder_text="Find...",
            font=("Arial", 16),
            height=40,
            fg_color=inner_color,
            border_width=0
        )
        self.find_entry.grid(row=0, column=0, columnspan=2, sticky="ew", padx=10, pady=(10, 5))
        self.find_entry.bind("<Return>", self.find_next)
        
        self.find_btn = ctk.CTkButton(
            main_frame,
            text="Find Next",
            command=self.find_next,
            font=("Arial", 16, "bold"),
            fg_color="green",
            hover_color="#45A049",
            height=40
        )
        self.find_btn.grid(row=1, column=0, sticky="ew", padx=(10, 5), pady=10)
        
        self.cancel_btn = ctk.CTkButton(
            main_frame,
            text="Cancel",
            command=self.on_close_popup,
            font=("Arial", 16),
            fg_color="#555555",
            hover_color="#777777",
            height=40,
            width=80
        )
        self.cancel_btn.grid(row=1, column=1, sticky="e", padx=(5, 10), pady=10)
        
        self.find_entry.focus_set() # Focus the entry box on open
        center_window(self)

    def find_next(self, event=None):
        search_term = self.find_entry.get()
        if not search_term:
            return

        # Clear previous highlight
        self.target.tag_remove("find_highlight", "1.0", "end")

        # Search from the last known index
        start_index = self.master_app.find_last_index
        pos = self.target.search(search_term, start_index, "end", nocase=True)

        if pos:
            # Match found
            end_pos = f"{pos}+{len(search_term)}c"
            
            # Configure and apply the tag
            self.target.tag_config("find_highlight", background="yellow", foreground="black")
            self.target.tag_add("find_highlight", pos, end_pos)
            
            self.target.see(pos) # Scroll to the match
            self.master_app.find_last_index = end_pos # Update index for next search
        else:
            # No match found from current position
            if messagebox.askyesno("Find", "End of document reached.\nSearch from beginning?", parent=self):
                # Reset index and search again
                self.master_app.find_last_index = "1.0"
                self.find_next()
            else:
                # User said no, reset index for next manual search
                self.master_app.find_last_index = "1.0"

    def on_close_popup(self):
        """Clears highlights and resets the search index on close."""
        self.target.tag_remove("find_highlight", "1.0", "end")
        self.master_app.find_last_index = "1.0"
        self.destroy()


class TableFindPopup(ctk.CTkToplevel):
    """A popup window for finding text in the ExcelTable."""
    def __init__(self, master, target_table, **kwargs):
        super().__init__(master, **kwargs)
        self.target_table = target_table # This is the ExcelTable instance
        self.master_app = master     # This is the WabulkXpressApp
        
        # Keep track of the last search position
        if not hasattr(self.master_app, 'excel_find_last_pos'):
            self.master_app.excel_find_last_pos = {"row": 0, "col_idx": -1}

        self.title("Find in Table")
        self.geometry("350x130")
        self.resizable(False, False)
        self.attributes("-topmost", True)
        self.transient(master)
        
        mode = ctk.get_appearance_mode().lower()
        bg_color = DARK_BG if mode == "dark" else LIGHT_BG
        fg_color = DARK_FG if mode == "dark" else LIGHT_FG
        inner_color = DARK_INNER if mode == "dark" else LIGHT_INNER
        
        self.configure(fg_color=bg_color)
        self.protocol("WM_DELETE_WINDOW", self.on_close_popup)
        self.bind("<Escape>", lambda e: self.on_close_popup())
        
        main_frame = ctk.CTkFrame(self, fg_color=fg_color, corner_radius=10)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        main_frame.grid_columnconfigure(0, weight=1)
        main_frame.grid_columnconfigure(1, weight=0)

        self.find_entry = ctk.CTkEntry(
            main_frame, placeholder_text="Find...",
            font=("Arial", 16), height=40,
            fg_color=inner_color, border_width=0
        )
        self.find_entry.grid(row=0, column=0, columnspan=2, sticky="ew", padx=10, pady=(10, 5))
        self.find_entry.bind("<Return>", self.find_next)
        
        self.find_btn = ctk.CTkButton(
            main_frame, text="Find Next", command=self.find_next,
            font=("Arial", 16, "bold"), fg_color="green",
            hover_color="#45A049", height=40
        )
        self.find_btn.grid(row=1, column=0, sticky="ew", padx=(10, 5), pady=10)
        
        self.cancel_btn = ctk.CTkButton(
            main_frame, text="Cancel", command=self.on_close_popup,
            font=("Arial", 16), fg_color="#555555",
            hover_color="#777777", height=40, width=80
        )
        self.cancel_btn.grid(row=1, column=1, sticky="e", padx=(5, 10), pady=10)
        
        self.find_entry.focus_set()
        center_window(self)

    def find_next(self, event=None):
        search_term = self.find_entry.get().lower()
        if not search_term:
            return

        # Clear previous highlights
        self.clear_all_highlights()

        # Define the order of columns to search
        search_cols = ["phone", "name", "custom1", "custom2"]
        
        start_row = self.master_app.excel_find_last_pos["row"]
        start_col_idx = self.master_app.excel_find_last_pos["col_idx"] + 1

        # Iterate through the table
        for r_idx in range(start_row, len(self.target_table.rows)):
            row = self.target_table.rows[r_idx]
            
            for c_idx in range(start_col_idx, len(search_cols)):
                col_name = search_cols[c_idx]
                var_name = f"{col_name}_var" # e.g., "phone_var"
                entry_name = col_name        # e.g., "phone"
                
                if var_name in row:
                    cell_content = row[var_name].get().lower()
                    if search_term in cell_content:
                        # --- Match Found! ---
                        widget = row[entry_name]
                        
                        # 1. Focus and Select
                        widget.focus_set()
                        start_pos = cell_content.find(search_term)
                        end_pos = start_pos + len(search_term)
                        widget.select_range(start_pos, end_pos) # CTkEntry uses select_range
                        
                        # 2. Scroll to the widget
                        try:
                            self.target_table.update_idletasks()
                            row_frame_y = row["row_frame"].winfo_y()
                            inner_frame_height = self.target_table._parent_canvas.winfo_children()[0].winfo_height()
                            if inner_frame_height > 0:
                                fraction = row_frame_y / inner_frame_height
                                self.target_table._parent_canvas.yview_moveto(fraction)
                        except Exception as e:
                            print(f"Error scrolling to widget: {e}")

                        # 3. Update last position
                        self.master_app.excel_find_last_pos = {"row": r_idx, "col_idx": c_idx}
                        return # Stop after finding one match
            
            # Reset col index for the next row
            start_col_idx = 0 

        # --- No Match Found ---
        if messagebox.askyesno("Find", "End of table reached.\nSearch from beginning?", parent=self):
            self.master_app.excel_find_last_pos = {"row": 0, "col_idx": -1} # Reset
            self.find_next()
        else:
            self.master_app.excel_find_last_pos = {"row": 0, "col_idx": -1} # Reset for next time

    def clear_all_highlights(self):
        """Iterates all entry widgets and clears selection."""
        for row in self.target_table.rows:
            for col_name in ["phone", "name", "custom1", "custom2"]:
                if col_name in row:
                    try:
                        row[col_name].select_clear()
                    except Exception:
                        pass # Ignore if widget doesn't exist

    def on_close_popup(self):
        self.clear_all_highlights()
        self.destroy()

class TranslatePopup(ctk.CTkToplevel):
    def __init__(self, master, process_callback, *args, **kwargs):
        super().__init__(master, *args, **kwargs)
        self.title("Translate Message")
        self.iconbitmap(TITLE_ICON_PATH)
        self.geometry("550x300") # New geometry
        self.resizable(False, False)
        self.wm_attributes("-topmost", True)
        self.process_callback = process_callback
        
        mode = ctk.get_appearance_mode().lower()
        bg_color = DARK_BG if mode == "dark" else LIGHT_BG
        fg_color = DARK_FG if mode == "dark" else LIGHT_FG # Frame background
        
        self.configure(fg_color=bg_color) # Theme
        
        # --- Main Layout (2 columns) ---
        self.grid_columnconfigure(0, weight=1) # Left controls
        self.grid_columnconfigure(1, weight=1) # Right image
        self.grid_rowconfigure(0, weight=1)
        
        # --- Left Frame (Controls) ---
        left_frame = ctk.CTkFrame(self, fg_color=fg_color, corner_radius=15)
        left_frame.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
        
        left_frame.grid_columnconfigure(0, weight=1)
        left_frame.grid_rowconfigure(0, weight=0) # Label
        left_frame.grid_rowconfigure(1, weight=1) # Dropdown (centered)
        left_frame.grid_rowconfigure(2, weight=1) # Buttons (centered)
        
        # 1. Label
        ctk.CTkLabel(left_frame, text="Select Target Language:", font=("Arial", 18, "bold")).grid(row=0, column=0, pady=(20, 10), padx=20)
        
        # 2. Languages List (Expanded)
        languages = [
            "English", "Hindi", "Marathi", "Spanish", "French", "German", "Italian", 
            "Portuguese", "Russian", "Chinese (Simplified)", "Japanese", "Korean", 
            "Arabic", "Bengali", "Gujarati", "Kannada", "Malayalam", "Odia", 
            "Punjabi", "Tamil", "Telugu", "Urdu", "Dutch", "Turkish", "Vietnamese"
        ]
        self.language_var = tk.StringVar(value="English")
        
        # Get theme colors for the dropdown
        inner_color = DARK_INNER if mode == "dark" else LIGHT_INNER
        text_color = "white" if mode == "dark" else "black"

        self.language_menu = ctk.CTkOptionMenu(
            left_frame, 
            values=languages, 
            variable=self.language_var,
            font=("Arial", 16, "bold"),         # Bigger font
            dropdown_font=("Arial", 16),      # Bigger font
            height=50,                        # Taller
            corner_radius=100,                # Full rounded edge
            fg_color="green",                 # Green color
            button_color="#45A049",           # Green arrow button
            button_hover_color="#3E8E41",     # Green arrow hover
            # --- UPDATED LINES ---
            dropdown_fg_color=DARK_BG,        # <-- CHANGED: Pitch black background
            dropdown_hover_color=inner_color, # Use the inner color for hover
            dropdown_text_color=text_color,   # Set appropriate text colo
        )
        # Added more space with pady=20
        self.language_menu.grid(row=1, column=0, pady=20, padx=30, sticky="ew")

        # 4. Button Frame
        button_frame = ctk.CTkFrame(left_frame, fg_color="transparent")
        button_frame.grid(row=2, column=0, pady=(10, 20))
        button_frame.grid_columnconfigure((0, 1), weight=1)

        btn_font = ("Arial", 16, "bold")
        btn_height = 50
        btn_width = 120
        btn_radius = 100

        # 5. OK Button
        self.ok_button = ctk.CTkButton(
            button_frame, 
            text="OK", 
            command=self.on_ok,
            font=btn_font,
            height=btn_height,
            width=btn_width,
            corner_radius=btn_radius,
            fg_color="green",
            hover_color="#45A049"
        )
        self.ok_button.grid(row=0, column=0, padx=5)

        # 6. Cancel Button
        self.cancel_button = ctk.CTkButton(
            button_frame, 
            text="Cancel", 
            command=self.destroy,
            font=btn_font,
            height=btn_height,
            width=btn_width,
            corner_radius=btn_radius,
            fg_color="#555555", # Neutral gray
            hover_color="#777777"
        )
        self.cancel_button.grid(row=0, column=1, padx=5)

        # --- Right Frame (Image) ---
        right_frame = ctk.CTkFrame(self, fg_color=fg_color, corner_radius=15)
        right_frame.grid(row=0, column=1, sticky="nsew", padx=(0, 20), pady=20)
        right_frame.grid_rowconfigure(0, weight=1)
        right_frame.grid_columnconfigure(0, weight=1)

        try:
            TRANSLATE_ICON_PATH = os.path.join(BIN_FOLDER, "translate_icon.png")
            img_size = (150, 150)
            img = Image.open(TRANSLATE_ICON_PATH).resize(img_size, Image.Resampling.LANCZOS)
            translate_icon = ctk.CTkImage(light_image=img, dark_image=img, size=img_size)
            
            img_label = ctk.CTkLabel(right_frame, image=translate_icon, text="")
            img_label.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
        except Exception as e:
            print(f"Error loading translate icon: {e}")
            fallback_label = ctk.CTkLabel(right_frame, text="Translate", font=("Arial", 24, "bold"))
            fallback_label.grid(row=0, column=0, sticky="nsew")

        # --- Finalize ---
        center_window(self)
        self.bind("<Return>", lambda e: self.on_ok())
        self.bind("<Escape>", lambda e: self.destroy())

    def on_ok(self):
        lang = self.language_var.get()
        self.process_callback(lang)
        self.destroy()




# --- (Make sure AI_ICON_PATH is defined globally near other paths) ---
AI_ICON_PATH = os.path.join(BIN_FOLDER, "ai_icon.png")

# --- (Replace the existing AIPopup class with this) ---
# --- (Make sure AI_ICON_PATH is defined globally near other paths) ---
AI_ICON_PATH = os.path.join(BIN_FOLDER, "ai_icon.png") 

# --- (Replace the existing AIPopup class with this) ---
class AIPopup(ctk.CTkToplevel):
    def __init__(self, master, button):
        super().__init__(master)
        self.master = master
        self.overrideredirect(True) # Frameless

        mode = ctk.get_appearance_mode().lower()
        bg_color = DARK_FG if mode == "dark" else LIGHT_FG
        cell_color = DARK_INNER if mode == "dark" else LIGHT_INNER
        cell_hover = "#333333" if mode == "dark" else "#E0E0E0"
        # Removed border_color as border_width is 0
        tk_cell_bg = cell_color # Get hex value before potentially changing cell_color on hover

        # Removed border_width=1
        self.frame = ctk.CTkFrame(self, corner_radius=15, fg_color=bg_color, border_width=0) 
        self.frame.pack(padx=5, pady=5, ipadx=10, ipady=10) # Internal padding for overall popup
        # Configure columns for button grid - uniform ensures equal width
        self.frame.columnconfigure((0, 1, 2), weight=1, uniform="cell_col") 
        # Configure rows containing cells - uniform ensures equal height
        self.frame.rowconfigure((1, 2), weight=1) # Rows for cells
        self.frame.rowconfigure(0, weight=0) # Row for header

        # --- Header: AI Icon (Left) + Text (Center/Right) ---
        header_frame = ctk.CTkFrame(self.frame, fg_color="transparent")
        header_frame.grid(row=0, column=0, columnspan=3, pady=(10, 15), sticky="ew") # Increased bottom padding
        header_frame.columnconfigure(0, weight=0) # Icon column
        header_frame.columnconfigure(1, weight=1) # Text column expands

        ai_popup_icon = None
        if os.path.exists(AI_ICON_PATH):
            try:
                # Icon size for header
                ai_img = Image.open(AI_ICON_PATH).resize((30,30), Image.Resampling.LANCZOS) 
                ai_popup_icon = ctk.CTkImage(ai_img, size=(30,30))
            except Exception as e:
                print(f"Error loading AI icon for popup: {e}")

        if ai_popup_icon:
            icon_label = ctk.CTkLabel(header_frame, image=ai_popup_icon, text="")
            icon_label.grid(row=0, column=0, padx=(10, 5), pady=5, sticky="w") # Place icon left
        
        # Add "Writing Tools" text
        header_text = ctk.CTkLabel(header_frame, text="Writing Tools", font=("Arial", 18, "bold"), anchor="w")
        header_text.grid(row=0, column=1, padx=(5, 10), pady=5, sticky="ew") # Place text next to icon


        # --- Action Buttons (Cells - 2 rows, 3 columns) ---
        # Using CTkImage for consistency now
        # emoji_font_config = ("Arial", 32) # Backup if images fail
        text_font = ctk.CTkFont(size=14) # Text font size

        # Updated data with image paths
        buttons_data = [
            ("Reframe", os.path.join(BIN_FOLDER, "reframe_icon.png"), "Reframe"),
            ("Emoji", os.path.join(BIN_FOLDER, "emoji_icon.png"), "Emoji"),
            ("Professional", os.path.join(BIN_FOLDER, "professional_icon.png"), "Professional"),
            ("Funny", os.path.join(BIN_FOLDER, "funny_icon.png"), "Funny"),
            ("Ask AI", os.path.join(BIN_FOLDER, "ask_ai_icon.png"), "Ask AI"),
            ("Translate", os.path.join(BIN_FOLDER, "translate_icon.png"), "Translate")
        ]

        # Define fixed size for cells and image size
        cell_width = 110 
        cell_height = 110
        image_size = (40, 40) # Size for icons within cells

        self.cell_frames = [] # To manage hover state

        for i, (text, img_path, key) in enumerate(buttons_data):
            # Create a frame for each button to act as the cell with FIXED size
            cell_frame = ctk.CTkFrame(self.frame, fg_color=cell_color, corner_radius=10, width=cell_width, height=cell_height)
            cell_frame.grid(row=(i // 3) + 1, column=(i % 3), sticky="nsew", padx=5, pady=5) 
            cell_frame.configure(cursor="hand2") # Make frame look clickable
            
            cell_frame.grid_propagate(False) # Enforce fixed size

            self.cell_frames.append({"frame": cell_frame, "original_color": cell_color, "hover_color": cell_hover})

            # --- Content inside the cell frame ---
            content_container = ctk.CTkFrame(cell_frame, fg_color="transparent")
            content_container.place(relx=0.5, rely=0.5, anchor="center") 

            # Image Label (Top) - Load image using CTkImage
            cell_icon = None
            emoji_label = None # Define variable outside try block
            if os.path.exists(img_path):
                try:
                    img = Image.open(img_path).resize(image_size, Image.Resampling.LANCZOS)
                    cell_icon = ctk.CTkImage(img, size=image_size)
                    emoji_label = ctk.CTkLabel(content_container, image=cell_icon, text="", anchor="center")
                    emoji_label.pack(pady=(10, 5)) # Adjusted padding
                except Exception as e:
                    print(f"Error loading cell icon {img_path}: {e}")
                    # Fallback to text if image fails
                    emoji_label = ctk.CTkLabel(content_container, text="?", font=("Arial", 30), anchor="center")
                    emoji_label.pack(pady=(10, 5)) 
            else:
                 # Fallback to text if path doesn't exist
                 print(f"Cell icon not found: {img_path}")
                 emoji_label = ctk.CTkLabel(content_container, text="X", font=("Arial", 30), anchor="center")
                 emoji_label.pack(pady=(10, 5)) 

            # Text Label (Bottom)
            text_label = ctk.CTkLabel(content_container, text=text, font=text_font, anchor="center")
            text_label.pack(pady=(0, 10)) # Adjusted padding

            # --- Bindings ---
            widgets_to_bind = [cell_frame, content_container, emoji_label, text_label]
            for widget in widgets_to_bind:
                 # Ensure widget is not None before binding
                 if widget: 
                    widget.bind("<Button-1>", lambda e, k=key, popup=self: self.master.process_ai(k, popup=popup))
                    
                    # Apply hover effect using the stored colors and helper methods
                    widget.bind("<Enter>", lambda e, frm=cell_frame, hov=cell_hover: self.on_cell_enter(frm, hov))
                    widget.bind("<Leave>", lambda e, frm=cell_frame, orig=cell_color: self.on_cell_leave(frm, orig))

        # --- Position Popup ---
        self.update_idletasks() # Ensure widgets are drawn
        self.frame.update_idletasks() # Ensure frame size is calculated based on grid
        
        popup_width = self.frame.winfo_reqwidth() + 10 # Add border/padding estimate
        popup_height = self.frame.winfo_reqheight() + 10 # Add border/padding estimate

        btn_x = button.winfo_rootx()
        btn_y = button.winfo_rooty()
        btn_width = button.winfo_width()
        
        # Position popup DIRECTLY ABOVE the button, centered horizontally
        x = btn_x + (btn_width // 2) - (popup_width // 2) 
        y = btn_y - popup_height - 5 # Place 5px above the button

        self.geometry(f"{popup_width}x{popup_height}+{x}+{y}")
        
        self.lift()
        self.attributes("-topmost", True)
        self.bind("<FocusOut>", lambda e: self.destroy_popup()) # Close when focus lost
        self.bind("<Escape>", lambda e: self.destroy_popup()) # Also close if Escape key is pressed
        self.focus_set()

    # --- Helper methods for hover effects (Simplified) ---
    def on_cell_enter(self, frame_widget, hover_color):
        # Only change the frame background on hover
        frame_widget.configure(fg_color=hover_color)

    def on_cell_leave(self, frame_widget, original_color):
        # Restore the original frame background
        frame_widget.configure(fg_color=original_color)

    def destroy_popup(self):
        # Check if the window still exists before destroying
        if self.winfo_exists():
            self.destroy()

# ----------------------- NEW UPDATE POPUP CLASS -----------------------

class UpdatePopup(ctk.CTkToplevel):
    def __init__(self, master, *args, **kwargs):
        super().__init__(master, *args, **kwargs)
        self.master_app = master  # Reference to the main WabulkXpressApp
        self.title("Check for Updates")
        self.iconbitmap(TITLE_ICON_PATH)
        self.geometry("600x400")
        self.resizable(False, False)
        self.wm_attributes("-topmost", True)
        self.transient(master)

        self.installer_url = None
        self.installer_name = None
        self.save_path = None
        
        # --- Theme Colors ---
        mode = ctk.get_appearance_mode().lower()
        self.bg_color = DARK_BG if mode == "dark" else LIGHT_BG
        self.fg_color = DARK_FG if mode == "dark" else LIGHT_FG
        self.inner_color = DARK_INNER if mode == "dark" else LIGHT_INNER
        
        self.configure(fg_color=self.bg_color)
        
        # --- Main Layout ---
        self.grid_columnconfigure(0, weight=1) # Left controls
        self.grid_columnconfigure(1, weight=1) # Right image
        self.grid_rowconfigure(0, weight=1)    # Main content row
        self.grid_rowconfigure(1, weight=0)    # Progress bar row

        # --- Left Frame (Controls) ---
        self.left_frame = ctk.CTkFrame(self, fg_color=self.fg_color, corner_radius=15)
        self.left_frame.grid(row=0, column=0, sticky="nsew", padx=(20, 10), pady=20)
        self.left_frame.grid_columnconfigure(0, weight=1)
        self.left_frame.grid_rowconfigure(0, weight=0) # Title
        self.left_frame.grid_rowconfigure(1, weight=1) # Content (Changelog/Status)
        self.left_frame.grid_rowconfigure(2, weight=0) # Button

        # --- Right Frame (Image) ---
        self.right_frame = ctk.CTkFrame(self, fg_color=self.fg_color, corner_radius=15)
        self.right_frame.grid(row=0, column=1, sticky="nsew", padx=(10, 20), pady=20)
        self.right_frame.grid_rowconfigure(0, weight=1)
        self.right_frame.grid_columnconfigure(0, weight=1)

        try:
            # You must create this image and place it in your 'bin' folder
            UPDATE_GRAPHIC_PATH = os.path.join(BIN_FOLDER, "update_graphic.png") 
            img_size = (200, 200)
            img = Image.open(UPDATE_GRAPHIC_PATH)
            
            # Create rounded corners
            mask = Image.new("L", img.size, 0)
            draw = ImageDraw.Draw(mask)
            draw.rounded_rectangle((0, 0) + img.size, radius=20, fill=255)
            img.putalpha(mask)
            
            update_icon = ctk.CTkImage(light_image=img, dark_image=img, size=img_size)
            
            img_label = ctk.CTkLabel(self.right_frame, image=update_icon, text="")
            img_label.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
        except Exception as e:
            print(f"Error loading update_graphic.png: {e}")
            fallback_label = ctk.CTkLabel(self.right_frame, text="Update", font=("Arial", 24, "bold"))
            fallback_label.grid(row=0, column=0, sticky="nsew")

        # --- Progress Bar (Spans both columns) ---
        self.progress_bar = ctk.CTkProgressBar(self, corner_radius=10, fg_color=self.inner_color)
        self.progress_bar.grid(row=1, column=0, columnspan=2, sticky="ew", padx=20, pady=(0, 20))
        
        # --- Initial UI State ---
        self.setup_checking_ui()
        
        # --- Start Check ---
        self.after(100, self.start_check)
        
        center_window(self)
        
    def setup_checking_ui(self):
        """Sets the UI to the 'Checking for update...' state."""
        self.status_label = ctk.CTkLabel(self.left_frame, text="Checking for updates...", font=("Arial", 18, "bold"))
        self.status_label.grid(row=0, column=0, pady=(20, 10), padx=20)
        
        self.changelog_box = ctk.CTkTextbox(self.left_frame, corner_radius=10, fg_color=self.inner_color, border_width=0, state="disabled")
        # Leave changelog box to maintain layout, but keep it empty
        self.changelog_box.grid(row=1, column=0, sticky="nsew", padx=20, pady=10)

        self.action_button = ctk.CTkButton(
            self.left_frame, text="Cancel", 
            command=self.destroy,
            height=40, font=("Arial", 16, "bold"),
            fg_color="#555555", hover_color="#777777"
        )
        self.action_button.grid(row=2, column=0, pady=(10, 20), padx=20, sticky="ew")
        
        self.progress_bar.configure(mode="indeterminate")
        self.progress_bar.start()

    def start_check(self):
        """Launches the update check in a separate thread."""
        threading.Thread(target=self.check_for_update_thread, daemon=True).start()

    def check_for_update_thread(self):
        """WORKER THREAD: Connects to GitHub API."""
        try:
            response = requests.get(GITHUB_API_URL, timeout=15)
            response.raise_for_status()
            data = response.json()
            latest_version_tag = data.get("tag_name", "0")
            
            # Update the last check time *after* a successful check
            try:
                with open(UPDATE_CHECK_FILE, 'w') as f:
                    f.write(str(time.time()))
            except Exception as e:
                print(f"DEBUG: Failed to update timestamp file: {e}")

            if float(latest_version_tag) > float(CURRENT_VERSION):
                # --- UPDATE AVAILABLE ---
                changelog_body = data.get("body", "No changelog available.")
                assets = data.get("assets", [])
                installer_url = None
                installer_name = None

                # Find the .exe installer in the assets
                for asset in assets:
                    if asset.get("name", "").endswith(".exe"):
                        installer_url = asset.get("browser_download_url")
                        installer_name = asset.get("name")
                        break
                
                if installer_url and installer_name:
                    self.after(0, self.show_update_available_ui, latest_version_tag, changelog_body, installer_url, installer_name)
                else:
                    self.after(0, self.show_error_ui, "Update found, but no .exe installer asset was present in the release.")
            else:
                # --- NO UPDATE ---
                self.after(0, self.show_no_update_ui)
                
        except requests.exceptions.Timeout:
            self.after(0, self.show_error_ui, "Update check timed out. Check your internet connection.")
        except Exception as e:
            self.after(0, self.show_error_ui, f"An error occurred: {e}")

    def show_no_update_ui(self):
        """Updates UI to show 'You are on the latest version'."""
        self.progress_bar.stop()
        self.progress_bar.set(0)
        self.status_label.configure(text=f"You're up to date! (v{CURRENT_VERSION})")
        self.changelog_box.configure(state="normal")
        self.changelog_box.delete("0.0", "end")
        self.changelog_box.insert("0.0", "You are running the latest version of WabulkXpress.")
        self.changelog_box.configure(state="disabled")
        self.action_button.configure(text="Close", command=self.destroy, fg_color="green", hover_color="#45A049")

    def show_error_ui(self, message):
        """Updates UI to show an error message."""
        self.progress_bar.stop()
        self.progress_bar.set(0)
        self.status_label.configure(text="Update Check Failed")
        self.changelog_box.configure(state="normal")
        self.changelog_box.delete("0.0", "end")
        self.changelog_box.insert("0.0", f"Error: {message}")
        self.changelog_box.configure(state="disabled")
        self.action_button.configure(text="Close", command=self.destroy)

    def show_update_available_ui(self, version, changelog, url, name):
        """Updates UI to show update details and Download button."""
        self.installer_url = url
        self.installer_name = name
        
        self.progress_bar.stop()
        self.progress_bar.set(0)
        self.status_label.configure(text=f"New Version Available: v{version}")
        
        self.changelog_box.configure(state="normal")
        self.changelog_box.delete("0.0", "end")
        self.changelog_box.insert("0.0", f"--- Changelog ---\n\n{changelog}")
        self.changelog_box.configure(state="disabled")
        
        self.action_button.configure(
            text="Download Update", 
            command=self.start_download, 
            fg_color="green", 
            hover_color="#45A049"
        )

    def start_download(self):
        """Disables button and starts the download thread."""
        if not self.installer_url:
            self.show_error_ui("Download URL is missing.")
            return

        self.action_button.configure(text="Downloading...", state="disabled")
        self.progress_bar.configure(mode="determinate")
        self.progress_bar.set(0)
        
        # Save the installer to the same directory as the app
        self.save_path = os.path.join(os.getcwd(), self.installer_name)
        
        threading.Thread(target=self.download_thread, daemon=True).start()

    def download_thread(self):
        """WORKER THREAD: Downloads the file."""
        try:
            response = requests.get(self.installer_url, stream=True, timeout=15)
            response.raise_for_status()
            
            total_size = int(response.headers.get('content-length', 0))
            downloaded_size = 0
            
            with open(self.save_path, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)
                    downloaded_size += len(chunk)
                    
                    if total_size > 0:
                        progress = downloaded_size / total_size
                        # Schedule progress update on main thread
                        self.after(0, self.update_download_progress, progress)

            # Download complete
            self.after(0, self.show_install_ui)
            
        except Exception as e:
            self.after(0, self.show_error_ui, f"Download failed: {e}")

    def update_download_progress(self, progress):
        """Updates the progress bar and button text."""
        self.progress_bar.set(progress)
        self.action_button.configure(text=f"Downloading... {int(progress * 100)}%")

    def show_install_ui(self):
        """Updates UI to show 'Install' button."""
        self.progress_bar.set(1)
        self.status_label.configure(text="Download Complete!")
        self.action_button.configure(
            text="Install and Relaunch",
            state="normal",
            command=self.install_and_close,
            fg_color="#006400", # Darker green
            hover_color="#008000"
        )

    def install_and_close(self):
        """Launches the installer and closes the app."""
        try:
            self.master_app.log_live("🚀 Launching installer...")
            os.startfile(self.save_path)
            # Call the main app's close method
            self.master_app.on_close() 
        except Exception as e:
            self.show_error_ui(f"Failed to launch installer: {e}")
            messagebox.showerror("Error", f"Failed to launch installer at {self.save_path}.\n\nError: {e}")
            self.action_button.configure(state="normal") # Re-enable button


class ProfileButton(ctk.CTkFrame):
    def __init__(self, master, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)
        self.profile_url = GITHUB_PROFILE_URL
        
        # --- Image Size & Animation Settings ---
        self.image_size_val = 30
        self.ring_thickness = 3
        self.image_final_size = self.image_size_val + self.ring_thickness * 2 # 36
        
        self.frame_height = 70 # Match other header buttons
        self.start_width = 70  # Collapsed width
        self.end_width = 220   # Expanded width
        
        # --- NEW: Step-based animation ---
        self.animation_steps = 10 # Total number of steps
        self.animation_delay = 15 # Milliseconds per step (approx 150ms total)
        self.width_step = (self.end_width - self.start_width) / self.animation_steps
        self.current_width = self.start_width
        self._animation_job = None # To store the .after() job ID
        # --- End New ---

        self.configure(height=self.frame_height)
        
        mode = ctk.get_appearance_mode().lower()
        self.bg_color = DARK_FG if mode == "dark" else LIGHT_FG
        
        # --- 1. Create Profile Image with Ring (Local) ---
        try:
            avatar = Image.open(MY_PROFILE_PIC).convert("RGBA")
            size = (self.image_size_val, self.image_size_val)
            final_size_tuple = (self.image_final_size, self.image_final_size)

            base = Image.new("RGBA", final_size_tuple, (0, 0, 0, 0))
            draw = ImageDraw.Draw(base)
            
            colors = ["#EA4335", "#FBBC05", "#34A853", "#4285F4"]
            bbox = [(0, 0), final_size_tuple]
            draw.arc(bbox, 270, 360, fill=colors[0], width=self.ring_thickness)
            draw.arc(bbox, 180, 270, fill=colors[1], width=self.ring_thickness)
            draw.arc(bbox, 90, 180, fill=colors[2], width=self.ring_thickness)
            draw.arc(bbox, 0, 90, fill=colors[3], width=self.ring_thickness)

            mask = Image.new("L", size, 0)
            draw_mask = ImageDraw.Draw(mask)
            draw_mask.ellipse([(0, 0), size], fill=255)
            
            avatar = avatar.resize(size, Image.Resampling.LANCZOS)
            base.paste(avatar, (self.ring_thickness, self.ring_thickness), mask)
            
            self.profile_image = ctk.CTkImage(light_image=base, dark_image=base, size=final_size_tuple)
        
        except Exception as e:
            print(f"Error loading local profile pic: {e}")
            logger.error(f"Error loading local profile pic: {e}")
            size = (self.image_final_size, self.image_final_size)
            placeholder = Image.new("RGBA", size, (0,0,0,0))
            draw = ImageDraw.Draw(placeholder)
            draw.ellipse([(0, 0), size], fill="#333333")
            self.profile_image = ctk.CTkImage(light_image=placeholder, dark_image=placeholder, size=size)

        # --- 2. Main Animated Frame (The "Unit") ---
        self.main_frame = ctk.CTkFrame(self, 
                                       width=self.start_width,
                                       height=self.frame_height, 
                                       corner_radius=35,
                                       fg_color=self.bg_color)
        self.main_frame.pack(side="right", fill="y", pady=0, padx=0)
        self.main_frame.pack_propagate(False)

        # --- 3. Text Label (Created but not placed) ---
        self.text_label = ctk.CTkLabel(self.main_frame, 
                                       text="Parth Sancheti", 
                                       font=("Arial", 16, "bold"),
                                       fg_color="transparent")
        
        # --- 4. Image Label (USE .place()) ---
        self.image_label = ctk.CTkLabel(self.main_frame, text="", image=self.profile_image, fg_color="transparent")
        self.image_label.place(relx=1.0, rely=0.5, x=-(self.start_width/2), anchor="center")

        # Bind events to all components
        self.bind_all("<Enter>", self.on_enter)
        self.bind_all("<Leave>", self.on_leave)
        self.bind_all("<Button-1>", self.open_profile)

    def bind_all(self, event, command):
        self.bind(event, command)
        self.main_frame.bind(event, command)
        self.text_label.bind(event, command)
        self.image_label.bind(event, command)

    def open_profile(self, event=None):
        webbrowser.open(self.profile_url)
        
    def on_enter(self, event=None):
        # Cancel any previous animation
        if self._animation_job:
            self.after_cancel(self._animation_job)
            
        # Place text, it will be invisible until frame expands
        self.text_label.place(relx=0, rely=0.5, x=20, anchor="w") 
        
        self.animate_expand()

    def on_leave(self, event=None):
        # Cancel any previous animation
        if self._animation_job:
            self.after_cancel(self._animation_job)
            
        self.text_label.place_forget()
        
        self.animate_collapse()

    # --- NEW: Step-Based Animation Functions ---
    def animate_expand(self):
        self.current_width += self.width_step
        if self.current_width > self.end_width:
            self.current_width = self.end_width
            
        self.main_frame.configure(width=self.current_width)
        
        if self.current_width < self.end_width:
            self._animation_job = self.after(self.animation_delay, self.animate_expand)
        else:
            self._animation_job = None

    def animate_collapse(self):
        self.current_width -= self.width_step
        if self.current_width < self.start_width:
            self.current_width = self.start_width
            
        self.main_frame.configure(width=self.current_width)
        
        if self.current_width > self.start_width:
            self._animation_job = self.after(self.animation_delay, self.animate_collapse)
        else:
            self._animation_job = None

    def update_theme(self):
        mode = ctk.get_appearance_mode().lower()
        self.bg_color = DARK_FG if mode == "dark" else LIGHT_FG
        self.main_frame.configure(fg_color=self.bg_color)
def generate_html_report(success, failure):
    total = success + failure
    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>WabulkXpress Messaging Analytics</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<style>
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{ font-family: 'Arial', sans-serif; color: #e0e0e0; position: relative; min-height: 100vh; overflow: hidden; padding: 20px; }}
    body::before {{ content: ""; background: url("bin/bg.jpg") no-repeat center center fixed; background-size: cover; filter: blur(8px); position: absolute; top: 0; left: 0; right: 0; bottom: 0; z-index: -2; }}
    body::after {{ content: ""; position: absolute; top: 0; left: 0; right: 0; bottom: 0; background-color: rgba(0, 0, 0, 0.6); z-index: -1; }}
    .container {{ display: flex; align-items: flex-start; justify-content: space-between; padding: 30px; gap: 20px; border-radius: 10px; background-color: rgba(30, 30, 30, 0.9); max-width: 1200px; margin: auto; box-shadow: 0 0 20px rgba(0,0,0,0.5); transition: transform 0.3s ease; }}
    .container:hover {{ transform: scale(1.02); }}
    .info {{ flex: 1; padding: 20px; background: rgba(0, 0, 0, 0.3); border-radius: 10px; margin-right: 20px; }}
    .info h2 {{ margin-bottom: 15px; }}
    .info p {{ margin-bottom: 10px; line-height: 1.5; }}
    .chart-container {{ flex: 1; position: relative; max-width: 400px; background: rgba(0, 0, 0, 0.3); padding: 15px; border-radius: 10px; box-shadow: 0 0 15px rgba(0,0,0,0.4); transition: box-shadow 0.3s ease; }}
    .chart-container:hover {{ box-shadow: 0 0 25px rgba(0,0,0,0.6); }}
    h1 {{ margin-bottom: 20px; text-align: center; font-size: 2em; }}
    button {{ padding: 10px 20px; border: none; border-radius: 5px; background-color: #0078D7; color: #fff; cursor: pointer; transition: background-color 0.3s ease, transform 0.3s ease; margin-top: 15px; }}
    button:hover {{ background-color: #005fa3; transform: scale(1.05); }}
    @media (max-width: 768px) {{ .container {{ flex-direction: column; }} .chart-container {{ margin-top: 20px; max-width: 100%; }} }}
</style>
</head>
<body onload="window.focus();">
<h1>WabulkXpress Messaging Analytics</h1>
<div class="container">
    <div class="info">
        <h2>Message Summary</h2>
        <p>Total Messages: <strong id="totalCount"></strong></p>
        <p>Success: <strong id="successCount"></strong></p>
        <p>Failure: <strong id="failureCount"></strong></p>
        <button onclick="window.close();">Close Report</button>
    </div>
    <div class="chart-container">
        <canvas id="pieChart"></canvas>
    </div>
</div>
<script>
    const total = {total};
    const success = {success};
    const failure = {failure};
    document.getElementById('totalCount').textContent = total;
    document.getElementById('successCount').textContent = success;
    document.getElementById('failureCount').textContent = failure;
    const ctx = document.getElementById('pieChart').getContext('2d');
    const data = {{
        labels: ['Success', 'Failure'],
        datasets: [{{
            data: [success, failure],
            backgroundColor: ['#4CAF50', '#F44336'],
            borderColor: ['#2E7D32', '#C62828'],
            borderWidth: 2,
        }}]
    }};
    const options = {{
        cutout: '70%',
        responsive: true,
        plugins: {{
            legend: {{
                position: 'bottom',
                labels: {{ color: '#e0e0e0' }}
            }}
        }}
    }};
    new Chart(ctx, {{ type: 'doughnut', data: data, options: options }});
</script>
</body>
</html>"""
    report_path = os.path.join(os.getcwd(), "Report.html")
    try:
        with open(report_path, "w", encoding="utf-8") as f:
            f.write(html_content)
        webbrowser.open("file:///" + report_path)
    except Exception as e:
        print(f"Error generating/opening HTML report: {e}")
class WabulkXpressApp(ctk.CTkFrame):       
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)
        
        # --- FIX 1: Define self.master FIRST ---
        self.master = master 
        
        # --- FIX 2: This block MUST come BEFORE creating self.title_bar ---
        # Get theme colors to set the title bar color correctly
        mode = ctk.get_appearance_mode().lower()
        self._set_title_bar_colors(mode) # This creates self.title_bar_color
        
        # --- Create the Custom Title Bar ---
        # This line will now work
        self.title_bar = ctk.CTkFrame(self.master, height=45, corner_radius=10, fg_color=self.title_bar_color)
        self.title_bar.pack(side="top", fill="x", padx=0, pady=(0, 10))
        self.title_bar.pack_propagate(False) # <-- ADD THIS LINE BACK

        # --- Create the Main App Frames ---
        self.animation_active = False
        self.ai_start_time = 0
        # 20-step Blue/Purple "Pulse" (2 colors, 20 shades)
        self.gradient_colors = [
            "#4285F4", "#5077ED", "#5E69E6", "#6C5CDF", "#7A4ED8",
            "#8841D1", "#9633CA", "#A426C3", "#B218BC", "#C00BB5",
            "#B218BC", "#A426C3", "#9633CA", "#8841D1", "#7A4ED8",
            "#6C5CDF", "#5E69E6", "#5077ED", "#4285F4", "#3473E7"
        ]
        self.attachments = {"Picture": None, "Video": None, "Document": None, "Any": None}
        self.all_hints = []
        self.find_last_index = "1.0"
        self.excel_find_last_pos = {"row": 0, "col_idx": -1}
        self.custom_image_enabled = False
        self.excel_data = []
        self.sending = False
        self.undo_stack = []
        self.redo_stack = []
        self.schedule_time = None
        self.first_cycle = True
        self.last_action = None
        
        self.sidebar = ctk.CTkFrame(self, width=250, corner_radius=15, fg_color=DARK_FG) # Theme
        self.sidebar.pack(side="left", fill="y", padx=10, pady=10)
        
        self.header = ctk.CTkFrame(self, height=120, corner_radius=15, fg_color=DARK_FG) # Theme
        self.header.pack(side="top", fill="x", padx=10, pady=(0,0)) # No vertical padding        
        
        # --- FIX 3: Define self.main_area BEFORE packing it ---
        self.main_area = ctk.CTkFrame(self, corner_radius=15, fg_color="transparent") # Theme
        self.main_area.pack(side="right", fill="both", expand=True, padx=10, pady=10)
        
        # --- Create Widgets ---
        self.create_sidebar()
        self.create_header()
        self.create_main_area()
        self._create_custom_title_bar() # Call the function to populate the title bar
        
        self.gemini_api_key = "AIzaSyAyy-ROEqCBloi8pglFZ5WkBuyPZmVXrzE"  # Fallback (replace with actual key or handle gracefully)
        self.gemini_api_url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={self.gemini_api_key}"
        self.ai_prompts = {
            "Reframe": "Rephrase the following message in a single cohesive paragraph with no extra disclaimers or stars:",
            "Emoji": "Rewrite the following message with a few relevant emojis, keeping it concise:",
            "Professional": "Rewrite the following message in a polite, professional tone, no extra disclaimers or bullet points:",
            "Funny": "Rewrite the following message with a light, humorous style, no extra disclaimers or bullet points:",
            "Ask AI": "Ask AI: Please answer the following message without adding any extra formatting or stars:",
            "Translate": "Please translate the following message into {lang}, ensuring that you preserve the original formatting exactly no extra disclaimers or stars or Any othermessage from your side only the text translated:",
        }
        self.schedule_background_update_check()
        self.refresh_icons()
        initial_mode = ctk.get_appearance_mode() # Gets "Light" or "Dark"
        self.apply_theme(initial_mode)
        # --- NEW LICENSE CHECK & THREAD FIX ---
        if not self.validate_local_license():
            # License is missing, invalid, or expired. Show the popup synchronously.
            FirstRunPopup(self, self.first_run_closed).wait_window()
        else:
            # License is valid, just log welcome.
            self.first_run_closed()
            
        # Schedule the background update check *after* initialization and the first run popup close.
        # This ensures the mainloop is running when the thread starts.
        self.after(500, self.schedule_background_update_check)
        # --- END NEW LICENSE CHECK & THREAD FIX ---
    # --- --------------------------------- ---
    # --- CUSTOM TITLE BAR FUNCTIONS      ---
    # --- --------------------------------- ---

    def _create_custom_title_bar(self):
        """Populates the custom title bar frame with slightly bigger icon and text."""
        
        # --- App Icon ---
        try:
            icon_path = TITLE_ICON_PATH
            if os.path.exists(icon_path):
                # *** FIX: Increase icon size from 20x20 to 24x24 ***
                img = Image.open(icon_path).resize((24, 24), Image.Resampling.LANCZOS)
                self.app_icon_image = ctk.CTkImage(light_image=img, dark_image=img, size=(24, 24))
                self.icon_label = ctk.CTkLabel(self.title_bar, image=self.app_icon_image, text="", fg_color="transparent")
                self.icon_label.pack(side="left", padx=(15, 10), pady=5)
            else:
                self.icon_label = None
        except Exception as e:
            print(f"Error loading title bar icon: {e}")
            self.icon_label = None

        # --- Title Label ---
        # *** FIX: Increase font size from 14 to 16 ***
        self.title_label = ctk.CTkLabel(self.title_bar, text="WabulkXpress", font=("Segoe UI", 16, "bold"), text_color=self.title_text_color, anchor="w")
        self.title_label.pack(side="left", fill="x", expand=True, pady=5, padx=5)

        # --- Window Control Buttons (Right to Left order) ---
        
        button_width = 50
        button_height = 45 
        button_pady = 0    
        
        # Close Button
        self.close_button = ctk.CTkButton(
            self.title_bar, text="✕", 
            width=button_width, height=button_height, corner_radius=0, 
            fg_color="transparent", 
            hover_color=self.close_hover_color, 
            text_color=self.title_text_color, 
            font=("Segoe UI Symbol", 12),
            command=self.on_close
        )
        self.close_button.pack(side="right", padx=0, pady=button_pady) 

        self.close_button.bind("<Enter>", 
            lambda e: self.close_button.configure(
                fg_color=self.close_hover_color, 
                text_color="white"
            )
        )
        self.close_button.bind("<Leave>", 
            lambda e: self.close_button.configure(
                fg_color="transparent", 
                text_color=self.title_text_color
            )
        )

        # Maximize/Restore Button
        self._is_maximized = False
        self._restore_geometry = ""
        self.maximize_button = ctk.CTkButton(
            self.title_bar, text="☐", 
            width=button_width, height=button_height, corner_radius=0, 
            fg_color="transparent", hover_color=self.button_hover_color,
            text_color=self.title_text_color, font=("Segoe UI Symbol", 11),
            command=self._toggle_maximize
        )
        self.maximize_button.pack(side="right", padx=0, pady=button_pady) 

        # Minimize Button
        self.minimize_button = ctk.CTkButton(
            self.title_bar, 
            text="−", 
            width=button_width, height=button_height, corner_radius=0, 
            fg_color="transparent", hover_color=self.button_hover_color,
            text_color=self.title_text_color, font=("Segoe UI Symbol", 11, "bold"),
            command=self._minimize_window
        )
        self.minimize_button.pack(side="right", padx=0, pady=button_pady) 

        # --- Drag Functionality ---
        self._offset_x = 0
        self._offset_y = 0
        widgets_to_bind = [self.title_bar, self.title_label]
        if self.icon_label:
             widgets_to_bind.append(self.icon_label)

        for widget in widgets_to_bind:
             if widget:
                widget.bind("<ButtonPress-1>", self._start_move)
                widget.bind("<ButtonRelease-1>", self._stop_move)
                widget.bind("<B1-Motion>", self._do_move)
                widget.bind("<Double-Button-1>", self._toggle_maximize)
    
    def _minimize_window(self):
        """Minimizes the main window using ctypes for borderless window."""
        if "win" not in sys.platform:
            self.master.iconify()
            return
        try:
            hwnd = ctypes.windll.user32.GetParent(self.master.winfo_id())
            ctypes.windll.user32.ShowWindow(hwnd, 6) # SW_MINIMIZE
        except Exception as e:
            print(f"Error minimizing window: {e}")
            self.master.iconify()

    def _toggle_maximize(self, event=None):
        """Maximizes or restores the main window."""
        if self._is_maximized:
            self.master.geometry(self._restore_geometry)
            self.maximize_button.configure(text="☐")
            self._is_maximized = False
        else:
            self._restore_geometry = self.master.geometry()
            screen_width = self.master.winfo_screenwidth()
            screen_height = self.master.winfo_screenheight()
            taskbar_height = 40
            max_height = screen_height - taskbar_height
            self.master.geometry(f"{screen_width}x{max_height}+0+0")
            self.maximize_button.configure(text="❐")
            self._is_maximized = True

    def _start_move(self, event):
        if self._is_maximized:
            return
        self._offset_x = event.x
        self._offset_y = event.y

    def _stop_move(self, event):
        self._offset_x = 0
        self._offset_y = 0

    def _do_move(self, event):
        if self._is_maximized:
            return
        new_x = self.master.winfo_x() + event.x - self._offset_x
        new_y = self.master.winfo_y() + event.y - self._offset_y
        self.master.geometry(f"+{new_x}+{new_y}")

    def _set_title_bar_colors(self, mode):
        """Sets the colors for the title bar and its buttons."""
        if mode == "dark":
            self.title_bar_color = "#000000" # Pitch Black
            self.title_text_color = "#FFFFFF"
            self.button_hover_color = DARK_INNER
            self.close_hover_color = "#C42B1C" # Red hover
        else:
            # --- THIS BLOCK WAS MISSING ---
            self.title_bar_color = LIGHT_FG  # Use the theme's light frame color (F5F5F5)
            self.title_text_color = "#000000" # Black text
            self.button_hover_color = LIGHT_INNER # (EAEAEA)
            self.close_hover_color = "#E81123" # Standard Windows light-mode close red

    def open_find_popup(self, event=None):
        FindPopup(self, target_textbox=self.message_text)
        
    def open_excel_find_popup(self, event=None):
        TableFindPopup(self, target_table=self.excel_table)

    def first_run_closed(self):
        self.log_live("Welcome to WabulkXpress!")
        self.master.lift()
        self.master.focus_force()

    def validate_local_license(self):
        if not os.path.exists(LICENSE_FILE):
            logger.info("license.dat not found. Showing FirstRunPopup.")
            return False
            
        try:
            with open(LICENSE_FILE, 'r') as f:
                data = json.load(f)
                
            expiry_date_str = data.get('expiry_date')
            key = data.get('key')
            
            if not expiry_date_str or not key:
                raise ValueError("License file is incomplete.")
                
            expiry_date = datetime.strptime(expiry_date_str, "%Y-%m-%d").date()
            today = datetime.now().date()
            
            if today > expiry_date:
                logger.warning(f"License for key {key} expired on {expiry_date_str}.")
                messagebox.showerror("License Expired", 
                                     f"Your license for key {key} expired on {expiry_date_str}.\n"
                                     "Please enter a new product key.")
                return False # Show popup
            else:
                days_left = (expiry_date - today).days
                self.log_live(f"✅ License for key {key} is valid. Expires in {days_left} days.")
                logger.info(f"License valid for {days_left} more days.")
                return True # License is valid!
                
        except json.JSONDecodeError:
            logger.error("Failed to decode license.dat. File is corrupt.")
            messagebox.showerror("License Error", "Your license file (license.dat) is corrupt. Please re-enter your key.")
        except Exception as e:
            logger.error(f"Error reading license file: {e}", exc_info=True)
            messagebox.showerror("License Error", f"An error occurred reading your license file: {e}\nPlease re-enter your key.")
            
        try:
            os.remove(LICENSE_FILE)
        except Exception as e:
            logger.error(f"Could not remove corrupt license file: {e}")
            
        return False


    def schedule_background_update_check(self):
        try:
            last_check_time = 0.0
            if os.path.exists(UPDATE_CHECK_FILE):
                with open(UPDATE_CHECK_FILE, 'r') as f:
                    content = f.read().strip()
                    if content:
                        last_check_time = float(content)

            seven_days_ago = time.time() - (2 * 24 * 60 * 60)

            if last_check_time < seven_days_ago:
                self.log_live("Performing automatic update check (last check > 2 days ago)...")                # Run the check in a separate thread to avoid blocking startup
                threading.Thread(target=self.check_for_update_background, daemon=True).start()
            else:
                last_check_date = datetime.fromtimestamp(last_check_time).strftime('%Y-%m-%d %H:%M')
                print(f"DEBUG: Skipping automatic update check. Last check was at {last_check_date}") # Console log

        except Exception as e:
            self.log_live(f"⚠️ Error scheduling background update check: {e}")
            print(f"DEBUG: Error reading update check file: {e}") # Console log


    def check_for_update_background(self):
        update_available = False
        latest_version_tag = "0"
        error_message = None # Store potential error message

        try:
            response = requests.get(GITHUB_API_URL, timeout=15)
            response.raise_for_status()
            data = response.json()
            latest_version_tag = data.get("tag_name", "0")

            if float(latest_version_tag) > float(CURRENT_VERSION):
                update_available = True
            else:
                log_msg = f"Automatic check: You are using the latest version ({CURRENT_VERSION})."
                self.after(0, lambda msg=log_msg: self.log_live(msg))

            # Update timestamp only after successful check
            with open(UPDATE_CHECK_FILE, 'w') as f:
                f.write(str(time.time()))
            print(f"DEBUG: Updated last update check timestamp to {datetime.now()}")

        except Exception as e:
            error_message = f"⚠️ Automatic update check failed: {e}"
            print(f"DEBUG: Background update check failed: {e}")
            self.after(0, lambda msg=error_message: self.log_live(msg))
            return # Exit if check failed

        # --- FIX: Use the Custom UpdatePopup for background update notification ---
        if update_available:
            log_msg = f"Automatic check: Update available: {latest_version_tag} (Current: {CURRENT_VERSION})"
            self.after(0, lambda msg=log_msg: self.log_live(msg))
            
            # **REPLACING the old messagebox call with the custom UI launch**
            self.after(0, self.open_update_prompt_ui)

    def open_update_prompt_ui(self):
        """
        FIX: Launches the custom UpdatePopup UI on the main thread.
        This function is called when an update is detected in the background.
        """
        # The UpdatePopup automatically starts the check, so we launch it directly.
        UpdatePopup(self)

    def prompt_for_update(self, latest_version_tag):
        """Shows the update prompt messagebox on the main thread."""
        answer = messagebox.askyesno(
            "Update Available",
            f"A new version ({latest_version_tag}) is available!\n"
            f"Current version: {CURRENT_VERSION}\n\n"
            "Would you like to visit the releases page to download it?"
        )
        if answer:
            webbrowser.open(GITHUB_RELEASES_URL)
            
    def refresh_icons(self):
        self.update_button.configure(
            image=self.get_icon("update"),
            fg_color="transparent",
            hover_color="#333333",
            corner_radius=0
        )
        self.theme_toggle_button.configure(
            image=self.get_icon("dark"),
            fg_color="transparent",
            hover_color="#333333",
            corner_radius=0
        )
        
    def create_sidebar(self):
        if os.path.exists(LOGO_PATH):
            img = Image.open(LOGO_PATH).resize((200,200), Image.Resampling.LANCZOS)
            self.sidebar_logo = ctk.CTkImage(img, size=(200,200))
            self.logo_label = ctk.CTkLabel(self.sidebar, image=self.sidebar_logo, text="")
            self.logo_label.pack(pady=(20,10))
        else:
            self.logo_label = ctk.CTkLabel(self.sidebar, text="Logo Missing")
            self.logo_label.pack(pady=(20,20))
        
        center_frame = ctk.CTkFrame(self.sidebar, corner_radius=0, fg_color="transparent") # Theme
        center_frame.pack(expand=True, fill="both")
        center_frame.grid_columnconfigure(0, weight=1)
        center_frame.grid_columnconfigure(1, weight=1)
        
        # Start button
        self.start_stop_button = CustomRoundedButton(
            center_frame, text="Start", corner_radius=100, height=50, width=100, command=self.toggle_sending
        )
        self.start_stop_button.grid(row=1, column=0, padx=0, pady=20)
        # Schedule button with an arrow icon and integrated into one button
        arrow_path = os.path.join(BIN_FOLDER, "down_arrow_dark.png")
        down_arrow_icon = None
        if os.path.exists(arrow_path):
            arrow_img = Image.open(arrow_path).resize((25, 25), Image.Resampling.LANCZOS)
            down_arrow_icon = ctk.CTkImage(arrow_img, size=(25, 25))
        self.schedule_button = CustomRoundedButton(
            center_frame,
            text="",
            image=down_arrow_icon,
            corner_radius=20,
            height=50,
            width=50,
            command=self.open_schedule_popup
        )
        self.schedule_button.grid(row=1, column=1, padx=0, pady=2)  # Reduced padx to 0 for smaller gap between start and schedule
        # --- ADD THIS LINE ---
        self.all_hints.append(HoverHint(self.schedule_button, 
                  "Schedule your messages to be sent at a specific time. The app must remain open.", 
                  os.path.join(BIN_FOLDER, "hint_schedule.png")))    # Login button with full-rounded corners, bigger font, and green default color
        self.login_button = CustomRoundedButton(
            center_frame, text="Login", corner_radius=100, height=50, width=120, command=self.launch_whatsapp_beta
        )
        self.login_button.grid(row=2, column=0, columnspan=2, padx=5, pady=(10, 5))  # Reduced pady for smaller gap to start/schedule buttons
        self.live_alerts = ctk.CTkTextbox(
            self.sidebar, 
            height=250, 
            corner_radius=10, 
            font=ctk.CTkFont(size=16, weight="normal"),  # Bigger font for live alerts and all text
            padx=10,  # Added padding
            pady=10,   # Added padding
            fg_color=DARK_INNER, # Theme
            border_width=0 # Theme
        )
        self.live_alerts.pack(side="bottom", pady=20, padx=10) # Added horizontal padding
        self.live_alerts.insert("0.0", "Live Alerts:\n")
        self.live_alerts.configure(state="disabled")

    def is_logged_in(self):
        """Checks if the selenium session folder exists and is not empty."""
        if not os.path.exists(SESSION_DIR) or not os.listdir(SESSION_DIR):
            self.log_live("❗️ Login required. Please use the Login button.")
            messagebox.showerror("Login Required", "You must log in to WhatsApp Web first. Please use the 'Login' button.")
            return False
        return True
        
    def open_schedule_popup(self):
        """
        Opens the scheduling pop-up after checking for login status, contacts, and content.
        Allows scheduling if a message, attachment, OR custom image is present.
        """
        if not self.is_logged_in():
            return  # Stop if not logged in

        contact_data = self.excel_table.get_data()
        if not contact_data:
            messagebox.showerror("Error", "No phone numbers loaded. Please add at least one contact to schedule.")
            return  # Stop if no data

        message_content = self.message_text.get("0.0", "end-1c").strip()
        single_attachment_present = "Any" in self.attachments and self.attachments["Any"]
        
        # Check if NO content is available (neither message, nor attachment, nor custom image)
        if not message_content and not self.custom_image_enabled and not single_attachment_present:
            messagebox.showerror(
                "Error", 
                "No content to send. Please add a message, attachment, or custom image to schedule."
            )
            return  # Stop if no content

        # Proceed to open the schedule popup
        SchedulePopup(self, self.set_schedule_time)

    def create_header(self):
        # Double the height of the header box (assuming default ~40, set to 80)
        
        
        # Use grid for better vertical centering control
        self.header.grid_columnconfigure(0, weight=1)
        self.header.grid_columnconfigure(1, weight=0)
        self.header.grid_rowconfigure(0, weight=1)
        
        # Left side: Vertical stack of labels
        left_frame = ctk.CTkFrame(self.header, fg_color="transparent")
        left_frame.grid(row=0, column=0, sticky="nw", padx=20, pady=10)
        left_frame.grid_rowconfigure(0, weight=0)
        left_frame.grid_rowconfigure(1, weight=0)
        
        self.welcome_label = ctk.CTkLabel(
            left_frame, 
            text="Welcome!", 
            font=("Arial", 30, "bold")  # Bigger font
        )
        self.welcome_label.grid(row=0, column=0, sticky="w", pady=(0, 5))
        
        self.subtitle_label = ctk.CTkLabel(
            left_frame, 
            text="To WaBulkExpress", 
            font=("Arial", 16)  # Smaller font
        )
        self.subtitle_label.grid(row=1, column=0, sticky="w")
        
        # Right side: Buttons vertically centered and bigger
        right_frame = ctk.CTkFrame(self.header, fg_color="transparent")
        right_frame.grid(row=0, column=1, sticky="e", padx=10, pady=10)
        right_frame.grid_rowconfigure(0, weight=1)
        right_frame.grid_rowconfigure(2, weight=1)
        right_frame.grid_rowconfigure(1, weight=0)
        right_frame.grid_columnconfigure(0, weight=0)
        right_frame.grid_columnconfigure(1, weight=0)
        right_frame.grid_columnconfigure(2, weight=0)
        
        self.theme_toggle_button = ctk.CTkButton(
            right_frame,
            text="",
            width=70,  # Bigger
            height=70,  # Bigger and taller for bigger icons and vertical centering
            command=self.toggle_theme,
        )
        self.theme_toggle_button.grid(row=1, column=0, padx=(0, 5), sticky="e")
        
        self.update_button = ctk.CTkButton(
            right_frame,
            text="",
            width=70,  # Bigger
            height=70,  # Bigger and taller
            command=self.check_for_update,
        )
        self.update_button.grid(row=1, column=1, padx=5, sticky="e")
        
        self.profile_button = ProfileButton(right_frame)
        self.profile_button.grid(row=1, column=2, padx=(5, 0), sticky="e")
        
        # Call refresh_icons (which was part of the original create_header)
        self.refresh_icons()

    def get_icon(self, icon_name):
        mode = ctk.get_appearance_mode().lower()
        if icon_name in ["github", "update"]:
            file_name = f"{icon_name}.png" if mode == "light" else f"{icon_name}_dark.png"
        elif icon_name == "dark":
            file_name = "dark.png" if mode == "light" else "light.png"
        else:
            file_name = f"{icon_name}.png"
        icon_path = os.path.join(os.getcwd(), "bin", file_name)
        size = (40,40)
        if os.path.exists(icon_path):
            img = Image.open(icon_path).resize(size, Image.Resampling.LANCZOS)
        else:
            img = Image.new("RGB", size, "#DBDBDB" if mode=="light" else "#2B2B2B")
        return ctk.CTkImage(img, size=size)
        
    def create_main_area(self):
        # Changed weight to 1:3 to give Excel area more space
        self.main_area.columnconfigure(0, weight=1) 
        self.main_area.columnconfigure(1, weight=3)
        
        self.message_frame = ctk.CTkFrame(self.main_area, corner_radius=15, fg_color=DARK_FG) # Theme
        self.message_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        self.create_message_area(self.message_frame)
        
        self.excel_frame = ctk.CTkFrame(self.main_area, corner_radius=15, fg_color=DARK_FG) # Theme
        self.excel_frame.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)
        self.create_excel_area(self.excel_frame)
        
        self.main_area.rowconfigure(0, weight=1)


    def create_message_area(self, parent):
            # --- Configure Parent Grid ---
            parent.grid_rowconfigure(3, weight=1) # Row 3 (text_area_frame) expands
            parent.grid_columnconfigure(0, weight=1) # Full width

            # --- Top Buttons (Side-by-side) ---
            top_button_frame = ctk.CTkFrame(parent, corner_radius=10, fg_color="transparent")
            top_button_frame.grid(row=0, column=0, sticky="ew", pady=(10, 5), padx=5) # Use grid
            top_button_frame.columnconfigure((0, 1), weight=1) # Make columns expand equally

            self.attachment_btn = ctk.CTkButton(
                top_button_frame,
                text="Select Attachment",
                corner_radius=100,
                height=40,
                command=self.handle_attachment,
                font=("Arial", 16, "bold"),
                fg_color="green",
                hover_color="#45A049"
            )
            self.attachment_btn.grid(row=0, column=0, padx=(10, 5), pady=5, sticky="ew") # Use grid

            self.custom_image_btn = ctk.CTkButton(
                top_button_frame,
                text="Custom Image Namer",
                corner_radius=100,
                command=self.open_custom_image_window,
                height=40,
                font=("Arial", 16, "bold"),
                fg_color="green",
                hover_color="#45A049"
            )
            self.custom_image_btn.grid(row=0, column=1, padx=(5, 10), pady=5, sticky="ew") # Use grid
            self.all_hints.append(HoverHint(self.custom_image_btn, "Automatically places the receiver’s name onto your custom image template — perfect for personalized visual messages!", os.path.join(os.getcwd(), "bin", "woi_ci.png")))

            # --- Formatting Buttons (Two Rows) ---
            fmt_frame = ctk.CTkFrame(parent, corner_radius=10, fg_color="transparent") # Theme
            fmt_frame.grid(row=1, column=0, sticky="ew", pady=5, padx=5) # Use grid
            # Configure columns for spacing in the first row
            fmt_frame.columnconfigure((0, 1, 2, 3), weight=0) # Left buttons don't expand
            fmt_frame.columnconfigure(4, weight=1) # Spacer expands
            fmt_frame.columnconfigure((5, 6), weight=0) # Right buttons don't expand

            # Style dictionary for small formatting buttons (width removed)
            btn_style_small = {
                "fg_color": "green",
                "hover_color": "#45A049",
                "corner_radius": 100,
                "font": ("Arial", 12, "bold"), # Smaller font
                # "width": 35, # REMOVED THIS LINE
                "height": 35 # Smaller height
            }

            # Row 0: Formatting Left, Undo/Redo Right
            self.bold_btn = ctk.CTkButton(fmt_frame, text="B", command=lambda: self.apply_formatting("*"), **btn_style_small, width=35) # Specify width here
            self.bold_btn.grid(row=0, column=0, padx=2, pady=2)

            self.italic_btn = ctk.CTkButton(fmt_frame, text="I", command=lambda: self.apply_formatting("_"), **btn_style_small, width=35) # Specify width here
            self.italic_btn.grid(row=0, column=1, padx=2, pady=2)

            self.strike_btn = ctk.CTkButton(fmt_frame, text="S", command=lambda: self.apply_formatting("~"), **btn_style_small, width=35) # Specify width here
            self.strike_btn.grid(row=0, column=2, padx=2, pady=2)

            self.mono_btn = ctk.CTkButton(fmt_frame, text="Code", command=lambda: self.apply_formatting("```"), **btn_style_small, width=40) # Specify width here
            self.mono_btn.grid(row=0, column=3, padx=2, pady=2)

            self.undo_btn = ctk.CTkButton(fmt_frame, text="Undo", command=self.undo, **btn_style_small, width=50) # Specify width here
            self.undo_btn.grid(row=0, column=5, padx=2, pady=2)

            self.redo_btn = ctk.CTkButton(fmt_frame, text="Redo", command=self.redo, **btn_style_small, width=50) # Specify width here
            self.redo_btn.grid(row=0, column=6, padx=(2,10), pady=2) # Add right padding

            # Row 1: Placeholders, evenly spaced using grid directly on fmt_frame
            placeholder_btn_style = {
                "fg_color": "green",
                "hover_color": "#45A049",
                "corner_radius": 100,
                "font": ("Arial", 12, "bold"), # Smaller font
                "height": 35 # Smaller height
            }
            # Create a sub-frame for row 1 buttons to manage spacing better
            row1_fmt_frame = ctk.CTkFrame(fmt_frame, fg_color="transparent")
            row1_fmt_frame.grid(row=1, column=0, columnspan=7, sticky="ew", pady=(5,0))
            row1_fmt_frame.columnconfigure((0,1,2), weight=1) # Even spacing within this sub-frame

            self.username_btn = ctk.CTkButton(row1_fmt_frame, text="User", command=lambda: self.insert_placeholder("{User_Name}"), **placeholder_btn_style)
            self.username_btn.grid(row=0, column=0, padx=5, pady=0, sticky="ew")
            self.all_hints.append(HoverHint(self.username_btn, "Inserts {User_Name} placeholder", os.path.join(os.getcwd(), "bin", "woi_un.png")))

            self.custom1_btn = ctk.CTkButton(row1_fmt_frame, text="Custom 1", command=lambda: self.insert_placeholder("{Custom1}"), **placeholder_btn_style)
            self.custom1_btn.grid(row=0, column=1, padx=5, pady=0, sticky="ew")
            self.all_hints.append(HoverHint(self.custom1_btn, "Inserts {Custom1} placeholder (from optional import column)", os.path.join(os.getcwd(), "bin", "woi_un.png")))
            self.custom2_btn = ctk.CTkButton(row1_fmt_frame, text="Custom 2", command=lambda: self.insert_placeholder("{Custom2}"), **placeholder_btn_style)
            self.custom2_btn.grid(row=0, column=2, padx=5, pady=0, sticky="ew")
            self.all_hints.append(HoverHint(self.custom2_btn, "Inserts {Custom2} placeholder (from optional import column)", os.path.join(os.getcwd(), "bin", "woi_un.png")))


            # --- Message Area ---
            msg_label = ctk.CTkLabel(parent, text="Message:", font=("Arial", 18, "bold"))
            msg_label.grid(row=2, column=0, sticky="w", padx=10, pady=(10, 5)) # Use grid

            # Get theme-aware color for the "invisible" border
            mode = ctk.get_appearance_mode().lower()
            # PARENT frame uses the main FG color
            text_frame_bg = DARK_FG if mode == "dark" else LIGHT_FG
            # CHILD text box uses the INNER color
            inner_bg = DARK_INNER if mode == "dark" else LIGHT_INNER

            self.text_area_frame = ctk.CTkFrame(
                parent, 
                corner_radius=24, # Outer frame has a 24px radius
                fg_color=text_frame_bg, # Use the main FG color
                border_width=0  # <-- CHANGED: We will animate fg_color instead
            )
            self.text_area_frame.grid(row=3, column=0, sticky="nsew", padx=10, pady=(0,10)) # Use grid

            # --- THIS IS THE CORNER FIX: Use .grid() to "inset" the text box ---
            # Configure the grid *inside* the text_area_frame
            self.text_area_frame.grid_rowconfigure(0, weight=1)
            self.text_area_frame.grid_columnconfigure(0, weight=1)

            self.message_text = ctk.CTkTextbox(
                self.text_area_frame, # Child of the animating frame
                height=100,
                corner_radius=16, # Inner radius = 24 (parent) - 8 (border)
                fg_color=inner_bg, # Use the INNER color
                border_width=0,
                padx=15, # INTERNAL text padding
                pady=15, # INTERNAL text padding
                font=("Arial", 16) 
            ) 
            # --- End of .grid() fix ---

            self.message_text.grid(row=0, column=0, sticky="nsew", padx=12, pady=12) 
            # --- End of .grid() fix ---

            self.message_text.bind("<Control-z>", self.undo)
            self.message_text.bind("<Control-y>", self.redo)
            self.message_text.bind("<KeyRelease>", self.save_state)
            self.message_text.bind("<Button-3>", self.show_context_menu)
            # ... (inside create_message_area)
            self.message_text.bind("<Button-3>", self.show_context_menu)
            self.message_text.bind("<Control-f>", self.open_find_popup) # <-- ADD THIS LINE

            self.context_menu = tk.Menu(self.message_text, tearoff=0)
            # ... (rest of the method) ...
            self.context_menu = tk.Menu(self.message_text, tearoff=0)
            self.context_menu.add_command(label="Bold", command=lambda: self.apply_formatting("*"))
            self.context_menu.add_command(label="Italic", command=lambda: self.apply_formatting("_"))
            self.context_menu.add_command(label="Strike", command=lambda: self.apply_formatting("~"))
            self.context_menu.add_command(label="Code", command=lambda: self.apply_formatting("```"))
            self.context_menu.add_separator()
            self.context_menu.add_command(label="Insert User", command=lambda: self.insert_placeholder("{User_Name}"))
            self.context_menu.add_command(label="Insert Custom 1", command=lambda: self.insert_placeholder("{Custom1}"))
            self.context_menu.add_command(label="Insert Custom 2", command=lambda: self.insert_placeholder("{Custom2}"))
            self.context_menu.add_separator()
            self.context_menu.add_command(label="Undo", command=self.undo)
            self.context_menu.add_command(label="Redo", command=self.redo)

            # --- AI Button ---
            ai_icon = None
            if os.path.exists(AI_ICON_PATH):
                try:
                    ai_img = Image.open(AI_ICON_PATH).resize((25,25), Image.Resampling.LANCZOS)
                    ai_icon = ctk.CTkImage(ai_img, size=(25,25))
                except Exception as e:
                    print(f"Error loading AI icon for button: {e}")

            mode = ctk.get_appearance_mode().lower()
            # The button's background (fg_color) should match the main frames (DIFFERENT from text box)
            ai_bg_color = DARK_FG if mode == "dark" else LIGHT_FG
            # The button's hover color should be a slightly different shade
            ai_hover_color = "#333333" if mode == "dark" else "#E0E0E0"

            self.ai_button = ctk.CTkButton(
                self.message_text,  # Place button inside message_text
                text="",
                image=ai_icon,
                fg_color=ai_bg_color,  # Use new base color
                hover_color=ai_hover_color, # Use new hover color
                corner_radius=20,  
                width=50,
                height=50,
                command=self.show_ai_popup
            )
            self.ai_button.place(relx=1.0, rely=1.0, x=-15, y=-15, anchor="se") # Place in corner of text box
            # Set initial bindings for the AI button click effect
            ai_press_color = "#555555" if mode == "dark" else "#CFCFCF" 
            self.ai_button.bind("<ButtonPress-1>", lambda e, c=ai_press_color: self.ai_button.configure(fg_color=c))
            self.ai_button.bind("<ButtonRelease-1>", lambda e, c=ai_hover_color: self.ai_button.configure(fg_color=c))

            # --- Delay Frame ---
            delay_frame = ctk.CTkFrame(parent, corner_radius=10, fg_color="transparent") # Theme
            delay_frame.grid(row=4, column=0, sticky="ew", padx=5, pady=(10, 5)) # Use grid

            delay_font = ("Arial", 14) # Increased font size

            ctk.CTkLabel(delay_frame, text="Min Delay (s)", font=delay_font).pack(side="left", padx=(10, 5))
            self.min_delay_entry = ctk.CTkEntry(delay_frame, placeholder_text="1", corner_radius=10, width=60, fg_color=DARK_INNER, border_width=0, font=delay_font) # Theme & Font
            self.min_delay_entry.insert(0, str(DEFAULT_MIN_DELAY))
            self.min_delay_entry.pack(side="left", padx=5)

            ctk.CTkLabel(delay_frame, text="Max Delay (s)", font=delay_font).pack(side="left", padx=(15, 5))
            self.max_delay_entry = ctk.CTkEntry(delay_frame, placeholder_text="10", corner_radius=10, width=60, fg_color=DARK_INNER, border_width=0, font=delay_font) # Theme & Font
            self.max_delay_entry.insert(0, str(DEFAULT_MAX_DELAY))
            self.max_delay_entry.pack(side="left", padx=5)

# --- ---------------------------------------------------- ---
    # --- NEW HELPER METHODS FOR THE MERGED HEADER/TITLE BAR ---
    # --- ---------------------------------------------------- ---

    def _minimize_window(self):
        """Minimizes the main window using ctypes for borderless window."""
        if "win" not in sys.platform:
            self.master.iconify()
            return
        try:
            hwnd = ctypes.windll.user32.GetParent(self.master.winfo_id())
            ctypes.windll.user32.ShowWindow(hwnd, 6) # SW_MINIMIZE
        except Exception as e:
            print(f"Error minimizing window: {e}")
            self.master.iconify()

    def _toggle_maximize(self, event=None):
        """Maximizes or restores the main window."""
        if self._is_maximized:
            # Restore
            self.master.geometry(self._restore_geometry)
            self.maximize_button.configure(text="☐") # Restore symbol
            self._is_maximized = False
        else:
            # Maximize
            self._restore_geometry = self.master.geometry() # Save current size/pos
            screen_width = self.master.winfo_screenwidth()
            screen_height = self.master.winfo_screenheight()
            taskbar_height = 40
            max_height = screen_height - taskbar_height
            self.master.geometry(f"{screen_width}x{max_height}+0+0")
            self.maximize_button.configure(text="❐") # Maximize symbol
            self._is_maximized = True

    def _start_move(self, event):
        if self._is_maximized:
            return
        self._offset_x = event.x
        self._offset_y = event.y

    def _stop_move(self, event):
        self._offset_x = 0
        self._offset_y = 0

    def _do_move(self, event):
        if self._is_maximized:
            return
        new_x = self.master.winfo_x() + event.x - self._offset_x
        new_y = self.master.winfo_y() + event.y - self._offset_y
        self.master.geometry(f"+{new_x}+{new_y}")

    def _set_header_button_colors(self, mode):
        """Helper to set colors for window controls."""
        if mode == "dark":
            self.title_text_color = "#FFFFFF"
            self.button_hover_color = DARK_INNER
            self.close_hover_color = "#C42B1C"
        else:
            self.title_text_color = "#000000"
            self.button_hover_color = LIGHT_INNER
            self.close_hover_color = "#E81123"


         
    def show_context_menu(self, event):
        try:
            # Update menu theme before showing
            mode = ctk.get_appearance_mode().lower()
            if mode == "dark":
                self.context_menu.configure(bg=DARK_FG, fg="white")
            else:
                self.context_menu.configure(bg=LIGHT_FG, fg="black")
            self.context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.context_menu.grab_release()
            
    def save_state(self, event=None):
        current_text = self.message_text.get("0.0", "end-1c")
        if not self.undo_stack or current_text != self.undo_stack[-1]: # Avoid duplicate states
            self.undo_stack.append(current_text)
            self.redo_stack.clear()
        
    def undo(self, event=None):
        if len(self.undo_stack) > 1: # Need at least one state to revert to
            current_text = self.undo_stack.pop()
            self.redo_stack.append(current_text)
            previous_text = self.undo_stack[-1]
            self.message_text.delete("0.0", "end")
            self.message_text.insert("0.0", previous_text)
            
    def redo(self, event=None):
        if self.redo_stack:
            next_text = self.redo_stack.pop()
            self.undo_stack.append(next_text)
            self.message_text.delete("0.0", "end")
            self.message_text.insert("0.0", next_text)
            
    # New generic method for placeholders
    def insert_placeholder(self, placeholder_text):
        self.save_state()
        self.message_text.insert("insert", placeholder_text)
        self.save_state() # Save state *after* inserting
        
    def apply_formatting(self, symbol):
        self.save_state()
        try:
            sel_start_str = self.message_text.index("sel.first")
            sel_end_str = self.message_text.index("sel.last")
            
            # Get text and delete selection BEFORE inserting to avoid index issues
            text = self.message_text.get(sel_start_str, sel_end_str).strip()
            self.message_text.delete(sel_start_str, sel_end_str) 
            
            formatted = f"{symbol}{text}{symbol}"
            self.message_text.insert(sel_start_str, formatted)

            # Re-select the inserted formatted text (optional, good UX)
            end_index = f"{sel_start_str}+{len(formatted)}c"
            self.message_text.tag_add("sel", sel_start_str, end_index)
            self.message_text.mark_set("insert", end_index)
            
        except tk.TclError: # Handle case where no text is selected
            # Insert the symbols at the cursor position
            cursor_index = self.message_text.index("insert")
            formatted = f"{symbol}{symbol}"
            self.message_text.insert(cursor_index, formatted)
            # Place cursor between the symbols
            self.message_text.mark_set("insert", f"{cursor_index}+{len(symbol)}c")
            
        self.save_state() # Save state *after* formatting

    def handle_attachment(self):
        # Check if an attachment is already selected
        if self.attachments.get("Any"):
            # --- Deselect Logic ---
            self.attachments["Any"] = None
            self.log_live("ℹ️ Attachment deselected.")
            
            # Reset button to "Select" state (green)
            self.attachment_btn.configure(
                text="Select Attachment",
                fg_color="green",
                hover_color="#45A049"
            )
            
            # If the last action was 'attachment', clear it
            if self.last_action == "attachment":
                self.last_action = None
                
        else:
            # --- Select Logic (No attachment) ---
            filetypes = [
                ("All Files", "*.*"),
                ("Image Files", "*.png;*.jpg;*.jpeg;*.gif;*.bmp"),
                ("Document Files", "*.pdf;*.docx;*.doc;*.txt;*.csv;*.xlsx;*.xls"),
                ("Audio/Video Files", "*.mp3;*.wav;*.m4a;*.mp4;*.mkv;*.avi")
            ]
            path = filedialog.askopenfilename(filetypes=filetypes)
            
            if path:
                # Clear other types just in case
                self.attachments["Picture"] = None
                self.attachments["Video"] = None
                self.attachments["Document"] = None
                # Set the new attachment
                self.attachments["Any"] = path
                
                self.log_live(f"📎 Attachment selected: {os.path.basename(path)}")
                self.last_action = "attachment"
                
                # Update button to "Deselect" state (red)
                self.attachment_btn.configure(
                    text="Deselect Attachment",
                    fg_color="#FF4040", # Red color (like stop button)
                    hover_color="#D32F2F"  # Darker red hover
                )
            
    def open_custom_image_window(self):
        if hasattr(self, "excel_table"):
            self.excel_data = self.excel_table.get_data()
        if not self.excel_data:
            messagebox.showerror("Error", "Please load or enter phone data first.")
            return
            
        self.custom_image_enabled = True
        self.last_action = "custom"
        
        # --- This is the FIRST and ONLY call ---
        CustomImageWindow(self, self.excel_data)
        
        # --- New Logic ---
        # 1. Update this button to "Deselect" state (red)
        self.custom_image_btn.configure(
            text="Deselect Custom",
            command=self.deselect_custom_images,  # Point to the new deselect method
            fg_color="#FF4040", # Red color
            hover_color="#D32F2F"
        )
        
        # 2. Deselect the single attachment if one was active
        if self.attachments.get("Any"):
            self.attachments["Any"] = None
            self.log_live("ℹ️ Single attachment deselected (Custom Images enabled).")
            # Reset the *other* button
            self.attachment_btn.configure(
                text="Select Attachment",
                fg_color="green",
                hover_color="#45A049"
            )
        # --- End New Logic ---

        # (The duplicate call at the end has been removed)

    def deselect_custom_images(self):
        """Deselects custom images, cleans the output folder, and resets the button."""
        if not self.custom_image_enabled:
            return  # Do nothing if already deselected

        try:
            if os.path.exists(OUTPUT_IMG_FOLDER):
                shutil.rmtree(OUTPUT_IMG_FOLDER)
                self.log_live(f"ℹ️ Cleared custom images from {OUTPUT_IMG_FOLDER}.")
            else:
                self.log_live("ℹ️ Custom images deselected. (Folder not found)")
            
            # Re-create the folder so it's ready for next time
            os.makedirs(OUTPUT_IMG_FOLDER, exist_ok=True) 

        except PermissionError:
            self.log_live(f"❌ Error: Could not delete {OUTPUT_IMG_FOLDER}. Files may be in use.")
            messagebox.showerror("Error", f"Could not clear custom images. Close any programs using files in {OUTPUT_IMG_FOLDER} and try again.")
            return  # Don't proceed if we failed to delete
        except Exception as e:
            self.log_live(f"❌ Error clearing images: {e}")
            # Still proceed to reset UI
        
        self.custom_image_enabled = False
        if self.last_action == "custom":
            self.last_action = None
        
        # Reset button to "Select" state
        self.custom_image_btn.configure(
            text="Custom Image Namer",
            command=self.open_custom_image_window,  # Point command back to the original
            fg_color="green",
            hover_color="#45A049"
        )
    
        
    def create_excel_area(self, parent):
        top_frame = ctk.CTkFrame(parent, corner_radius=10, fg_color="transparent")
        top_frame.pack(fill="x", padx=10, pady=10)
        
        country_codes = ["+91", "None", "+1", "+44", "+61", "+81", "+49", "+33", "+86", "+7"]
        self.country_code_var = ctk.StringVar(value="+91")
        # Get theme colors for the dropdown
        mode = ctk.get_appearance_mode().lower()
        # Use DARK_BG for pitch black popup
        inner_color = DARK_INNER if mode == "dark" else LIGHT_INNER
        text_color = "white" if mode == "dark" else "black"

        self.country_code_dropdown = ctk.CTkOptionMenu(
            top_frame, 
            values=country_codes, 
            variable=self.country_code_var,
            font=("Arial", 16), # Bigger font
            dropdown_font=("Arial", 16), # Bigger font
            height=40,
            corner_radius=100,                # <-- CHANGED: Full rounded edge
            fg_color="green",                 # Main button background
            button_color="#45A049",           # Color for the arrow button
            button_hover_color="#3E8E41",     # Darker green for arrow hover
            # --- UPDATED LINES ---
            dropdown_fg_color=DARK_BG,        # <-- CHANGED: Pitch black background
            dropdown_hover_color=inner_color, # Use the inner color for hover
            dropdown_text_color=text_color,   # Set appropriate text color
        )
        # Added more space with padx=10
        self.country_code_dropdown.pack(side="left", padx=10, pady=5)
        
        self.import_db_btn = ctk.CTkButton(
            top_frame, 
            text="Import DataBase", 
            corner_radius=10, 
            command=self.open_import_popup,
            font=("Arial", 16, "bold"), # Bigger font
            height=40,
            fg_color="green",
            hover_color="#45A049"
        )
        self.import_db_btn.pack(side="left", padx=5, pady=5)
        # --- ADD THIS LINE ---
        self.all_hints.append(HoverHint(self.import_db_btn, 
                  "Import contacts from Excel or CSV files. You can map columns for phone, name, and two custom fields.", 
                  os.path.join(BIN_FOLDER, "hint_import.png")))
        
        # --- NEW FIND BUTTON ---
        self.find_btn = ctk.CTkButton(
            top_frame,
            text="Find",
            corner_radius=10,
            command=self.open_excel_find_popup,
            font=("Arial", 16, "bold"),
            height=40,
            width=80, # Smaller width
            fg_color="green", # <-- CHANGED
            hover_color="#45A049" # <-- CHANGED
        )
        self.find_btn.pack(side="left", padx=5, pady=5)
        # --- END NEW BUTTON ---
        mode = ctk.get_appearance_mode().lower()
        excel_bg = DARK_FG if mode == "dark" else LIGHT_FG
        
        self.excel_table = ExcelTable(parent, main_app=self, fg_color=excel_bg)
        self.excel_table.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        
        # --- ADD THIS BINDING ---
        # Bind to the scrollable frame itself
        self.excel_table.bind("<Control-f>", self.open_excel_find_popup)



    def parse_row_ranges(self, range_str):
        """Parses a row range string (e.g., '2-10, 15, 21-30') into a set of integers."""
        if not range_str:
            return None # No range specified, process all
            
        selected_rows = set()
        try:
            parts = range_str.split(',')
            for part in parts:
                part = part.strip()
                if not part:
                    continue
                if '-' in part:
                    # It's a range
                    start, end = part.split('-')
                    start_num = int(start.strip())
                    end_num = int(end.strip())
                    if start_num <= 1 or end_num < start_num: # Rows start at 2 (after header)
                        raise ValueError(f"Invalid range '{part}'. Rows must be > 1.")
                    selected_rows.update(range(start_num, end_num + 1))
                else:
                    # It's a single number
                    num = int(part.strip())
                    if num <= 1:
                         raise ValueError(f"Invalid row number '{part}'. Rows must be > 1.")
                    selected_rows.add(num)
            
            if not selected_rows:
                return None # Empty range string
                
            return selected_rows
        except Exception as e:
            self.log_live(f"⚠️ Invalid row range '{range_str}': {e}. Processing all rows.")
            messagebox.showwarning("Range Error", f"Invalid row range: {e}\nProcessing all rows instead.")
            return None
        

        
    def open_import_popup(self):
        NewImportPopup(self, on_success_callback=self.open_mapping_popup)
    
    def open_mapping_popup(self, file_path=None, csv_data=None):
        """
        Handles the transition from file selection to column mapping.
        If data exists, prompts user to Merge or Replace.
        """
        
        def _launch_mapper_with_mode(merge_mode):
            """Helper to launch the MappingPopup."""
            MappingPopup(
                self, 
                on_import_callback=self.load_excel_data, 
                merge_mode=merge_mode, 
                file_path=file_path, 
                csv_data=csv_data
            )

        if self.excel_data:
            # Data exists, show Merge/Replace choice using the new themed popup
            MergeReplacePopup(self, _launch_mapper_with_mode)
            
        else:
            # No existing data, proceed directly to mapping in "replace" mode
            _launch_mapper_with_mode(merge_mode=False)


    def load_excel_data(self, file_path, phone_col, name_col, custom1_col, custom2_col, merge_mode, row_range_str=None, csv_data=None):
        """
        Loads data from file or CSV string, applies column and row mapping, 
        and updates the main Excel table.
        """
        try:
            prog = ProgressPopup(self, "Loading Data", total=1)
            prog.geometry(
                f"500x300+{self.winfo_rootx() + (self.winfo_width()-500)//2}"
                f"+{self.winfo_rooty() + (self.winfo_height()-300)//2}"
            )
            
            def perform_loading():
                new_data = []
                parsed_ranges = self.parse_row_ranges(row_range_str) # Parse ranges
                is_csv = (csv_data is not None) or (file_path and file_path.endswith('.csv'))

                # Helper to convert col letter to index (A=0, B=1, ...)
                def col_to_index(col_letter):
                    if not col_letter or not col_letter.isalpha() or len(col_letter) > 1:
                        return -1
                    return ord(col_letter.upper()) - 65

                # --- Column Indices ---
                phone_idx = -1
                name_idx = -1
                custom1_idx = -1
                custom2_idx = -1
                
                header = [] # To store header row

                # --- Data Loading Logic ---
                try:
                    if csv_data: # Handle in-memory CSV data
                        f = StringIO(csv_data)
                        try:
                            dialect = csv.Sniffer().sniff(f.read(2048))
                        except csv.Error:
                            f.seek(0)
                            dialect = 'excel' # Fallback
                        f.seek(0)
                        reader = csv.reader(f, dialect)
                        
                        header = next(reader) # Read header row
                        
                        # Find indices by name
                        try: phone_idx = header.index(phone_col)
                        except ValueError:
                            self.log_live(f"❌ Error: Phone column '{phone_col}' not found in CSV header.")
                            self.after(0, lambda: prog.close())
                            messagebox.showerror("Import Error", f"Phone column '{phone_col}' not found in CSV header: {header}")
                            return
                        try: name_idx = header.index(name_col)
                        except ValueError: name_idx = -1
                        try: custom1_idx = header.index(custom1_col)
                        except ValueError: custom1_idx = -1
                        try: custom2_idx = header.index(custom2_col)
                        except ValueError: custom2_idx = -1
                        
                        current_row_num = 2 # Start from 2 (row 1 was header)
                        for row in reader:
                            if parsed_ranges and current_row_num not in parsed_ranges:
                                current_row_num += 1
                                continue # Skip row if not in range
                            
                            entry = {}
                            try:
                                phone = row[phone_idx].strip() if len(row) > phone_idx else ""
                                if not phone:
                                    current_row_num += 1
                                    continue
                                entry["phone"] = phone
                                if name_idx != -1 and len(row) > name_idx:
                                    entry["name"] = row[name_idx].strip()
                                if custom1_idx != -1 and len(row) > custom1_idx:
                                    entry["custom1"] = row[custom1_idx].strip()
                                if custom2_idx != -1 and len(row) > custom2_idx:
                                    entry["custom2"] = row[custom2_idx].strip()
                                new_data.append(entry)
                            except Exception as e:
                                self.log_live(f"Error reading CSV row {current_row_num}: {e}")
                            current_row_num += 1
                    
                    elif file_path.endswith('.csv'): # Handle CSV file
                        with open(file_path, newline='', encoding='utf-8-sig') as csvfile:
                            try:
                                dialect = csv.Sniffer().sniff(csvfile.read(2048))
                            except csv.Error:
                                csvfile.seek(0)
                                dialect = 'excel' # Fallback
                            csvfile.seek(0)
                            reader = csv.reader(csvfile, dialect)
                            
                            header = next(reader) # Read header row
                            
                            # Find indices by name
                            try: phone_idx = header.index(phone_col)
                            except ValueError:
                                self.log_live(f"❌ Error: Phone column '{phone_col}' not found in CSV header.")
                                self.after(0, lambda: prog.close())
                                messagebox.showerror("Import Error", f"Phone column '{phone_col}' not found in CSV header: {header}")
                                return
                            try: name_idx = header.index(name_col)
                            except ValueError: name_idx = -1
                            try: custom1_idx = header.index(custom1_col)
                            except ValueError: custom1_idx = -1
                            try: custom2_idx = header.index(custom2_col)
                            except ValueError: custom2_idx = -1
                            
                            current_row_num = 2 # Start from 2
                            for row in reader:
                                if parsed_ranges and current_row_num not in parsed_ranges:
                                    current_row_num += 1
                                    continue # Skip row
                                
                                entry = {}
                                try:
                                    phone = row[phone_idx].strip() if len(row) > phone_idx else ""
                                    if not phone:
                                        current_row_num += 1
                                        continue
                                    entry["phone"] = phone
                                    if name_idx != -1 and len(row) > name_idx:
                                        entry["name"] = row[name_idx].strip()
                                    if custom1_idx != -1 and len(row) > custom1_idx:
                                        entry["custom1"] = row[custom1_idx].strip()
                                    if custom2_idx != -1 and len(row) > custom2_idx:
                                        entry["custom2"] = row[custom2_idx].strip()
                                    new_data.append(entry)
                                except Exception as e:
                                    self.log_live(f"Error reading CSV row {current_row_num}: {e}")
                                current_row_num += 1

                    elif file_path.endswith(('.xlsx', '.xls')): # Handle Excel file
                        wb = openpyxl.load_workbook(file_path, data_only=True)
                        sheet = wb.active
                        
                        # Get header
                        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
                        header = [str(cell) if cell is not None else "" for cell in header_row]
                        
                        # --- Determine indices ---
                        # Try by letter first
                        phone_idx = col_to_index(phone_col)
                        name_idx = col_to_index(name_col)
                        custom1_idx = col_to_index(custom1_col)
                        custom2_idx = col_to_index(custom2_col)
                        
                        # If not a valid letter, try by name
                        if phone_idx == -1:
                            try: phone_idx = header.index(phone_col)
                            except ValueError: phone_idx = -1 # Final check
                        if name_idx == -1 and name_col:
                            try: name_idx = header.index(name_col)
                            except ValueError: name_idx = -1
                        if custom1_idx == -1 and custom1_col:
                            try: custom1_idx = header.index(custom1_col)
                            except ValueError: custom1_idx = -1
                        if custom2_idx == -1 and custom2_col:
                            try: custom2_idx = header.index(custom2_col)
                            except ValueError: custom2_idx = -1
                        
                        if phone_idx == -1:
                            self.log_live(f"❌ Error: Phone column '{phone_col}' not found in Excel.")
                            self.after(0, lambda: prog.close())
                            messagebox.showerror("Import Error", f"Phone column '{phone_col}' not found in Excel header: {header}")
                            return
                        
                        current_row_num = 2 # Start from 2
                        for row in sheet.iter_rows(min_row=2, values_only=True):
                            if parsed_ranges and current_row_num not in parsed_ranges:
                                current_row_num += 1
                                continue # Skip row
                            
                            entry = {}
                            try:
                                phone_val = row[phone_idx] if len(row) > phone_idx else ""
                                phone = str(phone_val).strip() if phone_val is not None else ""
                                if not phone:
                                    current_row_num += 1
                                    continue
                                entry["phone"] = phone
                                if name_idx != -1 and len(row) > name_idx:
                                    name_val = row[name_idx]
                                    entry["name"] = str(name_val).strip() if name_val is not None else ""
                                if custom1_idx != -1 and len(row) > custom1_idx:
                                    c1_val = row[custom1_idx]
                                    entry["custom1"] = str(c1_val).strip() if c1_val is not None else ""
                                if custom2_idx != -1 and len(row) > custom2_idx:
                                    c2_val = row[custom2_idx]
                                    entry["custom2"] = str(c2_val).strip() if c2_val is not None else ""
                                new_data.append(entry)
                            except Exception as e:
                                self.log_live(f"Error reading Excel row {current_row_num}: {e}")
                            current_row_num += 1
                
                except Exception as e:
                     self.log_live(f"Error opening/reading file: {e}")
                     messagebox.showerror("File Error", f"Could not read file: {e}")
                     self.after(0, lambda: prog.close())
                     return

                # --- Handle Merge/Replace ---
                if merge_mode:
                    self.excel_data = self.excel_data + new_data
                    self.log_live(f"✅ Merged {len(new_data)} new contacts. Total: {len(self.excel_data)}")
                else:
                    self.excel_data = new_data
                    log_source = f"Google Sheet" if csv_data else f"{os.path.basename(file_path)}"
                    self.log_live(f"✅ Imported {len(new_data)} contacts from {log_source}")

                # --- Load data into the dynamic table ---
                self.after(0, lambda: self.excel_table.load_data(self.excel_data))
                self.after(0, lambda: prog.close())
                
            threading.Thread(target=perform_loading, daemon=True).start()
            
        except Exception as e:
            self.log_live(f"Error loading file: {e}")
            messagebox.showerror("Error", f"Failed to load file: {e}")



    def launch_whatsapp_beta(self):
        # --- New Logic to Clear Session ---
        try:
            if os.path.exists(SESSION_DIR):
                self.log_live("Clearing previous session for fresh login...")
                # Use shutil.rmtree to delete a non-empty directory
                shutil.rmtree(SESSION_DIR)
                self.log_live("Previous session cleared.")
            
            # Re-create the directory immediately
            if not os.path.exists(SESSION_DIR):
                 os.makedirs(SESSION_DIR)
                 
        except PermissionError:
            self.log_live(f"❌ Error: Session files are in use. Please close Chrome.")
            messagebox.showerror("Session Error", "Could not clear the session. Please close any open Chrome windows (especially WhatsApp Web) and try again.")
            return # Stop if we can't clear the session
        except Exception as e:
            self.log_live(f"❌ Error clearing session: {e}")
            messagebox.showerror("Error", f"An unexpected error occurred while clearing the session: {e}")
            return # Stop on other errors
        # --- End of New Logic ---
            
        def run_login():
            logger = GuiLogger(self)
            selenium_login(logger)
        
        threading.Thread(target=run_login, daemon=True).start()
        
    def toggle_sending(self):
        if not self.sending:
            self.start_sending()
        else:
            self.stop_sending()
            
    def start_sending(self):
        # --- NEW CHECKS ---
        if not self.is_logged_in():
            return # Stop if not logged in

        data = self.excel_table.get_data()
        if not data:
            messagebox.showerror("Error", "No phone numbers loaded. Please add at least one contact.")
            return # Stop if no data

        msg = self.message_text.get("0.0", "end-1c").strip()
        attachment_present = "Any" in self.attachments and self.attachments["Any"]
        if not msg:
            msg = "*"  # Default message if empty
        # --- FIX: Only block if NO content (neither message, nor attachment, nor custom image) is present ---
        if not msg and not attachment_present and not self.custom_image_enabled:
            messagebox.showerror("Error", "No content to send. Please add a message, attachment, or custom image.")
            self.stop_sending() # Reset button
            return # Stop if no content
        # --- END FIX ---

        self.sending = True
        self.start_stop_button.configure(text="Stop", fg_color="#FF4040")
        
        attachment_path = self.attachments["Any"] if "Any" in self.attachments else None
        
        # --- Get Delay Settings ---
        try:
            min_delay = int(self.min_delay_entry.get())
            if min_delay <= 0: min_delay = DEFAULT_MIN_DELAY # Ensure positive
        except ValueError:
            min_delay = DEFAULT_MIN_DELAY
            self.log_live(f"⚠️ Invalid Min Delay, using default: {DEFAULT_MIN_DELAY}s")
            
        try:
            max_delay = int(self.max_delay_entry.get())
            if max_delay <= 0: max_delay = DEFAULT_MAX_DELAY # Ensure positive
        except ValueError:
            max_delay = DEFAULT_MAX_DELAY
            self.log_live(f"⚠️ Invalid Max Delay, using default: {DEFAULT_MAX_DELAY}s")

        if min_delay > max_delay:
            self.log_live(f"⚠️ Min Delay ({min_delay}) > Max Delay ({max_delay}). Swapping values.")
            min_delay, max_delay = max_delay, min_delay # Swap them
        
        self.log_live(f"🚀 Using random delay range: {min_delay}s to {max_delay}s")
        # --- End Delay Settings ---

        # Filter out skipped entries
        valid_data = [entry for entry in data if not entry.get("skip", False)]
        if not valid_data:
             messagebox.showerror("Error", "All entries are marked to be skipped.")
             self.stop_sending()
             return

        numbers = [normalize_phone(entry["phone"]) for entry in valid_data if entry.get("phone")] # Ensure phone exists
        
        # --- *** UPDATED PERSONALIZATION *** ---
        personalized_msgs = []
        for entry in valid_data:
            if not entry.get("phone"): # Skip if no phone number for this entry
                continue
                
            temp_msg = msg.replace("{User_Name}", entry.get("name", ""))
            
            # Add logic to replace {Custom1} and {Custom2}
            if "{Custom1}" in temp_msg:
                temp_msg = temp_msg.replace("{Custom1}", entry.get("custom1", ""))
            if "{Custom2}" in temp_msg:
                temp_msg = temp_msg.replace("{Custom2}", entry.get("custom2", ""))
                
            personalized_msgs.append(temp_msg)
        # --- *** END OF UPDATE *** ---

        # Check if numbers and messages lists are aligned
        if len(numbers) != len(personalized_msgs):
            self.log_live(f"❌ Error: Mismatch between numbers ({len(numbers)}) and messages ({len(personalized_msgs)}).")
            messagebox.showerror("Error", "Mismatch in contact data. Check for entries without phone numbers.")
            self.stop_sending()
            return

        def schedule_and_send():
            # --- *** NEW SCHEDULER LOGIC V3 *** ---
            # 1. Atomically get the schedule time and clear it from the class
            # This prevents race conditions with old threads or new schedules
            target_time = self.schedule_time
            self.schedule_time = None 
            # --- *** END NEW V3 LOGIC *** ---

            print(f"DEBUG: Schedule thread started. Target time: {target_time}, Sending flag: {self.sending}")

            # --- *** NEW SCHEDULER LOGIC *** ---
            if target_time: # Use the local variable
                now = datetime.now()
                wait_seconds = (target_time - now).total_seconds()

                if wait_seconds > 0:
                    self.log_live(f"⏳ Waiting... Sending will start in {int(wait_seconds)} seconds at {target_time.strftime('%I:%M %p')}.")
                    print(f"DEBUG: Entering wait loop for {wait_seconds} seconds. Target: {target_time}")

                    # Sleep in 1-second intervals to keep the 'Stop' button responsive
                    for i in range(int(wait_seconds)):
                        if not self.sending: # Check if user clicked Stop
                            self.log_live("🛑 Schedule cancelled by user during wait.")
                            print("DEBUG: Sending flag became False during wait loop.")
                            self.after(0, self.stop_sending) # Reset UI
                            return # Exit the thread
                        
                        time.sleep(1)
                    
                    print(f"DEBUG: Wait loop finished. Proceeding to send.")
                    self.log_live("🚀 Scheduled time reached! Starting sending...")
                
                else:
                    # Time is already in the past, start immediately
                    self.log_live("🚀 Scheduled time is in the past! Starting sending now...")
                    print(f"DEBUG: Scheduled time {target_time} is in the past. Sending immediately.")

            else: # No schedule was set, start immediately
                 print("DEBUG: No schedule time set, starting immediately.")
                 self.log_live("🚀 Starting sending now...")
            # --- *** END OF NEW LOGIC *** ---


            # --- Rest of the function (preparing attachments, calling selenium_send_bulk) ---
            logger = GuiLogger(self)
            send_attachments = []

            # Custom Image Attachment Logic
            if self.custom_image_enabled:
                self.log_live("📎 Attaching custom images...")
                generated_image_paths = {}
                if hasattr(self, 'excel_data') and self.excel_data:
                    for item in self.excel_data:
                        phone_norm = normalize_phone(item.get('phone', ''))
                        if 'image_path' in item and item['image_path']:
                            generated_image_paths[phone_norm] = item['image_path']
                for num in numbers:
                    path = generated_image_paths.get(num)
                    if path and os.path.exists(path): send_attachments.append(path)
                    else: send_attachments.append(None)

            # Single Attachment Logic
            elif attachment_path:
                self.log_live(f"📎 Attaching single file: {os.path.basename(attachment_path)}")
                send_attachments = [attachment_path] * len(numbers)
            else:
                send_attachments = None

            # --- Start Selenium Send ---
            if not self.sending: # Final check
                print("DEBUG: Sending flag became False just before calling selenium_send_bulk.")
                self.after(0, self.stop_sending)
                return

            print(f"DEBUG: Calling selenium_send_bulk with {len(numbers)} numbers, min_delay={min_delay}, max_delay={max_delay}")

            # Make the call to the Selenium function
            success, failure = selenium_send_bulk(
                numbers,
                personalized_msgs,
                send_attachments,
                min_delay,
                max_delay,
                logger
            )

            # --- Report and Cleanup ---
            if (success + failure) > 0:
                generate_html_report(success, failure)
            else:
                self.log_live("ℹ️ No messages were attempted.")

            # Ensure stop_sending is called on the main thread after completion
            self.after(0, self.stop_sending)
            # self.schedule_time = None # Already cleared at the start
            print("DEBUG: Schedule thread finished.")

        # Start the thread
        threading.Thread(target=schedule_and_send, daemon=True).start()

        
    def stop_sending(self):
        self.sending = False
        self.start_stop_button.configure(text="Start", fg_color="green") # Changed to green
        self.log_live("🛑 Sending stopped.") # Changed log message
        
    def set_schedule_time(self, schedule_time):
        # Check if a process is already running
        if self.sending:
            self.log_live("❗️ Cannot set schedule while sending is already in progress.")
            messagebox.showwarning("In Progress", "Cannot set a new schedule while a send is in progress. Please stop the current process first.")
            return

        # Set the schedule time
        self.schedule_time = schedule_time
        self.log_live(f"⏰ Scheduled sending for {schedule_time.strftime('%I:%M %p, %b %d, %Y')}")
        
        # --- THIS IS THE FIX ---
        # Automatically call start_sending() to begin the "waiting" state.
        # The start_sending function will see self.schedule_time and
        # launch the thread that waits.
        self.start_sending()
        
    # New method to show the AI popup
    def show_ai_popup(self):
        AIPopup(self, self.ai_button)

    def start_ai_animation(self):
        """Starts the border color cycling animation."""
        if self.animation_active:
            return
        self.log_live("🤖 AI processing started...")
        self.animation_active = True
        self.ai_start_time = time.time()
        self.animate_glow_border() # Start the loop

    def stop_ai_animation(self):
        """Stops the animation and resets the border."""
        self.animation_active = False
        try:
            # Reset the FRAME'S BACKGROUND (which is our new border)
            mode = ctk.get_appearance_mode().lower()
            text_frame_bg = DARK_FG if mode == "dark" else LIGHT_FG
            self.text_area_frame.configure(
                fg_color=text_frame_bg
            )
        except Exception as e:
            print(f"Error stopping animation: {e}") # Handle if widget is destroyed

    def animate_glow_border(self):
        """The animation loop. Cycles border color."""
        if not self.animation_active:
            self.stop_ai_animation() # Ensure it resets
            return

        # Cycle through colors based on time for "rotation" effect
        color_index = int((time.time() * 5) % len(self.gradient_colors)) # Slower pulse
        color = self.gradient_colors[color_index]
        
        try:
            self.text_area_frame.configure(fg_color=color) # <-- CHANGED: Animate fg_color
        except Exception:
            pass # Ignore errors if widget is destroyed
        
        # Loop at ~30 FPS
        self.after(33, self.animate_glow_border)

    def type_text_animation(self, text, index=0):
        """Animates text insertion into the message box."""
        try:
            if index < len(text):
                self.message_text.insert("end", text[index])
                self.message_text.see("end") # Auto-scroll
                
                # Randomize typing speed slightly (10ms to 40ms)
                delay = random.randint(10, 40) 
                self.after(delay, self.type_text_animation, text, index + 1)
            else:
                self.save_state() # Save state after typing is complete
        except Exception as e:
            print(f"Typing animation error: {e}") # Catch errors if window is closed

    # Modified method to handle popup closing
    def process_ai(self, option, popup=None):
        if popup:
            popup.destroy() # Close the popup when an option is clicked
            
        text = self.message_text.get("0.0", "end-1c").strip()
        if not text and option != "Ask AI": # Allow "Ask AI" to run on empty text
            messagebox.showerror("Error", "No message text to process.")
            return
            
        if option == "Translate":
            TranslatePopup(self, self.translate_message)
        else:
            threading.Thread(target=self.call_gemini_api, args=(option, text), daemon=True).start()
            
    def translate_message(self, lang):
        text = self.message_text.get("0.0", "end-1c").strip()
        if not text:
            messagebox.showerror("Error", "No message text to translate.")
            return
        threading.Thread(target=self.call_gemini_api, args=("Translate", text, lang), daemon=True).start()
        
    # Add this import at the top of the file with other imports


    # In the __init__ method of WabulkXpressApp, add:
    genai.configure(api_key="AIzaSyBR9Os8LKyNQAwkeLKXO1OyGQnRjnaMe8w")

    # Ensure this method is properly indented inside the WabulkXpressApp class
    # (4 spaces or consistent with class indentation)
    def call_gemini_api(self, option, text, lang=None):
        # --- 1. Start Animation on main thread ---
        self.after(0, self.start_ai_animation)
        
        def ai_thread():
            processed_text = ""
            error_message = ""
            
            try:
                # This code runs in a separate thread
                prompt = self.ai_prompts[option]
                if option == "Translate":
                    prompt = prompt.format(lang=lang)
                
                full_prompt = f"{prompt}\n\n{text}"
                
                model = genai.GenerativeModel('gemini-2.0-flash')
                response = model.generate_content(full_prompt)
                
                processed_text = response.text
                
            except Exception as e:
                error_message = f"❌ Error in AI processing: {e}"
                logger.error(f"Unexpected AI error: {e}", exc_info=True)

            # --- 2. Enforce 3-second minimum ---
            elapsed = time.time() - self.ai_start_time
            min_wait = 3.0
            if elapsed < min_wait:
                sleep_time = min_wait - elapsed
                print(f"DEBUG: AI finished in {elapsed:.2f}s, waiting {sleep_time:.2f}s for animation.")
                time.sleep(sleep_time) # This blocks the AI thread, which is fine
            
            # --- 3. Stop animation and show results on main thread ---
            def update_ui():
                self.stop_ai_animation()
                
                if processed_text:
                    self.save_state() # Save before changing
                    self.message_text.delete("0.0", "end")
                    self.type_text_animation(processed_text) # Call new typing function
                    self.log_live(f"✅ AI {option} applied successfully!")
                elif error_message:
                    self.log_live(error_message)
                    messagebox.showerror("AI Error", f"Failed to process with AI: {error_message}")
                else:
                    self.log_live(f"⚠️ AI processing returned empty result for {option}.")
            
            # Schedule the UI update on the main thread
            self.after(0, update_ui)

        # Start the AI processing in a new thread
        threading.Thread(target=ai_thread, daemon=True).start()



    def load_profile_picture(self):
        try:
            # 1. Fetch GitHub API
            api_response = requests.get(GITHUB_AVATAR_API, timeout=10)
            api_response.raise_for_status()
            avatar_url = api_response.json()['avatar_url']
            
            # 2. Download Avatar Image
            image_data = requests.get(avatar_url, timeout=10).content
            avatar = Image.open(BytesIO(image_data)).convert("RGBA")
            
            # 3. Create the "Gemini Ring" Image
            size = (60, 60) # Avatar size
            ring_thickness = 4
            final_size_val = size[0] + ring_thickness * 2
            final_size = (final_size_val, final_size_val) # Final image size

            base = Image.new("RGBA", final_size, (0, 0, 0, 0)) # Transparent base
            draw = ImageDraw.Draw(base)
            
            # R, Y, G, B
            colors = ["#EA4335", "#FBBC05", "#34A853", "#4285F4"] 
            
            # 4. Draw the 4 colored arcs
            bbox = [(0, 0), final_size]
            draw.arc(bbox, 270, 360, fill=colors[0], width=ring_thickness) # Top-Right, Red
            draw.arc(bbox, 180, 270, fill=colors[1], width=ring_thickness) # Top-Left, Yellow
            draw.arc(bbox, 90, 180, fill=colors[2], width=ring_thickness) # Bottom-Left, Green
            draw.arc(bbox, 0, 90, fill=colors[3], width=ring_thickness) # Bottom-Right, Blue

            # 5. Create circular mask for avatar
            mask = Image.new("L", size, 0)
            draw_mask = ImageDraw.Draw(mask)
            draw_mask.ellipse((0, 0), size, fill=255)
            
            avatar = avatar.resize(size, Image.Resampling.LANCZOS)
            
            # 6. Paste avatar inside the ring
            base.paste(avatar, (ring_thickness, ring_thickness), mask)
            
            # 7. Create the CTkImage
            self.profile_image = ctk.CTkImage(light_image=base, dark_image=base, size=final_size)
            
            # 8. Schedule the update on the main GUI thread
            if self.profile_image_label:
                self.after(0, self.update_profile_image)
                
        except Exception as e:
            print(f"Failed to load profile pic: {e}")
            logger.error(f"Failed to load GitHub profile picture: {e}")

    def update_profile_image(self):
        # This runs on the main thread to safely update the UI
        if self.profile_image and self.profile_image_label:
            self.profile_image_label.configure(image=self.profile_image, text="")


            
    def update_message_text(self, text):
        """Clears and inserts text using the typing animation."""
        self.save_state() # Save before changing
        self.message_text.delete("0.0", "end")
        self.type_text_animation(text) # Use the new animation
        
    def log_live(self, message):
        if not self.winfo_exists():
            return
        # Simple check to prevent excessive logging of the same message rapidly
        try:
             last_log = self.live_alerts.get("end-2l linestart", "end-1l lineend") # Get second to last line
             if message == last_log:
                  return 
        except tk.TclError:
             pass # Ignore error if textbox is empty or has only one line
             
        self.live_alerts.configure(state="normal")
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.live_alerts.insert("end", f"[{timestamp}] {message}\n")
        self.live_alerts.see("end")
        self.live_alerts.configure(state="disabled")
        
    def check_for_update(self):
        """Launches the new custom update popup window."""
        UpdatePopup(self)

    def apply_theme(self, new_mode):
        """Applies the specified theme (Light or Dark) to all widgets."""
        
        # --- This block updates the root window background ---
        if new_mode == "Light":
            self.master.configure(bg=LIGHT_BG)
        else: # new_mode == "Dark"
            self.master.configure(bg=DARK_BG)
        
        # --- Update Custom Title Bar Colors ---
        self._set_title_bar_colors(new_mode.lower())
        if hasattr(self, 'title_bar') and self.title_bar.winfo_exists():
            self.title_bar.configure(fg_color=self.title_bar_color)
            if hasattr(self, 'title_label') and self.title_label.winfo_exists():
                self.title_label.configure(text_color=self.title_text_color)
            if hasattr(self, 'minimize_button') and self.minimize_button.winfo_exists():
                self.minimize_button.configure(hover_color=self.button_hover_color, text_color=self.title_text_color)
            if hasattr(self, 'maximize_button') and self.maximize_button.winfo_exists():
                self.maximize_button.configure(hover_color=self.button_hover_color, text_color=self.title_text_color)
            if hasattr(self, 'close_button') and self.close_button.winfo_exists():
                # --- THIS IS THE FIX FOR THE RED HOVER ---
                self.close_button.configure(hover_color=self.close_hover_color) 
                self.close_button.configure(text_color=self.title_text_color)
        # --- End Title Bar Update ---

        if hasattr(self, "profile_button"):
            self.profile_button.update_theme()

        # --- Update Main App Widgets based on the new theme ---
        if new_mode == "Light":
            self.configure(fg_color=LIGHT_BG)
            self.sidebar.configure(fg_color=LIGHT_FG)
            self.header.configure(fg_color=LIGHT_FG)
            self.main_area.configure(fg_color="transparent")
            self.message_frame.configure(fg_color=LIGHT_FG)
            self.excel_frame.configure(fg_color=LIGHT_FG)
            self.live_alerts.configure(fg_color=LIGHT_INNER)
            if not self.animation_active:
                self.text_area_frame.configure(fg_color=LIGHT_FG)
            self.message_text.configure(fg_color=LIGHT_INNER)
            self.min_delay_entry.configure(fg_color=LIGHT_INNER)
            self.max_delay_entry.configure(fg_color=LIGHT_INNER)
            if hasattr(self, 'excel_table'):
                self.excel_table.configure(fg_color=LIGHT_FG, scrollbar_button_color="#AAAAAA", scrollbar_button_hover_color="#888888")
                for row in self.excel_table.rows:
                    if "phone" in row and row["phone"].winfo_exists(): row["phone"].configure(fg_color=LIGHT_INNER)
                    if "name" in row and row["name"].winfo_exists(): row["name"].configure(fg_color=LIGHT_INNER)
                    if "custom1" in row and row["custom1"].winfo_exists(): row["custom1"].configure(fg_color=LIGHT_INNER)
                    if "custom2" in row and row["custom2"].winfo_exists(): row["custom2"].configure(fg_color=LIGHT_INNER)
                    if "indicator" in row and row["indicator"].winfo_exists(): row["indicator"].configure(bg=LIGHT_BG)
            if hasattr(self, 'ai_button'):
                ai_bg, ai_hover, ai_press = LIGHT_FG, "#E0E0E0", "#CFCFCF"
                self.ai_button.configure(fg_color=ai_bg, hover_color=ai_hover)
                self.ai_button.bind("<ButtonPress-1>", lambda e, c=ai_press: self.ai_button.configure(fg_color=c))
                self.ai_button.bind("<ButtonRelease-1>", lambda e, c=ai_hover: self.ai_button.configure(fg_color=c))
        else: # new_mode == "Dark"
            self.configure(fg_color=DARK_BG)
            self.sidebar.configure(fg_color=DARK_FG)
            self.header.configure(fg_color=DARK_FG)
            self.main_area.configure(fg_color="transparent")
            self.message_frame.configure(fg_color=DARK_FG)
            self.excel_frame.configure(fg_color=DARK_FG)
            self.live_alerts.configure(fg_color=DARK_INNER)
            if not self.animation_active:
                self.text_area_frame.configure(fg_color=DARK_FG)
            self.message_text.configure(fg_color=DARK_INNER)
            self.min_delay_entry.configure(fg_color=DARK_INNER)
            self.max_delay_entry.configure(fg_color=DARK_INNER)
            if hasattr(self, 'excel_table'):
                self.excel_table.configure(fg_color=DARK_FG, scrollbar_button_color="#555555", scrollbar_button_hover_color="#333333")
                for row in self.excel_table.rows:
                    if "phone" in row and row["phone"].winfo_exists(): row["phone"].configure(fg_color=DARK_INNER)
                    if "name" in row and row["name"].winfo_exists(): row["name"].configure(fg_color=DARK_INNER)
                    if "custom1" in row and row["custom1"].winfo_exists(): row["custom1"].configure(fg_color=DARK_INNER)
                    if "custom2" in row and row["custom2"].winfo_exists(): row["custom2"].configure(fg_color=DARK_INNER)
                    if "indicator" in row and row["indicator"].winfo_exists(): row["indicator"].configure(bg=DARK_BG)
            if hasattr(self, 'ai_button'):
                ai_bg, ai_hover, ai_press = DARK_FG, "#333333", "#555555"
                self.ai_button.configure(fg_color=ai_bg, hover_color=ai_hover)
                self.ai_button.bind("<ButtonPress-1>", lambda e, c=ai_press: self.ai_button.configure(fg_color=c))
                self.ai_button.bind("<ButtonRelease-1>", lambda e, c=ai_hover: self.ai_button.configure(fg_color=c))

        self.refresh_icons()
        if hasattr(self, 'all_hints'):
            for hint in self.all_hints:
                if hint.winfo_exists():
                    hint.update_theme()

    def network_check_and_wait(self, current_index, total_count):
        """
        Runs network checks with 5-minute wait loops, timing out after 1 hour.
        
        Returns:
            bool: True if network restored within 1 hour, False otherwise.
        """
        # Ensure 'self' is passed for logging and 'sending' flag access
        gui_logger = GuiLogger(self)
        
        TIMEOUT_SECONDS = 3600  # 1 hour
        WAIT_INTERVAL_SECONDS = 300 # 5 minutes
        start_time = time.time()
        
        while time.time() - start_time < TIMEOUT_SECONDS:
            if not self.sending:
                gui_logger.log("🛑 Stop signal received. Halting network wait...")
                return False

            if network_check(gui_logger):
                gui_logger.log("🌐 Network Restored! Resuming send process...")
                return True
            else:
                remaining_timeout = int(TIMEOUT_SECONDS - (time.time() - start_time))
                gui_logger.log(
                    f"⚠️ Network Down! Message was sent up to contact {current_index}/{total_count}. "
                    f"Waiting 5 minutes (Timeout in {remaining_timeout}s)..."
                )
                
                # Sleep in 1-second intervals to check for the 'Stop' button
                sleep_start = time.time()
                while time.time() - sleep_start < WAIT_INTERVAL_SECONDS:
                    if not self.sending:
                        gui_logger.log("🛑 Stop signal received. Halting network wait...")
                        return False
                    time.sleep(1) # Check stop flag every second
        
        # 1-hour timeout reached
        gui_logger.log(
            f"❌ Network Timeout (1 Hour). Stopping bulk send process. "
            f"Last successfully processed contact index was {current_index-1}."
        )
        self.after(0, self.stop_sending) # Schedule stop on main thread
        return False
    

    def toggle_theme(self):
        current_mode = ctk.get_appearance_mode().lower()
        new_mode = "Light" if current_mode == "dark" else "Dark"
        ctk.set_appearance_mode(new_mode) # Set the global theme
        
        # Call the new helper function
        self.apply_theme(new_mode)

    def on_close(self):
        # Ensure cleanup happens before destroying the window
        if self.sending:
             # Optionally ask the user if they want to stop sending
             if messagebox.askyesno("Confirm Exit", "Sending is in progress. Are you sure you want to exit?"):
                  self.stop_sending() # Attempt graceful stop if possible
             else:
                  return # Prevent closing if user cancels

        # Stop video if playing
        if hasattr(self, "video_player") and self.video_player:
            try:
                self.video_player.stop()
            except Exception as e:
                print(f"Error stopping video player: {e}")
        
        # Add any other necessary cleanup here (e.g., close Selenium driver if running)
        
    # ... (cleanup logic)
        self.master.destroy() # Destroy the main window
        sys.exit(0) # Ensure the script fully terminates

def _apply_window_fixes(root_window):
    """Applies all necessary Windows API fixes for the borderless window."""
    try:
        # --- 1. Get Window Handle (HWND) by Title ---
        window_title = ctypes.c_wchar_p("WabulkXpress")
        hwnd = ctypes.windll.user32.FindWindowW(None, window_title)
        
        if not hwnd:
            print("Error: Could not find window handle by title 'WabulkXpress'. Trying fallback.")
            root_window.update_idletasks()
            hwnd = ctypes.windll.user32.GetParent(root_window.winfo_id())

        if not hwnd:
            print("Error: Could not get window handle.")
            return

        # --- 2. Fix Taskbar Icon and Alt-Tab ---
        GWL_EXSTYLE = -20
        WS_EX_APPWINDOW = 0x00040000
        WS_EX_TOOLWINDOW = 0x00000080
        style = ctypes.windll.user32.GetWindowLongW(hwnd, GWL_EXSTYLE)
        style = (style | WS_EX_APPWINDOW) & ~WS_EX_TOOLWINDOW
        ctypes.windll.user32.SetWindowLongW(hwnd, GWL_EXSTYLE, style)

        # --- 3. Apply Rounded Corners (Win 11+) ---
        if sys.platform == "win32":
            corner_preference = ctypes.c_int(2) # DWMWCP_ROUND
            ctypes.windll.dwmapi.DwmSetWindowAttribute(
                hwnd, 
                33, # DWMWA_WINDOW_CORNER_PREFERENCE
                ctypes.byref(corner_preference), 
                ctypes.sizeof(corner_preference)
            )

    except Exception as e:
        print(f"Error applying window fixes: {e}")

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    root = tkinterdnd2.TkinterDnD.Tk()
    print(f"DEBUG: 'root' variable created: {root}")
    root.overrideredirect(True)
    root.withdraw()

    ctk.set_appearance_mode("System")
    detected_mode = ctk.get_appearance_mode().lower()
    
    root_bg_color = DARK_BG if detected_mode == "dark" else LIGHT_BG
    root.configure(bg=root_bg_color)
    root.title("WabulkXpress") 
    root.geometry("1800x900")

    if os.path.exists(TITLE_ICON_PATH):
        try:
            icon_img = Image.open(TITLE_ICON_PATH).resize((32,32), Image.Resampling.LANCZOS)
            icon_tk = ImageTk.PhotoImage(icon_img)
            root.wm_iconphoto(False, icon_tk)
        except Exception as e:
            print(f"Error loading title icon: {e}")

    loco_icon = os.path.join(os.getcwd(), "bin", "loco.ico")
    if os.path.exists(loco_icon):
        try:
            root.iconbitmap(loco_icon)
        except Exception as e:
            print(f"Error loading bitmap icon: {e}")

    logger.info("Starting WabulkXpress Application...")

    # Create the App Frame inside the Root
    app = WabulkXpressApp(master=root, fg_color="transparent")
    # Pack the main app frame *below* the custom title bar
    app.pack(fill="both", expand=True)

    # Set the close protocol to the app's method
    root.protocol("WM_DELETE_WINDOW", app.on_close)
    _apply_window_fixes(root)
    root.deiconify()
    try:
        root.mainloop() # Run the root window's mainloop
    except Exception as e:
         logger.error("An unhandled exception occurred in the main loop:", exc_info=True)
         messagebox.showerror("Fatal Error", f"A critical error occurred:\n{e}\n\nPlease check the logs for details.")
    finally:
         logger.info("WabulkXpress Application finished.")
